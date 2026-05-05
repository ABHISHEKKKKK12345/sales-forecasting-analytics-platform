"""
Sales Forecasting & Automated Reporting System  v2.0
═══════════════════════════════════════════════════════════════════════════════
Models    : ARIMA (MLE grid) · Holt-Winters (add/mul) · Random Forest
            Gradient Boosting · Ridge Regression · XGBoost · LightGBM
            Prophet · Stacked Ensemble
Reports   : Multi-page PDF · 9-sheet Excel workbook · JSON summary
Config    : TOML file — settings persist between runs
            (~/.config/sales_forecast/config.toml or portable ./sf_config.toml)
Author    : Abhishek
Version   : 2.0.0
═══════════════════════════════════════════════════════════════════════════════
"""

"""
Sales Forecasting System — Development Version

Base System       : forecasting_experimental.py
Development Type  : Pre-production
Internal Version  : v2

Description:
This file is a development iteration of the experimental forecasting model.
It is used for testing and enhancements and is not intended for production use.
"""

# ─── Standard library ─────────────────────────────────────────────────────────
import io
import json
import logging
import logging.handlers
import math
import os
import queue
import shutil
import sys
import tempfile
import threading
import time
import traceback
import warnings
from contextlib import contextmanager
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

warnings.filterwarnings("ignore")

# ── Python 3.11+ has tomllib in stdlib; fall back to tomli on older Pythons ──
try:
    import tomllib as _tomllib_read
except ImportError:
    try:
        import tomli as _tomllib_read  # type: ignore[no-redef]
    except ImportError:
        _tomllib_read = None  # type: ignore[assignment]

try:
    import tomli_w as _tomllib_write  # type: ignore
except ImportError:
    _tomllib_write = None  # type: ignore[assignment]


# ── Logging: stream + rotating file ─────────────────────────────────────────
_LOG_DIR = Path.home() / "SalesForecastReports"
_LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.handlers.RotatingFileHandler(
            _LOG_DIR / "sales_forecast.log",
            maxBytes=5 * 1024 * 1024,
            backupCount=3,
            encoding="utf-8",
        ),
    ],
)
logger = logging.getLogger("SalesForecast")


def _global_exception_handler(exc_type, exc_value, exc_tb):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_tb)
        return
    logger.critical("Unhandled exception", exc_info=(exc_type, exc_value, exc_tb))


sys.excepthook = _global_exception_handler

# ─── GUI ──────────────────────────────────────────────────────────────────────
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    from tkinter.font import Font as TkFont
    _TK_AVAILABLE = True
except ImportError:
    _TK_AVAILABLE = False
    print(
        "ERROR: tkinter is not available.\n"
        "  Ubuntu/Debian : sudo apt install python3-tk\n"
        "  macOS         : brew install python-tk\n"
        "  Windows       : Reinstall Python with 'tcl/tk and IDLE' checked."
    )
    sys.exit(1)

# ─── Third-party ──────────────────────────────────────────────────────────────
_IMPORT_ERROR: Optional[str] = None
_XGBOOST_AVAILABLE   = False
_LIGHTGBM_AVAILABLE  = False
_PROPHET_AVAILABLE   = False

try:
    import numpy as np
    import pandas as pd
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import seaborn as sns
    from scipy import stats
    from scipy.optimize import minimize
    from scipy.signal import periodogram
    from sklearn.ensemble import GradientBoostingRegressor, RandomForestRegressor
    from sklearn.linear_model import Ridge
    from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
    from sklearn.model_selection import TimeSeriesSplit
    from sklearn.preprocessing import StandardScaler
    from sklearn.base import clone as sklearn_clone
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        HRFlowable, Image as RLImage, KeepTogether,
        PageBreak, Paragraph, SimpleDocTemplate,
        Spacer, Table, TableStyle,
    )
except ImportError as _e:
    _IMPORT_ERROR = str(_e)

# Optional heavy models — degrade gracefully if not installed
if not _IMPORT_ERROR:
    try:
        import xgboost as xgb
        _XGBOOST_AVAILABLE = True
    except ImportError:
        pass

    try:
        import lightgbm as lgb
        _LIGHTGBM_AVAILABLE = True
    except ImportError:
        pass

    try:
        from prophet import Prophet
        _PROPHET_AVAILABLE = True
    except ImportError:
        pass


# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS & THEME
# ══════════════════════════════════════════════════════════════════════════════

APP_TITLE   = "Sales Forecasting System"
APP_VERSION = "2.0.0"

_CONFIG_DIR  = Path.home() / ".config" / "sales_forecast"
_CONFIG_FILE = _CONFIG_DIR / "config.toml"
_LOCAL_CONFIG = Path("sf_config.toml")


class _NpEncoder(json.JSONEncoder):
    def default(self, obj: Any) -> Any:
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        if isinstance(obj, Path):
            return str(obj)
        return super().default(obj)


def _fmt_month(value: Any) -> str:
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%b %Y")
    if isinstance(value, str):
        try:
            return pd.Timestamp(value).strftime("%b %Y")
        except Exception:
            return value
    return str(value)


_PANDAS_GTE_2: bool = False
if not _IMPORT_ERROR:
    try:
        _PANDAS_GTE_2 = int(pd.__version__.split(".")[0]) >= 2
    except Exception:
        pass


class Theme:
    BG        = "#0D1117"
    SURFACE   = "#161B22"
    SURFACE2  = "#1C2128"
    BORDER    = "#30363D"
    BORDER2   = "#21262D"
    ACCENT    = "#238636"
    ACCENT_HV = "#2EA043"
    BLUE      = "#1F6FEB"
    BLUE_HV   = "#388BFD"
    WARN      = "#E3B341"
    DANGER    = "#DA3633"
    SUCCESS   = "#3FB950"
    TEXT      = "#E6EDF3"
    TEXT2     = "#C9D1D9"
    MUTED     = "#8B949E"
    MUTED2    = "#6E7681"
    PURPLE    = "#8B5CF6"
    TEAL      = "#00A0C7"
    CHART     = ["#1F6FEB", "#00A0C7", "#E8730A", "#238636",
                 "#DA3633", "#8B5CF6", "#E3B341", "#00796B"]


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION — with TOML persistence
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ForecastConfig:
    output_path:            Path  = field(default_factory=lambda: Path.home() / "SalesForecastReports" / "report")
    charts_dir:             Path  = field(default_factory=lambda: Path(tempfile.gettempdir()) / "sf_charts_tmp")
    forecast_horizon:       int   = 12
    confidence_level:       float = 0.95
    seasonality_period:     int   = 12
    use_arima:              bool  = True
    use_random_forest:      bool  = True
    use_gradient_boosting:  bool  = True
    use_ridge:              bool  = True
    use_exp_smoothing:      bool  = True
    use_xgboost:            bool  = True
    use_lightgbm:           bool  = True
    use_prophet:            bool  = True
    company_name:           str   = "My Company"
    report_title:           str   = "Sales Forecasting & Demand Analysis Report"
    analyst_name:           str   = "Automated Analytics Engine"
    currency_symbol:        str   = "$"
    report_format:          List[str] = field(default_factory=lambda: ["pdf", "excel"])
    primary_color:          Tuple = (0.12, 0.29, 0.49)
    accent_color:           Tuple = (0.0,  0.63, 0.79)
    n_bootstrap:            int   = 300
    cv_folds:               int   = 3
    prophet_changepoint_scale: float = 0.05
    prophet_seasonality_mode:  str   = "additive"

    def ensure_dirs(self) -> None:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.charts_dir = Path(tempfile.mkdtemp(prefix="sf_charts_"))

    def validate(self) -> List[str]:
        errs: List[str] = []
        if self.forecast_horizon < 1 or self.forecast_horizon > 60:
            errs.append("Forecast horizon must be 1–60 months.")
        if not (0 < self.confidence_level < 1):
            errs.append("Confidence level must be between 0 and 1.")
        if not self.report_format:
            errs.append("At least one report format must be selected.")
        model_flags = [
            self.use_arima, self.use_random_forest, self.use_gradient_boosting,
            self.use_ridge, self.use_exp_smoothing, self.use_xgboost,
            self.use_lightgbm, self.use_prophet,
        ]
        if not any(model_flags):
            errs.append("At least one forecasting model must be enabled.")
        if self.n_bootstrap < 50:
            errs.append("n_bootstrap must be >= 50.")
        if self.n_bootstrap > 5_000:
            errs.append("n_bootstrap must be <= 5,000.")
        if self.cv_folds < 2:
            errs.append("cv_folds must be >= 2.")
        if self.seasonality_period != 0 and not (2 <= self.seasonality_period <= 60):
            errs.append("seasonality_period must be 0 (auto-detect) or 2–60.")
        if len(self.company_name.strip()) == 0:
            errs.append("Company name must not be empty.")
        if len(self.company_name) > 120:
            errs.append("Company name must be <= 120 characters.")
        if len(self.analyst_name) > 120:
            errs.append("Analyst name must be <= 120 characters.")
        if len(self.currency_symbol.strip()) == 0:
            errs.append("Currency symbol must not be empty.")
        if self.prophet_seasonality_mode not in ("additive", "multiplicative"):
            errs.append("prophet_seasonality_mode must be 'additive' or 'multiplicative'.")
        return errs

    def to_toml_dict(self) -> Dict[str, Any]:
        return {
            "forecast": {
                "horizon":            self.forecast_horizon,
                "confidence_level":   self.confidence_level,
                "seasonality_period": self.seasonality_period,
                "n_bootstrap":        self.n_bootstrap,
                "cv_folds":           self.cv_folds,
            },
            "models": {
                "arima":            self.use_arima,
                "exp_smoothing":    self.use_exp_smoothing,
                "random_forest":    self.use_random_forest,
                "gradient_boosting":self.use_gradient_boosting,
                "ridge":            self.use_ridge,
                "xgboost":          self.use_xgboost,
                "lightgbm":         self.use_lightgbm,
                "prophet":          self.use_prophet,
            },
            "prophet": {
                "changepoint_prior_scale": self.prophet_changepoint_scale,
                "seasonality_mode":        self.prophet_seasonality_mode,
            },
            "report": {
                "company_name":    self.company_name,
                "report_title":    self.report_title,
                "analyst_name":    self.analyst_name,
                "currency_symbol": self.currency_symbol,
                "formats":         self.report_format,
                "output_path":     str(self.output_path),
            },
        }

    @classmethod
    def from_toml_dict(cls, data: Dict[str, Any]) -> "ForecastConfig":
        cfg = cls()
        fc  = data.get("forecast", {})
        mc  = data.get("models",   {})
        pc  = data.get("prophet",  {})
        rc  = data.get("report",   {})
        cfg.forecast_horizon          = int(fc.get("horizon",            cfg.forecast_horizon))
        cfg.confidence_level          = float(fc.get("confidence_level", cfg.confidence_level))
        cfg.seasonality_period        = int(fc.get("seasonality_period", cfg.seasonality_period))
        cfg.n_bootstrap               = int(fc.get("n_bootstrap",        cfg.n_bootstrap))
        cfg.cv_folds                  = int(fc.get("cv_folds",           cfg.cv_folds))
        cfg.use_arima                 = bool(mc.get("arima",             cfg.use_arima))
        cfg.use_exp_smoothing         = bool(mc.get("exp_smoothing",     cfg.use_exp_smoothing))
        cfg.use_random_forest         = bool(mc.get("random_forest",     cfg.use_random_forest))
        cfg.use_gradient_boosting     = bool(mc.get("gradient_boosting", cfg.use_gradient_boosting))
        cfg.use_ridge                 = bool(mc.get("ridge",             cfg.use_ridge))
        cfg.use_xgboost               = bool(mc.get("xgboost",          cfg.use_xgboost))
        cfg.use_lightgbm              = bool(mc.get("lightgbm",         cfg.use_lightgbm))
        cfg.use_prophet               = bool(mc.get("prophet",          cfg.use_prophet))
        cfg.prophet_changepoint_scale = float(pc.get("changepoint_prior_scale", cfg.prophet_changepoint_scale))
        cfg.prophet_seasonality_mode  = str(pc.get("seasonality_mode",          cfg.prophet_seasonality_mode))
        cfg.company_name              = str(rc.get("company_name",    cfg.company_name))
        cfg.report_title              = str(rc.get("report_title",    cfg.report_title))
        cfg.analyst_name              = str(rc.get("analyst_name",    cfg.analyst_name))
        cfg.currency_symbol           = str(rc.get("currency_symbol", cfg.currency_symbol))
        cfg.report_format             = list(rc.get("formats",        cfg.report_format))
        op = rc.get("output_path")
        if op:
            cfg.output_path = Path(op)
        return cfg

    def save(self, path: Optional[Path] = None) -> bool:
        target = path or (_LOCAL_CONFIG if _LOCAL_CONFIG.parent.is_dir() else _CONFIG_FILE)
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            d = self.to_toml_dict()
            if _tomllib_write is not None:
                with open(target, "wb") as f:
                    _tomllib_write.dump(d, f)
            else:
                lines = _dict_to_toml_lines(d)
                target.write_text("\n".join(lines) + "\n", encoding="utf-8")
            logger.info("Config saved to %s", target)
            return True
        except Exception as exc:
            logger.warning("Could not save config: %s", exc)
            return False

    @classmethod
    def load(cls, path: Optional[Path] = None) -> "ForecastConfig":
        candidates = [path] if path else [_LOCAL_CONFIG, _CONFIG_FILE]
        for p in candidates:
            if p and p.exists():
                try:
                    if _tomllib_read is not None:
                        with open(p, "rb") as f:
                            data = _tomllib_read.load(f)
                    else:
                        data = _parse_simple_toml(p.read_text(encoding="utf-8"))
                    cfg = cls.from_toml_dict(data)
                    logger.info("Config loaded from %s", p)
                    return cfg
                except Exception as exc:
                    logger.warning("Could not load config from %s: %s", p, exc)
        return cls()


def _dict_to_toml_lines(d: Dict, prefix: str = "") -> List[str]:
    lines: List[str] = []
    scalars: List[Tuple[str, Any]] = []
    tables:  List[Tuple[str, Dict]] = []
    for k, v in d.items():
        if isinstance(v, dict):
            tables.append((k, v))
        else:
            scalars.append((k, v))
    for k, v in scalars:
        if isinstance(v, bool):
            lines.append(f"{k} = {'true' if v else 'false'}")
        elif isinstance(v, str):
            lines.append(f'{k} = "{v}"')
        elif isinstance(v, list):
            items = ", ".join(f'"{x}"' if isinstance(x, str) else str(x) for x in v)
            lines.append(f"{k} = [{items}]")
        else:
            lines.append(f"{k} = {v}")
    for k, v in tables:
        section = f"{prefix}{k}" if prefix else k
        lines.append(f"\n[{section}]")
        lines.extend(_dict_to_toml_lines(v, f"{section}."))
    return lines


def _parse_simple_toml(text: str) -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    current: Dict[str, Any] = result
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("[") and line.endswith("]"):
            section_path = line[1:-1].strip().split(".")
            current = result
            for part in section_path:
                current = current.setdefault(part, {})
            continue
        if "=" in line:
            key, _, val_str = line.partition("=")
            key = key.strip()
            val_str = val_str.strip()
            if val_str.lower() == "true":
                current[key] = True
            elif val_str.lower() == "false":
                current[key] = False
            elif val_str.startswith('"') and val_str.endswith('"'):
                current[key] = val_str[1:-1]
            elif val_str.startswith("[") and val_str.endswith("]"):
                inner = val_str[1:-1]
                parts = [x.strip().strip('"') for x in inner.split(",") if x.strip()]
                current[key] = parts
            else:
                try:
                    current[key] = int(val_str)
                except ValueError:
                    try:
                        current[key] = float(val_str)
                    except ValueError:
                        current[key] = val_str
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  CUSTOM EXCEPTIONS
# ══════════════════════════════════════════════════════════════════════════════

class DataLoadError(Exception):
    pass


class ForecastError(Exception):
    pass


class ReportGenerationError(Exception):
    pass


# ══════════════════════════════════════════════════════════════════════════════
#  DATA MODELS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ResidualDiagnostics:
    ljung_box_stat:  float = 0.0
    ljung_box_pval:  float = 1.0
    shapiro_stat:    float = 0.0
    shapiro_pval:    float = 1.0
    durbin_watson:   float = 2.0
    heteroscedastic: bool  = False

    @property
    def is_adequate(self) -> bool:
        return self.ljung_box_pval > 0.05 and 1.0 < self.durbin_watson < 3.0

    def summary(self) -> str:
        lb = "✓" if self.ljung_box_pval > 0.05 else "✗"
        sw = "✓" if self.shapiro_pval   > 0.05 else "✗"
        dw = "✓" if 1.5 < self.durbin_watson < 2.5 else "✗"
        return (f"Ljung-Box {lb} (p={self.ljung_box_pval:.3f})  "
                f"Shapiro {sw}  DW={self.durbin_watson:.2f} {dw}")


@dataclass
class ForecastResult:
    model_name:         str
    predictions:        "np.ndarray"
    lower_ci:           "np.ndarray"
    upper_ci:           "np.ndarray"
    mae:                float = 0.0
    rmse:               float = 0.0
    smape:              float = 100.0
    mase:               float = 1.0
    mape:               float = 100.0
    r2:                 float = -1.0
    aic:                Optional[float] = None
    bic:                Optional[float] = None
    cv_smape:           Optional[float] = None
    feature_importance: Optional[Dict[str, float]] = None
    diagnostics:        Optional[ResidualDiagnostics] = None
    metadata:           Dict[str, Any] = field(default_factory=dict)
    train_time_s:       float = 0.0

    @property
    def is_valid(self) -> bool:
        return (self.predictions is not None and
                len(self.predictions) > 0 and
                np.all(np.isfinite(self.predictions)) and
                math.isfinite(self.smape))

    @property
    def accuracy_score(self) -> float:
        return max(0.0, 100.0 - self.smape)


@dataclass
class SalesAnalysis:
    raw_data:          "pd.DataFrame"
    trend:             "np.ndarray"
    seasonal:          "np.ndarray"
    residual:          "np.ndarray"
    anomalies:         "pd.DataFrame"
    yoy_growth:        "pd.Series"
    mom_growth:        "pd.Series"
    rolling_avg:       "pd.Series"
    statistics:        Dict[str, Any]
    forecasts:         Dict[str, Optional[ForecastResult]]
    best_model:        str
    ensemble_forecast: ForecastResult
    detected_period:   int


# ══════════════════════════════════════════════════════════════════════════════
#  STATISTICAL UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

class StatUtils:
    @staticmethod
    def smape(actual: "np.ndarray", pred: "np.ndarray") -> float:
        a, p = np.asarray(actual, float), np.asarray(pred, float)
        denom = np.abs(a) + np.abs(p)
        mask  = denom > 1e-10
        if not mask.any():
            return 0.0
        return float(np.mean(2.0 * np.abs(a[mask] - p[mask]) / denom[mask]) * 100.0)

    @staticmethod
    def mase(actual: "np.ndarray", pred: "np.ndarray",
             train: "np.ndarray", period: int = 12) -> float:
        a, p = np.asarray(actual, float), np.asarray(pred, float)
        tr   = np.asarray(train, float)
        if len(tr) <= period:
            return float("nan")
        naive_err = float(np.mean(np.abs(tr[period:] - tr[:-period])))
        if naive_err < 1e-10:
            return float("nan")
        return float(np.mean(np.abs(a - p)) / naive_err)

    @staticmethod
    def all_metrics(
        actual: "np.ndarray", pred: "np.ndarray",
        train: "np.ndarray", period: int = 12,
    ) -> Tuple[float, float, float, float, float, float]:
        a, p = np.asarray(actual, float), np.asarray(pred, float)
        mae   = float(mean_absolute_error(a, p))
        rmse  = float(np.sqrt(mean_squared_error(a, p)))
        smape = StatUtils.smape(a, p)
        mase_ = StatUtils.mase(a, p, train, period)
        denom = np.abs(a)
        mask  = denom > 1e-10
        mape  = float(np.mean(np.abs(a[mask] - p[mask]) / denom[mask]) * 100.0) if mask.any() else float("nan")
        try:
            r2 = float(r2_score(a, p))
        except Exception:
            r2 = float("nan")
        return mae, rmse, smape, mase_, mape, r2

    @staticmethod
    def bootstrap_ci(
        preds: "np.ndarray", residuals: "np.ndarray",
        alpha: float = 0.05, n_boot: int = 500,
    ) -> Tuple["np.ndarray", "np.ndarray"]:
        rng   = np.random.default_rng(42)
        n     = len(preds)
        boots = np.zeros((n_boot, n))
        for b in range(n_boot):
            sample = rng.choice(residuals, size=n, replace=True)
            boots[b] = preds + sample
        lo = np.percentile(boots, 100.0 * alpha / 2,       axis=0)
        hi = np.percentile(boots, 100.0 * (1 - alpha / 2), axis=0)
        return lo, hi

    @staticmethod
    def compute_diagnostics(residuals: "np.ndarray") -> ResidualDiagnostics:
        d    = ResidualDiagnostics()
        resid = np.asarray(residuals, float)
        n     = len(resid)
        if n < 4:
            return d
        resid = resid[np.isfinite(resid)]
        n     = len(resid)
        if n < 4:
            return d
        try:
            max_lag = max(1, min(20, n // 2 - 2))
            acf_sq_sum = 0.0
            for k in range(1, max_lag + 1):
                if n - k >= 2:
                    r = np.corrcoef(resid[:n - k], resid[k:])[0, 1]
                    if np.isfinite(r):
                        acf_sq_sum += r ** 2
            lb_stat = float(n * (n + 2) * acf_sq_sum)
            from scipy.stats import chi2
            lb_pval = float(1.0 - chi2.cdf(lb_stat, df=max_lag))
            d.ljung_box_stat = lb_stat
            d.ljung_box_pval = lb_pval
        except Exception:
            pass
        try:
            sw_stat, sw_pval = stats.shapiro(resid[:min(5000, n)])
            d.shapiro_stat = float(sw_stat)
            d.shapiro_pval = float(sw_pval)
        except Exception:
            pass
        try:
            dw = float(np.sum(np.diff(resid) ** 2) / max(np.dot(resid, resid), 1e-20))
            d.durbin_watson = dw
        except Exception:
            pass
        try:
            half = n // 2
            if half >= 4:
                var1 = float(np.var(resid[:half], ddof=1))
                var2 = float(np.var(resid[half:], ddof=1))
                ratio = max(var1, var2) / max(min(var1, var2), 1e-20)
                d.heteroscedastic = ratio > 4.0
        except Exception:
            pass
        return d

    @staticmethod
    def detect_seasonality(y: "np.ndarray") -> int:
        y = np.asarray(y, float)
        n = len(y)
        candidates = [12, 4, 6, 3, 52, 7, 24]
        try:
            freqs, power = periodogram(y - float(np.mean(y)), fs=1.0)
            total = float(np.sum(power))
            for period in sorted(candidates, reverse=True):
                if period >= n // 2 or period < 2:
                    continue
                target_freq = 1.0 / period
                idx    = int(np.argmin(np.abs(freqs - target_freq)))
                window = max(1, int(len(freqs) * 0.01))
                band   = power[max(0, idx - window): idx + window + 1]
                if float(np.sum(band)) / total > 0.15:
                    return period
        except Exception:
            pass
        return 12

    @staticmethod
    def marketing_roi_with_ci(
        sales: "np.ndarray", spend: "np.ndarray",
        n_boot: int = 1000, alpha: float = 0.05,
    ) -> Dict[str, float]:
        if len(sales) < 4 or np.sum(spend) < 1e-6:
            return {"point": 0.0, "ci_lo": 0.0, "ci_hi": 0.0}
        rng       = np.random.default_rng(42)
        n         = len(sales)
        roi_boots = np.zeros(n_boot)
        for b in range(n_boot):
            idx          = rng.integers(0, n, size=n)
            s_b          = float(np.sum(spend[idx]))
            r_b          = float(np.sum(sales[idx]))
            roi_boots[b] = (r_b - s_b) / max(s_b, 1e-10)
        point = float((np.sum(sales) - np.sum(spend)) / max(float(np.sum(spend)), 1e-10))
        lo    = float(np.percentile(roi_boots, 100.0 * alpha / 2))
        hi    = float(np.percentile(roi_boots, 100.0 * (1.0 - alpha / 2)))
        return {"point": point, "ci_lo": lo, "ci_hi": hi}


# ══════════════════════════════════════════════════════════════════════════════
#  DATA GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class SalesDataGenerator:
    _PRODUCTS = {
        "Product A": {"base": 150_000, "trend": 3_000,  "seasonal_amp": 0.18},
        "Product B": {"base":  95_000, "trend": 1_800,  "seasonal_amp": 0.25},
        "Product C": {"base": 210_000, "trend": 4_500,  "seasonal_amp": 0.12},
        "Product D": {"base":  70_000, "trend": 2_200,  "seasonal_amp": 0.30},
        "Product E": {"base": 130_000, "trend":  -500,  "seasonal_amp": 0.20},
    }
    _REGIONS = ["North", "South", "East", "West"]

    def __init__(self, seed: Optional[int] = None):
        actual_seed = seed if seed is not None else int(time.time()) % 100_000
        self.rng = np.random.default_rng(actual_seed)

    def generate(
        self, n_months: int = 60, n_products: int = 5,
        start_date: str = "2020-01-01",
    ) -> "pd.DataFrame":
        dates    = pd.date_range(start=start_date, periods=n_months, freq="MS")
        t        = np.arange(n_months, dtype=float)
        products = list(self._PRODUCTS.items())[:n_products]
        records: List[Dict] = []
        for prod_name, params in products:
            base          = float(params["base"])
            trend_comp    = base + float(params["trend"]) * t
            phase         = float(self.rng.uniform(0, np.pi))
            seasonal_comp = float(params["seasonal_amp"]) * base * np.sin(2 * np.pi * t / 12 + phase)
            holiday = np.array([
                0.15 * base  if d.month in (11, 12)
                else -0.08 * base if d.month in (1, 2)
                else 0.0
                for d in dates
            ])
            noise = self.rng.normal(0, 0.04 * base, n_months)
            sales = np.maximum(trend_comp + seasonal_comp + holiday + noise, 0.0)
            for ai in self.rng.integers(0, n_months, size=2):
                sales[ai] *= float(self.rng.choice([0.4, 1.8]))
            for d, s in zip(dates, sales):
                price_per_unit = float(self.rng.uniform(25, 60))
                records.append({
                    "date":            d,
                    "product":         prod_name,
                    "region":          str(self.rng.choice(self._REGIONS)),
                    "sales":           round(float(s), 2),
                    "units":           max(1, int(s / price_per_unit)),
                    "returns":         round(float(s) * float(self.rng.uniform(0.01, 0.05)), 2),
                    "marketing_spend": round(float(s) * float(self.rng.uniform(0.08, 0.15)), 2),
                })
        df = pd.DataFrame(records)
        df["net_revenue"] = df["sales"] - df["returns"]
        return df


# ══════════════════════════════════════════════════════════════════════════════
#  DATA LOADER & VALIDATOR
# ══════════════════════════════════════════════════════════════════════════════

class DataLoader:
    REQUIRED_COLS     = {"date", "sales"}
    OPTIONAL_DEFAULTS: Dict[str, Any] = {"product": "Default", "region": "Default"}
    MIN_RECORDS       = 12

    _DATE_ALIASES: List[str] = [
        "date", "order_date", "order_date_time", "transaction_date",
        "sale_date", "invoice_date", "purchase_date", "period", "month",
        "year_month", "timestamp", "time", "datetime", "dt", "dates",
        "fiscal_date", "week", "day", "reporting_date", "report_date",
        "order_placed_date", "created_at", "created_date",
    ]
    _SALES_ALIASES: List[str] = [
        "sales", "revenue", "amount", "total", "total_sales", "total_revenue",
        "net_sales", "gross_sales", "sale_amount", "invoice_amount",
        "transaction_amount", "order_amount", "order_value", "value",
        "price", "income", "earnings", "receipts", "turnover",
        "sales_amount", "sale_value", "qty_value", "extended_price",
        "subtotal", "grand_total", "payment", "payment_amount",
    ]
    _PRODUCT_ALIASES: List[str] = [
        "product", "product_name", "item", "item_name", "sku", "category",
        "product_category", "product_type", "goods", "commodity", "line",
        "product_line", "description", "product_description", "service",
    ]
    _REGION_ALIASES: List[str] = [
        "region", "area", "territory", "zone", "location", "state",
        "country", "city", "market", "branch", "office", "division",
        "geo", "geography", "district", "sector",
    ]
    _UNITS_ALIASES: List[str] = [
        "units", "quantity", "qty", "count", "volume", "items",
        "units_sold", "quantity_sold", "num_units", "pieces",
    ]
    _RETURNS_ALIASES: List[str] = [
        "returns", "refunds", "returns_amount", "refund_amount",
        "returned_amount", "chargebacks", "return_value",
    ]
    _MARKETING_ALIASES: List[str] = [
        "marketing_spend", "marketing", "ad_spend", "advertising",
        "marketing_cost", "promo_spend", "promotion_spend", "ad_cost",
        "advertising_spend", "campaign_spend",
    ]
    _ALIAS_MAP: Dict[str, str] = {}

    @classmethod
    def _build_alias_map(cls) -> None:
        mapping: List[Tuple[str, List[str]]] = [
            ("date",            cls._DATE_ALIASES),
            ("sales",           cls._SALES_ALIASES),
            ("product",         cls._PRODUCT_ALIASES),
            ("region",          cls._REGION_ALIASES),
            ("units",           cls._UNITS_ALIASES),
            ("returns",         cls._RETURNS_ALIASES),
            ("marketing_spend", cls._MARKETING_ALIASES),
        ]
        for canonical, aliases in mapping:
            for alias in aliases:
                cls._ALIAS_MAP[alias] = canonical

    @classmethod
    def _normalise_col(cls, raw: str) -> str:
        return (
            str(raw).strip().lstrip("\ufeff").lower()
            .replace(" ", "_").replace("-", "_")
            .replace(".", "_").replace("/", "_").replace("\\", "_")
        )

    @classmethod
    def _map_columns(cls, df: "pd.DataFrame") -> Tuple["pd.DataFrame", Dict[str, str]]:
        if not cls._ALIAS_MAP:
            cls._build_alias_map()
        norm_to_orig: Dict[str, str] = {}
        for orig in df.columns:
            norm = cls._normalise_col(orig)
            if norm not in norm_to_orig:
                norm_to_orig[norm] = orig
        rename: Dict[str, str] = {}
        mapped: Dict[str, str] = {}
        for norm, orig in norm_to_orig.items():
            canonical = cls._ALIAS_MAP.get(norm)
            if canonical and canonical not in mapped:
                rename[orig] = canonical
                mapped[canonical] = orig
        targets_needed = {"date", "sales", "product", "region",
                          "units", "returns", "marketing_spend"} - set(mapped.keys())
        for norm, orig in norm_to_orig.items():
            if orig in rename:
                continue
            for canonical in list(targets_needed):
                keywords = {
                    "date":            ["date", "time", "period", "month", "day", "week"],
                    "sales":           ["sale", "revenue", "amount", "total", "value", "price"],
                    "product":         ["product", "item", "sku", "category", "good"],
                    "region":          ["region", "area", "territ", "zone", "locat", "state"],
                    "units":           ["unit", "qty", "quant", "count", "volume", "piece"],
                    "returns":         ["return", "refund", "chargeback"],
                    "marketing_spend": ["market", "ad_spend", "adverti", "promo"],
                }.get(canonical, [])
                if any(kw in norm for kw in keywords):
                    rename[orig] = canonical
                    mapped[canonical] = orig
                    targets_needed.discard(canonical)
                    break
        if "sales" not in mapped:
            numeric_cols = [
                orig for orig in df.columns
                if orig not in rename and pd.api.types.is_numeric_dtype(df[orig])
            ]
            if numeric_cols:
                best = numeric_cols[0]
                rename[best] = "sales"
                mapped["sales"] = best
                logger.warning("No 'sales' column found; using '%s'.", best)
        if "date" not in mapped:
            for orig in df.columns:
                if orig in rename:
                    continue
                try:
                    parsed = pd.to_datetime(df[orig], errors="coerce")
                    if parsed.notna().mean() >= 0.8:
                        rename[orig] = "date"
                        mapped["date"] = orig
                        break
                except Exception:
                    continue
        df = df.rename(columns=rename)
        return df, rename

    @classmethod
    def load(cls, source: Optional[str] = None) -> "pd.DataFrame":
        if source is None:
            return SalesDataGenerator(seed=42).generate(n_months=60, n_products=5)
        path = Path(source)
        if not path.exists():
            raise DataLoadError(f"File not found: {path}")
        if path.stat().st_size == 0:
            raise DataLoadError(f"File is empty: {path.name}")
        try:
            df = cls._read_file(path)
        except DataLoadError:
            raise
        except Exception as exc:
            raise DataLoadError(f"Failed to read '{path.name}': {exc}") from exc
        return cls._validate_and_clean(df)

    @staticmethod
    def _read_file(path: Path) -> "pd.DataFrame":
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xlsm"):
            try:
                xl = pd.ExcelFile(path, engine="openpyxl")
                for sheet in xl.sheet_names:
                    df = xl.parse(sheet)
                    if not df.empty and len(df.columns) > 1:
                        return df
                return xl.parse(xl.sheet_names[0])
            except Exception as exc:
                raise DataLoadError(f"Could not read Excel: {exc}") from exc
        if suffix == ".xls":
            for engine in ("xlrd", "openpyxl"):
                try:
                    df = pd.read_excel(path, engine=engine)
                    if not df.empty:
                        return df
                except Exception:
                    continue
            raise DataLoadError(f"Could not read .xls '{path.name}'. Install xlrd.")
        if suffix in (".csv", ".tsv", ".txt"):
            sep = "\t" if suffix == ".tsv" else None
            sep_candidates = [sep] if sep else [None, ",", ";", "|", "\t"]
            encodings = ("utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-15")
            for enc in encodings:
                for sep_try in sep_candidates:
                    try:
                        kwargs: Dict[str, Any] = dict(encoding=enc, on_bad_lines="skip")
                        if sep_try is not None:
                            kwargs["sep"] = sep_try
                            kwargs["engine"] = "c"
                        else:
                            kwargs["sep"] = None
                            kwargs["engine"] = "python"
                        df = pd.read_csv(path, **kwargs)
                        if len(df.columns) > 1 and not df.empty:
                            return df
                    except Exception:
                        continue
            raise DataLoadError(f"Could not decode CSV/TSV: {path.name}")
        if suffix == ".json":
            raw = path.read_bytes()
            try:
                df = pd.read_json(io.BytesIO(raw), lines=True)
                if not df.empty and len(df.columns) > 1:
                    return df
            except Exception:
                pass
            for orient in (None, "records", "split", "index", "values", "table"):
                try:
                    kwargs_j: Dict[str, Any] = {}
                    if orient:
                        kwargs_j["orient"] = orient
                    df = pd.read_json(io.BytesIO(raw), **kwargs_j)
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        return df
                except Exception:
                    continue
            raise DataLoadError(f"Could not parse JSON: {path.name}")
        if suffix == ".parquet":
            for engine in ("pyarrow", "fastparquet"):
                try:
                    return pd.read_parquet(path, engine=engine)
                except ImportError:
                    continue
                except Exception as exc:
                    raise DataLoadError(f"Parquet read failed: {exc}") from exc
            raise DataLoadError(
                "No parquet engine available. Install: pip install pyarrow")
        raise DataLoadError(f"Unsupported file type: {suffix}")

    @classmethod
    def _validate_and_clean(cls, df: "pd.DataFrame") -> "pd.DataFrame":
        df, rename_log = cls._map_columns(df)
        for col in ("date", "sales"):
            if col not in df.columns:
                raise DataLoadError(
                    f"Required column '{col}' not found. "
                    f"Columns detected: {list(df.columns)[:15]}")
        df["date"]  = pd.to_datetime(df["date"], errors="coerce", infer_datetime_format=True)
        df["sales"] = pd.to_numeric(df["sales"], errors="coerce")
        df = df.dropna(subset=["date", "sales"])
        if df.empty:
            raise DataLoadError("No valid rows after parsing dates and sales.")
        df = df[df["sales"] >= 0.0].copy()
        df = df.sort_values("date").reset_index(drop=True)
        df["date"] = df["date"].dt.to_period("M").dt.to_timestamp()
        for col, default in cls.OPTIONAL_DEFAULTS.items():
            if col not in df.columns:
                df[col] = default
        for col in ("units", "returns", "marketing_spend"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            else:
                df[col] = 0.0
        if len(df) < cls.MIN_RECORDS:
            raise DataLoadError(
                f"Dataset has only {len(df)} rows; minimum {cls.MIN_RECORDS} required.")
        return df


# ══════════════════════════════════════════════════════════════════════════════
#  ARIMA (custom pure-numpy implementation)
# ══════════════════════════════════════════════════════════════════════════════

class ARIMA:
    _ORDERS = [(1,1,1),(1,1,0),(0,1,1),(2,1,1),(1,1,2),(2,1,2),
               (0,1,2),(2,1,0),(1,2,1),(0,2,1)]

    def __init__(self, p: int = 1, d: int = 1, q: int = 1):
        self.p, self.d, self.q = p, d, q
        self._params: Optional["np.ndarray"] = None
        self._residuals: "np.ndarray" = np.array([])
        self._orig:      "np.ndarray" = np.array([])
        self._diff_series: "np.ndarray" = np.array([])
        self._aic: float = np.inf
        self._bic: float = np.inf

    @staticmethod
    def _difference(y: "np.ndarray", d: int) -> "np.ndarray":
        yd = y.astype(np.float64)
        for _ in range(d):
            yd = np.diff(yd)
        return yd

    @staticmethod
    def _undifference(diffed: "np.ndarray", orig: "np.ndarray", d: int) -> "np.ndarray":
        out = np.asarray(diffed, dtype=np.float64).copy()
        for step in range(d):
            seed = orig[-(d - step)]
            out  = np.concatenate([[seed], out])
            out  = np.cumsum(out)[1:]
        return out

    def _neg_log_likelihood(self, params: "np.ndarray", yd: "np.ndarray") -> float:
        p, q   = self.p, self.q
        ar_c   = params[:p]
        ma_c   = params[p:p + q]
        n      = len(yd)
        resid  = np.zeros(n)
        errors = np.zeros(n)
        for t in range(n):
            ar_part = sum(ar_c[i] * yd[t - i - 1]    for i in range(p) if t - i - 1 >= 0)
            ma_part = sum(ma_c[i] * errors[t - i - 1] for i in range(q) if t - i - 1 >= 0)
            resid[t]  = yd[t] - ar_part - ma_part
            errors[t] = resid[t]
        sigma2 = max(float(np.var(resid, ddof=1)), 1e-10)
        nll    = 0.5 * n * (math.log(2 * math.pi * sigma2) + 1.0)
        return nll if math.isfinite(nll) else 1e20

    def fit(self, y: "np.ndarray") -> "ARIMA":
        self._orig        = np.asarray(y, dtype=np.float64)
        self._diff_series = self._difference(self._orig, self.d)
        yd    = self._diff_series
        n_par = self.p + self.q
        if n_par == 0:
            self._params    = np.array([])
            self._residuals = yd.copy()
        else:
            x0     = np.zeros(n_par)
            bounds = [(-0.99, 0.99)] * n_par
            best_f = np.inf
            best_x = x0.copy()
            for start in [x0, np.ones(n_par) * 0.1, np.ones(n_par) * -0.1,
                           np.random.default_rng(42).uniform(-0.3, 0.3, n_par)]:
                try:
                    res = minimize(
                        self._neg_log_likelihood, start, args=(yd,),
                        method="L-BFGS-B", bounds=bounds,
                        options={"maxiter": 500, "ftol": 1e-12})
                    if np.isfinite(res.fun) and res.fun < best_f:
                        best_f, best_x = res.fun, res.x
                except Exception:
                    pass
            self._params = best_x
            ar_c = best_x[:self.p]
            ma_c = best_x[self.p:]
            n    = len(yd)
            resid  = np.zeros(n)
            errors = np.zeros(n)
            for t in range(n):
                ar_part = sum(ar_c[i] * yd[t-i-1]    for i in range(self.p) if t-i-1 >= 0)
                ma_part = sum(ma_c[i] * errors[t-i-1] for i in range(self.q) if t-i-1 >= 0)
                resid[t]  = yd[t] - ar_part - ma_part
                errors[t] = resid[t]
            self._residuals = resid
        n      = len(yd)
        k      = self.p + self.q + 1
        sigma2 = max(float(np.var(self._residuals, ddof=1)), 1e-10)
        ll     = -0.5 * (n * math.log(2 * math.pi * sigma2) + float(np.sum(self._residuals**2)) / sigma2)
        self._aic = float(-2.0 * ll + 2.0 * k)
        self._bic = float(-2.0 * ll + k * math.log(max(n, 2)))
        return self

    def predict(self, h: int) -> "np.ndarray":
        yd      = self._diff_series.copy()
        ar_c    = self._params[:self.p] if self._params is not None else np.array([])
        ma_c    = self._params[self.p:self.p + self.q] if self._params is not None else np.array([])
        errors  = self._residuals.copy() if len(self._residuals) > 0 else np.zeros(len(yd))
        diff_preds: List[float] = []
        for _ in range(h):
            full_yd = np.concatenate([yd, np.array(diff_preds, dtype=np.float64)])
            full_er = np.concatenate([errors, np.zeros(len(diff_preds))])
            t       = len(full_yd)
            ar_part = sum(float(ar_c[i]) * float(full_yd[t-i-1]) for i in range(self.p) if t-i-1 >= 0)
            ma_part = sum(float(ma_c[i]) * float(full_er[t-i-1]) for i in range(self.q) if t-i-1 >= 0)
            diff_preds.append(ar_part + ma_part)
        dp_arr = np.array(diff_preds, dtype=np.float64)
        return self._undifference(dp_arr, self._orig, self.d)

    def forecast_ci(self, h: int, alpha: float = 0.05) -> Tuple["np.ndarray", "np.ndarray"]:
        preds  = self.predict(h)
        sigma2 = max(float(np.var(self._residuals, ddof=1)), 1e-10) if len(self._residuals) > 0 else 1.0
        z      = float(stats.norm.ppf(1.0 - alpha / 2))
        psi    = np.zeros(h + 1)
        psi[0] = 1.0
        ar_c   = self._params[:self.p] if self._params is not None else np.array([])
        ma_c   = self._params[self.p:self.p + self.q] if self._params is not None else np.array([])
        for j in range(1, h + 1):
            s = sum(float(ar_c[i]) * psi[j-i-1] for i in range(self.p) if j-i-1 >= 0)
            s += (float(ma_c[j-1]) if j-1 < self.q else 0.0)
            psi[j] = s
        var_h  = sigma2 * np.array([float(np.sum(psi[:j+1]**2)) for j in range(h)])
        margin = z * np.sqrt(np.maximum(var_h, 0.0))
        return np.maximum(preds - margin, 0.0), preds + margin

    @classmethod
    def auto_select(cls, y: "np.ndarray") -> "ARIMA":
        best_aic = np.inf
        best_mdl: Optional["ARIMA"] = None
        for p, d, q in cls._ORDERS:
            yd_len = len(y) - d
            if yd_len < max(p, q, 1) + 2:
                continue
            try:
                mdl = cls(p, d, q).fit(y)
                if np.isfinite(mdl._aic) and mdl._aic < best_aic:
                    best_aic = mdl._aic
                    best_mdl = mdl
            except Exception:
                pass
        if best_mdl is None:
            best_mdl = cls(1, 1, 1).fit(y)
        return best_mdl


# ══════════════════════════════════════════════════════════════════════════════
#  HOLT-WINTERS
# ══════════════════════════════════════════════════════════════════════════════

class HoltWinters:
    def __init__(self, season_len: int = 12, multiplicative: bool = False):
        self.season_len    = max(2, season_len)
        self.multiplicative = multiplicative
        self.alpha = self.beta = self.gamma = 0.3
        self._level = self._trend = 0.0
        self._seasonal: List[float] = []
        self._n_obs    = 0
        self._resid:   "np.ndarray" = np.array([])
        self._sigma2   = 1.0
        self._aic      = np.inf
        self._fitted   = False

    @staticmethod
    def _smooth(
        y: "np.ndarray", alpha: float, beta: float, gamma: float,
        season_len: int, multiplicative: bool,
    ) -> Tuple[float, float, List[float], "np.ndarray"]:
        n = len(y)
        m = season_len
        if multiplicative:
            level  = float(np.mean(y[:m]))
            trend  = float((np.mean(y[m:2*m]) - np.mean(y[:m])) / m) if n >= 2 * m else 0.0
            seas   = [float(y[i]) / max(level, 1e-10) for i in range(m)]
        else:
            level  = float(np.mean(y[:m]))
            trend  = float((np.mean(y[m:2*m]) - np.mean(y[:m])) / m) if n >= 2 * m else 0.0
            seas   = [float(y[i]) - level for i in range(m)]
        fitted = np.empty(n, dtype=np.float64)
        for t in range(n):
            s_idx = t % m
            if multiplicative:
                fitted[t] = (level + trend) * max(seas[s_idx], 1e-9)
                err       = float(y[t])
                l_new     = alpha * (err / max(seas[s_idx], 1e-9)) + (1 - alpha) * (level + trend)
                b_new     = beta  * (l_new - level) + (1 - beta) * trend
                s_new     = gamma * (err / max(l_new, 1e-9)) + (1 - gamma) * seas[s_idx]
            else:
                fitted[t] = level + trend + seas[s_idx]
                err       = float(y[t])
                l_new     = alpha * (err - seas[s_idx]) + (1 - alpha) * (level + trend)
                b_new     = beta  * (l_new - level) + (1 - beta) * trend
                s_new     = gamma * (err - l_new) + (1 - gamma) * seas[s_idx]
            seas[s_idx] = s_new
            level, trend = l_new, b_new
        return level, trend, seas, fitted

    def _sse(self, params: List[float], y: "np.ndarray") -> float:
        a, b, g = params
        try:
            n = len(y)
            if n < 2 * self.season_len:
                return 1e20
            _, _, _, fitted = self._smooth(y, a, b, g, self.season_len, self.multiplicative)
            return float(np.sum((y - fitted) ** 2))
        except Exception:
            return 1e20

    def fit(self, y: "np.ndarray") -> "HoltWinters":
        best_sse    = np.inf
        best_params = [0.3, 0.1, 0.2]
        starts      = [
            (0.3, 0.1, 0.2), (0.5, 0.2, 0.3), (0.1, 0.05, 0.1),
            (0.7, 0.3, 0.4), (0.2, 0.05, 0.15), (0.4, 0.15, 0.25),
            (0.6, 0.25, 0.35), (0.15, 0.08, 0.12),
        ]
        for start in starts:
            try:
                res = minimize(
                    self._sse, list(start), args=(y,),
                    bounds=[(0.01, 0.9999)] * 3, method="L-BFGS-B",
                    options={"maxiter": 400, "ftol": 1e-12})
                if np.isfinite(res.fun) and res.fun < best_sse:
                    best_sse, best_params = res.fun, res.x.tolist()
            except Exception:
                pass
        self.alpha, self.beta, self.gamma = [float(x) for x in best_params]
        level, trend, seas, fitted = self._smooth(
            y, self.alpha, self.beta, self.gamma, self.season_len, self.multiplicative)
        self._level    = level
        self._trend    = trend
        self._seasonal = seas
        self._n_obs    = len(y)
        resid          = y - fitted
        self._resid    = resid
        n              = len(y)
        k              = 4
        sigma2         = max(float(np.var(resid, ddof=1)), 1e-20)
        self._sigma2   = sigma2
        ll             = -0.5 * (n * math.log(2 * math.pi * sigma2)
                                  + float(best_sse) / sigma2)
        self._aic      = float(-2.0 * ll + 2.0 * k)
        self._fitted   = True
        return self

    def predict(self, h: int) -> "np.ndarray":
        if not self._fitted:
            raise RuntimeError("Model not fitted.")
        m     = self.season_len
        preds = []
        for i in range(1, h + 1):
            s_idx = (i - 1) % m
            s_v   = self._seasonal[s_idx]
            base  = self._level + i * self._trend
            if self.multiplicative:
                p = base * (s_v if abs(s_v) > 1e-9 else 1.0)
            else:
                p = base + s_v
            preds.append(p)
        return np.array(preds, dtype=float)

    def forecast_ci(self, h: int, alpha: float = 0.05) -> Tuple["np.ndarray", "np.ndarray"]:
        preds = self.predict(h)
        if not self.multiplicative:
            z     = float(stats.norm.ppf(1.0 - alpha / 2))
            a, b  = self.alpha, self.beta
            var_h = np.array([
                self._sigma2 * max(
                    1.0 + max(i - 1, 0) * (a**2 + a * b * i + b**2 * i * (2*i - 1) / 6.0),
                    1e-9)
                for i in range(1, h + 1)
            ])
            margin = z * np.sqrt(var_h)
            return np.maximum(preds - margin, 0.0), preds + margin
        else:
            lo, hi = StatUtils.bootstrap_ci(preds, self._resid, alpha=alpha, n_boot=400)
            return np.maximum(lo, 0.0), hi

    @classmethod
    def auto_select(cls, y: "np.ndarray", season_len: int = 12) -> "HoltWinters":
        best_aic = np.inf
        best_mdl: Optional["HoltWinters"] = None
        for mul in (False, True):
            try:
                mdl = cls(season_len=season_len, multiplicative=mul)
                mdl.fit(y)
                if np.isfinite(mdl._aic) and mdl._aic < best_aic:
                    best_aic = mdl._aic
                    best_mdl = mdl
            except Exception:
                pass
        if best_mdl is None:
            best_mdl = cls(season_len=season_len).fit(y)
        return best_mdl


# ══════════════════════════════════════════════════════════════════════════════
#  ML FEATURE ENGINEERING
# ══════════════════════════════════════════════════════════════════════════════

class MLFeatureEngineering:
    @staticmethod
    def _pacf(y: "np.ndarray", max_lag: int) -> "np.ndarray":
        n    = len(y)
        pacf = np.zeros(max_lag + 1)
        pacf[0] = 1.0
        if n < 4 or max_lag < 1:
            return pacf
        ym  = y - float(np.mean(y))
        var = float(np.dot(ym, ym)) / n
        if var < 1e-14:
            return pacf
        acv      = np.array([float(np.dot(ym[:n - k], ym[k:])) / n for k in range(max_lag + 1)])
        acf_vals = acv / max(acv[0], 1e-14)
        phi      = np.zeros((max_lag + 1, max_lag + 1))
        phi[1, 1] = acf_vals[1]
        pacf[1]   = acf_vals[1]
        for k in range(2, max_lag + 1):
            num = acf_vals[k] - np.dot(phi[k-1, 1:k], acf_vals[k-1:0:-1])
            den = 1.0 - np.dot(phi[k-1, 1:k], acf_vals[1:k])
            if abs(den) < 1e-12:
                break
            phi[k, k] = num / den
            for j in range(1, k):
                phi[k, j] = phi[k-1, j] - phi[k, k] * phi[k-1, k-j]
            pacf[k] = phi[k, k]
        return pacf

    @classmethod
    def select_lags(cls, y: "np.ndarray", max_lag: int = 24, period: int = 12) -> List[int]:
        n         = len(y)
        max_lag   = min(max_lag, n // 3)
        if max_lag < 1:
            return [1]
        threshold = 2.0 / math.sqrt(max(n, 4))
        try:
            pacf_vals = cls._pacf(y, max_lag)
        except Exception:
            pacf_vals = np.zeros(max_lag + 1)
        selected: set = {1}
        for lag in range(1, max_lag + 1):
            if abs(pacf_vals[lag]) > threshold:
                selected.add(lag)
            if lag % period == 0:
                selected.add(lag)
        return sorted(selected)

    @classmethod
    def build_features(
        cls, y: "np.ndarray", period: int = 12,
    ) -> Tuple["np.ndarray", "np.ndarray", List[str]]:
        y    = np.asarray(y, dtype=np.float64)
        n    = len(y)
        lags = cls.select_lags(y, max_lag=min(24, max(1, n // 3)), period=period)
        feats: Dict[str, "np.ndarray"] = {}
        for lag in lags:
            if 0 < lag < n:
                padded      = np.empty(n, dtype=np.float64)
                padded[:lag] = np.nan
                padded[lag:] = y[:n - lag]
                feats[f"lag_{lag}"] = padded
        s = pd.Series(y)
        for w in [3, 6, 12]:
            if w <= n:
                feats[f"roll_mean_{w}"] = s.rolling(w, min_periods=1).mean().to_numpy(dtype=np.float64)
                feats[f"roll_std_{w}"]  = s.rolling(w, min_periods=1).std().fillna(0).to_numpy(dtype=np.float64)
                feats[f"roll_min_{w}"]  = s.rolling(w, min_periods=1).min().to_numpy(dtype=np.float64)
                feats[f"roll_max_{w}"]  = s.rolling(w, min_periods=1).max().to_numpy(dtype=np.float64)
        t_arr = np.arange(n, dtype=np.float64)
        feats["t"]      = t_arr
        feats["t_sq"]   = t_arr ** 2
        feats["t_sqrt"] = np.sqrt(t_arr)
        for harmonic in range(1, 3):
            feats[f"sin_{harmonic}"] = np.sin(2 * np.pi * harmonic * t_arr / period)
            feats[f"cos_{harmonic}"] = np.cos(2 * np.pi * harmonic * t_arr / period)
        df           = pd.DataFrame(feats)
        df["target"] = y
        df           = df.dropna()
        if df.empty:
            raise ValueError("No complete rows after feature engineering.")
        cols    = [c for c in df.columns if c != "target"]
        X       = df[cols].to_numpy(dtype=np.float64)
        y_clean = df["target"].to_numpy(dtype=np.float64)
        return X, y_clean, cols


# ══════════════════════════════════════════════════════════════════════════════
#  WALK-FORWARD CROSS VALIDATION
# ══════════════════════════════════════════════════════════════════════════════

class WalkForwardCV:
    @staticmethod
    def evaluate(
        model_fn: Callable, y: "np.ndarray",
        period: int = 12, n_splits: int = 3, horizon: int = 6,
    ) -> Dict[str, float]:
        n         = len(y)
        min_train = max(period * 2, 24)
        horizon   = min(horizon, n // 4)
        if n < min_train + horizon:
            return {"smape": float("nan"), "mase": float("nan"), "rmse": float("nan")}
        available  = n - min_train - horizon
        fold_step  = max(1, available // n_splits)
        split_ends = [min_train + (i + 1) * fold_step for i in range(n_splits)]
        split_ends = [s for s in split_ends if s + horizon <= n]
        if not split_ends:
            return {"smape": float("nan"), "mase": float("nan"), "rmse": float("nan")}
        smapes, mases, rmses = [], [], []
        for end in split_ends:
            train  = y[:end]
            actual = y[end: end + horizon]
            if len(actual) == 0:
                continue
            try:
                raw_preds = model_fn(train)
                preds     = np.asarray(raw_preds, dtype=float)[:len(actual)]
                if len(preds) < len(actual) or np.any(~np.isfinite(preds)):
                    continue
                smapes.append(StatUtils.smape(actual, preds))
                mases.append(StatUtils.mase(actual, preds, train, period))
                rmses.append(float(np.sqrt(mean_squared_error(actual, preds))))
            except Exception:
                pass
        if not smapes:
            return {"smape": float("nan"), "mase": float("nan"), "rmse": float("nan")}
        return {
            "smape": float(np.mean(smapes)),
            "mase":  float(np.mean(mases)),
            "rmse":  float(np.mean(rmses)),
        }


# ══════════════════════════════════════════════════════════════════════════════
#  CANCELLATION TOKEN
# ══════════════════════════════════════════════════════════════════════════════

class CancelToken:
    def __init__(self):
        self._event = threading.Event()

    def cancel(self) -> None:
        self._event.set()

    @property
    def is_cancelled(self) -> bool:
        return self._event.is_set()

    def check(self) -> None:
        if self._event.is_set():
            raise InterruptedError("Analysis cancelled by user.")


# ══════════════════════════════════════════════════════════════════════════════
#  FORECASTING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class ForecastingEngine:
    def __init__(
        self, config: ForecastConfig,
        progress_cb: Optional[Callable] = None,
        cancel_token: Optional[CancelToken] = None,
    ):
        self.config  = config
        self._cb     = progress_cb or (lambda msg, pct: None)
        self._cancel = cancel_token or CancelToken()

    def _report(self, msg: str, pct: Optional[float] = None) -> None:
        try:
            self._cb(msg, pct)
        except Exception:
            pass

    def _check(self) -> None:
        self._cancel.check()

    def run_arima(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_arima:
            return None
        self._report("Fitting ARIMA (MLE grid search) …")
        t0 = time.perf_counter()
        try:
            self._check()
            h     = self.config.forecast_horizon
            alpha = 1.0 - self.config.confidence_level

            def arima_fn(y_tr: "np.ndarray") -> "np.ndarray":
                return ARIMA.auto_select(y_tr).predict(h)

            cv_scores = WalkForwardCV.evaluate(
                arima_fn, y, period=period,
                n_splits=self.config.cv_folds, horizon=min(h, len(y) // 4))
            split       = max(h, len(y) // 5)
            train, test = y[:-split], y[-split:]
            val_mdl     = ARIMA.auto_select(train)
            val_preds   = np.maximum(val_mdl.predict(split)[:len(test)], 0.0)
            mae, rmse, smape, mase_, mape, r2 = StatUtils.all_metrics(test, val_preds, train, period)
            final  = ARIMA(val_mdl.p, val_mdl.d, val_mdl.q).fit(y)
            preds  = np.maximum(final.predict(h), 0.0)
            lo, hi = final.forecast_ci(h, alpha=alpha)
            lo     = np.maximum(lo, 0.0)
            diag   = StatUtils.compute_diagnostics(final._residuals)
            return ForecastResult(
                "ARIMA", preds, lo, hi, mae, rmse, smape, mase_, mape, r2,
                aic=final._aic, bic=final._bic,
                cv_smape=cv_scores.get("smape"),
                diagnostics=diag,
                metadata={"order": (final.p, final.d, final.q)},
                train_time_s=time.perf_counter() - t0,
            )
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"ARIMA failed: {exc}")
            logger.exception("ARIMA error")
            return None

    def run_exp_smoothing(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_exp_smoothing:
            return None
        self._report("Fitting Holt-Winters (auto add/mul) …")
        t0 = time.perf_counter()
        try:
            self._check()
            h     = self.config.forecast_horizon
            alpha = 1.0 - self.config.confidence_level

            def hw_fn(y_tr: "np.ndarray") -> "np.ndarray":
                return np.maximum(HoltWinters.auto_select(y_tr, season_len=period).predict(h), 0.0)

            cv_scores   = WalkForwardCV.evaluate(hw_fn, y, period=period,
                              n_splits=self.config.cv_folds, horizon=min(h, len(y) // 4))
            split       = max(h, len(y) // 5)
            train, test = y[:-split], y[-split:]
            val_mdl     = HoltWinters.auto_select(train, season_len=period)
            val_preds   = np.maximum(val_mdl.predict(split)[:len(test)], 0.0)
            mae, rmse, smape, mase_, mape, r2 = StatUtils.all_metrics(test, val_preds, train, period)
            final   = HoltWinters.auto_select(y, season_len=period)
            preds   = np.maximum(final.predict(h), 0.0)
            lo, hi  = final.forecast_ci(h, alpha=alpha)
            lo      = np.maximum(lo, 0.0)
            diag    = StatUtils.compute_diagnostics(final._resid)
            variant = "Multiplicative" if final.multiplicative else "Additive"
            return ForecastResult(
                "Holt-Winters", preds, lo, hi, mae, rmse, smape, mase_, mape, r2,
                aic=final._aic,
                cv_smape=cv_scores.get("smape"),
                diagnostics=diag,
                metadata={"variant": variant,
                           "alpha": round(final.alpha, 4),
                           "beta":  round(final.beta, 4),
                           "gamma": round(final.gamma, 4)},
                train_time_s=time.perf_counter() - t0,
            )
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"Holt-Winters failed: {exc}")
            return None

    def _run_ml_model(
        self, y: "np.ndarray",
        base_model: Any,
        model_name: str,
        period: int = 12,
    ) -> Optional[ForecastResult]:
        self._report(f"Fitting {model_name} …")
        t0 = time.perf_counter()
        try:
            self._check()
            h     = self.config.forecast_horizon
            alpha = 1.0 - self.config.confidence_level
            X_full, y_full, feat_names = MLFeatureEngineering.build_features(y, period=period)
            if len(X_full) < 12:
                self._report(f"{model_name}: insufficient data.")
                return None
            n_splits = min(self.config.cv_folds, max(2, len(X_full) // (h + 1)))
            tscv     = TimeSeriesSplit(n_splits=n_splits,
                                       test_size=min(h, max(1, len(X_full) // 4)))
            cv_preds, cv_actuals = [], []
            for tr_idx, te_idx in tscv.split(X_full):
                X_tr, X_te = X_full[tr_idx], X_full[te_idx]
                y_tr, y_te = y_full[tr_idx], y_full[te_idx]
                sc       = StandardScaler()
                fold_mdl = sklearn_clone(base_model)
                X_tr_sc  = sc.fit_transform(X_tr)
                X_te_sc  = sc.transform(X_te)
                fold_mdl.fit(X_tr_sc, y_tr)
                prd = np.maximum(fold_mdl.predict(X_te_sc), 0.0)
                cv_preds.extend(prd.tolist())
                cv_actuals.extend(y_te.tolist())
            cv_smape = (StatUtils.smape(np.array(cv_actuals), np.array(cv_preds))
                        if cv_preds else None)
            split      = max(h, len(X_full) // 5)
            X_tr, X_te = X_full[:-split], X_full[-split:]
            y_tr, y_te = y_full[:-split], y_full[-split:]
            val_scaler = StandardScaler()
            val_mdl    = sklearn_clone(base_model)
            val_mdl.fit(val_scaler.fit_transform(X_tr), y_tr)
            val_preds  = np.maximum(val_mdl.predict(val_scaler.transform(X_te)), 0.0)
            val_resid  = y_te - val_preds
            mae, rmse, smape, mase_, mape, r2 = StatUtils.all_metrics(y_te, val_preds, y_tr, period)
            self._check()
            full_scaler = StandardScaler()
            full_mdl    = sklearn_clone(base_model)
            full_mdl.fit(full_scaler.fit_transform(X_full), y_full)
            future_y = list(y.copy().astype(np.float64))
            preds: List[float] = []
            for step in range(h):
                try:
                    X_fut, _, new_feat_names = MLFeatureEngineering.build_features(
                        np.array(future_y, dtype=np.float64), period=period)
                except ValueError:
                    break
                X_fut_df  = pd.DataFrame(X_fut, columns=new_feat_names)
                X_fut_df  = X_fut_df.reindex(columns=feat_names, fill_value=0.0)
                X_fut_arr = X_fut_df.to_numpy(dtype=np.float64)
                if X_fut_arr.shape[1] != len(feat_names):
                    break
                X_fut_arr = np.nan_to_num(X_fut_arr, nan=0.0, posinf=0.0, neginf=0.0)
                if len(X_fut_arr) == 0:
                    break
                p = float(max(0.0, full_mdl.predict(full_scaler.transform(X_fut_arr[-1:]))[0]))
                preds.append(p)
                future_y.append(p)
            if not preds:
                return None
            preds_arr = np.array(preds[:h], dtype=float)
            lo, hi    = StatUtils.bootstrap_ci(preds_arr, val_resid, alpha=alpha,
                                               n_boot=self.config.n_bootstrap)
            lo = np.maximum(lo, 0.0)
            diag     = StatUtils.compute_diagnostics(val_resid)
            feat_imp = None
            if hasattr(full_mdl, "feature_importances_"):
                pairs    = zip(feat_names, full_mdl.feature_importances_)
                feat_imp = dict(sorted(pairs, key=lambda x: x[1], reverse=True)[:10])
            return ForecastResult(
                model_name, preds_arr, lo, hi, mae, rmse, smape, mase_, mape, r2,
                cv_smape=cv_smape,
                feature_importance=feat_imp,
                diagnostics=diag,
                train_time_s=time.perf_counter() - t0,
            )
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"{model_name} failed: {exc}")
            logger.exception("%s error", model_name)
            return None

    def run_random_forest(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_random_forest:
            return None
        return self._run_ml_model(
            y,
            RandomForestRegressor(n_estimators=200, max_depth=8, random_state=42,
                                  min_samples_leaf=2, max_features=0.8, n_jobs=-1),
            "Random Forest", period,
        )

    def run_gradient_boosting(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_gradient_boosting:
            return None
        return self._run_ml_model(
            y,
            GradientBoostingRegressor(n_estimators=200, learning_rate=0.05,
                                      max_depth=4, subsample=0.8, random_state=42,
                                      min_samples_leaf=3),
            "Gradient Boosting", period,
        )

    def run_ridge(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_ridge:
            return None
        return self._run_ml_model(y, Ridge(alpha=1.0), "Ridge Regression", period)

    def run_xgboost(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_xgboost or not _XGBOOST_AVAILABLE:
            return None
        try:
            model = xgb.XGBRegressor(
                n_estimators=300, learning_rate=0.05, max_depth=4,
                subsample=0.8, colsample_bytree=0.8, random_state=42,
                n_jobs=-1, verbosity=0, objective="reg:squarederror",
                reg_alpha=0.1, reg_lambda=1.0,
            )
            return self._run_ml_model(y, model, "XGBoost", period)
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"XGBoost failed: {exc}")
            return None

    def run_lightgbm(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_lightgbm or not _LIGHTGBM_AVAILABLE:
            return None
        try:
            model = lgb.LGBMRegressor(
                n_estimators=300, learning_rate=0.05, num_leaves=31,
                subsample=0.8, colsample_bytree=0.8, random_state=42,
                n_jobs=-1, verbose=-1, reg_alpha=0.1, reg_lambda=1.0,
            )
            return self._run_ml_model(y, model, "LightGBM", period)
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"LightGBM failed: {exc}")
            return None

    def run_prophet(
        self, y: "np.ndarray", dates: "pd.DatetimeIndex",
        period: int = 12,
    ) -> Optional[ForecastResult]:
        if not self.config.use_prophet or not _PROPHET_AVAILABLE:
            return None
        self._report("Fitting Prophet (trend + seasonality) …")
        t0 = time.perf_counter()
        try:
            self._check()
            h     = self.config.forecast_horizon
            alpha = 1.0 - self.config.confidence_level

            import logging as _logging
            _logging.getLogger("prophet").setLevel(_logging.ERROR)
            _logging.getLogger("cmdstanpy").setLevel(_logging.ERROR)

            prophet_df = pd.DataFrame({"ds": dates, "y": y})

            split_n   = max(h, len(y) // 5)
            train_df  = prophet_df.iloc[:-split_n]
            test_df   = prophet_df.iloc[-split_n:]

            def _fit_prophet(df_train: "pd.DataFrame") -> "Prophet":
                m = Prophet(
                    changepoint_prior_scale=self.config.prophet_changepoint_scale,
                    seasonality_mode=self.config.prophet_seasonality_mode,
                    yearly_seasonality=True,
                    weekly_seasonality=False,
                    daily_seasonality=False,
                    interval_width=self.config.confidence_level,
                )
                if period == 12:
                    m.add_seasonality(name="monthly", period=30.5, fourier_order=5)
                m.fit(df_train)
                return m

            val_mdl    = _fit_prophet(train_df)
            val_future = val_mdl.make_future_dataframe(periods=split_n, freq="MS")
            val_fc     = val_mdl.predict(val_future)
            val_preds  = val_fc["yhat"].values[-split_n:][:len(test_df)]
            val_preds  = np.maximum(val_preds, 0.0).astype(float)
            y_te       = test_df["y"].values
            val_resid  = y_te - val_preds[:len(y_te)]
            mae, rmse, smape, mase_, mape, r2 = StatUtils.all_metrics(
                y_te, val_preds[:len(y_te)], train_df["y"].values, period)

            self._check()
            final_mdl    = _fit_prophet(prophet_df)
            future_frame = final_mdl.make_future_dataframe(periods=h, freq="MS")
            forecast_df  = final_mdl.predict(future_frame)
            fc_tail      = forecast_df.tail(h)
            preds        = np.maximum(fc_tail["yhat"].values, 0.0).astype(float)
            lo_raw       = fc_tail["yhat_lower"].values
            hi_raw       = fc_tail["yhat_upper"].values
            lo           = np.maximum(lo_raw, 0.0).astype(float)
            hi           = hi_raw.astype(float)

            meta = {
                "changepoint_prior_scale": self.config.prophet_changepoint_scale,
                "seasonality_mode":        self.config.prophet_seasonality_mode,
            }
            return ForecastResult(
                "Prophet", preds, lo, hi, mae, rmse, smape, mase_, mape, r2,
                cv_smape=None,
                diagnostics=StatUtils.compute_diagnostics(val_resid),
                metadata=meta,
                train_time_s=time.perf_counter() - t0,
            )
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"Prophet failed: {exc}")
            logger.exception("Prophet error")
            return None

    def build_ensemble(
        self, results: Dict[str, Optional[ForecastResult]],
        y: "np.ndarray", period: int = 12,
    ) -> ForecastResult:
        valid = {k: v for k, v in results.items() if v is not None and v.is_valid}
        if not valid:
            n = len(y)
            h = self.config.forecast_horizon
            naive = np.full(h, float(np.mean(y[-min(12, n):])))
            return ForecastResult("Ensemble (Stacked)", naive,
                                  naive * 0.85, naive * 1.15,
                                  smape=100.0, mase=1.0, mape=100.0, r2=0.0)
        if len(valid) == 1:
            only = next(iter(valid.values()))
            return ForecastResult(
                "Ensemble (Stacked)", only.predictions.copy(),
                only.lower_ci.copy(), only.upper_ci.copy(),
                mae=only.mae, rmse=only.rmse, smape=only.smape,
                mase=only.mase, mape=only.mape, r2=only.r2,
            )
        scores: Dict[str, float] = {}
        for k, v in valid.items():
            s = v.cv_smape if (v.cv_smape is not None and math.isfinite(v.cv_smape)) else v.smape
            scores[k] = max(s, 1e-6)
        inv    = {k: 1.0 / s for k, s in scores.items()}
        total  = sum(inv.values())
        weights = {k: inv[k] / total for k in inv}
        h         = self.config.forecast_horizon
        ens_preds = np.zeros(h)
        ens_lo    = np.zeros(h)
        ens_hi    = np.zeros(h)
        for k, v in valid.items():
            w          = weights[k]
            n          = min(len(v.predictions), h)
            ens_preds[:n] += w * v.predictions[:n]
            ens_lo[:n]    += w * v.lower_ci[:n]
            ens_hi[:n]    += w * v.upper_ci[:n]
        split       = max(h, len(y) // 5)
        train_e, test_e = y[:-split], y[-split:]
        ens_val_preds   = np.zeros(split)
        for k, v in valid.items():
            try:
                if k == "ARIMA":
                    sub = ARIMA(v.metadata.get("order", (1,1,1))[0],
                                v.metadata.get("order", (1,1,1))[1],
                                v.metadata.get("order", (1,1,1))[2]).fit(train_e).predict(split)
                elif k == "Holt-Winters":
                    sub = HoltWinters.auto_select(train_e, season_len=period).predict(split)
                else:
                    sub = np.full(split, float(np.mean(train_e)))
                ens_val_preds += weights[k] * np.maximum(sub[:split], 0.0)
            except Exception:
                pass
        mae, rmse, smape, mase_, mape, r2 = StatUtils.all_metrics(
            test_e, ens_val_preds[:len(test_e)], train_e, period)
        return ForecastResult(
            "Ensemble (Stacked)", ens_preds, ens_lo, ens_hi,
            mae=mae, rmse=rmse, smape=smape, mase=mase_, mape=mape, r2=r2,
            metadata={"weights": {k: round(w, 4) for k, w in weights.items()}},
        )

    def forecast(
        self, y: "np.ndarray", period: int = 12,
        dates: Optional["pd.DatetimeIndex"] = None,
    ) -> Tuple[Dict[str, Optional[ForecastResult]], ForecastResult, str]:
        runners = [
            ("ARIMA",             lambda: self.run_arima(y, period)),
            ("Holt-Winters",      lambda: self.run_exp_smoothing(y, period)),
            ("Random Forest",     lambda: self.run_random_forest(y, period)),
            ("Gradient Boosting", lambda: self.run_gradient_boosting(y, period)),
            ("Ridge Regression",  lambda: self.run_ridge(y, period)),
            ("XGBoost",           lambda: self.run_xgboost(y, period)),
            ("LightGBM",          lambda: self.run_lightgbm(y, period)),
            ("Prophet",           lambda: self.run_prophet(y, dates or pd.DatetimeIndex([]), period)),
        ]
        results: Dict[str, Optional[ForecastResult]] = {}
        total = len(runners)
        for i, (name, fn) in enumerate(runners):
            self._report(f"Running {name} …", 30 + int(i / total * 55))
            try:
                self._check()
                results[name] = fn()
            except InterruptedError:
                raise
            except Exception as exc:
                self._report(f"{name} error: {exc}")
                results[name] = None
        ensemble = self.build_ensemble(results, y, period)
        valid    = {k: v for k, v in results.items() if v is not None and v.is_valid}
        best     = min(valid, key=lambda k: valid[k].smape) if valid else "Ensemble (Stacked)"
        return results, ensemble, best


# ══════════════════════════════════════════════════════════════════════════════
#  ANALYSIS ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class SalesAnalyzer:
    def __init__(
        self, config: ForecastConfig,
        progress_cb: Optional[Callable] = None,
        cancel_token: Optional[CancelToken] = None,
    ):
        self.config  = config
        self._cb     = progress_cb or (lambda msg, pct: None)
        self._cancel = cancel_token or CancelToken()

    def _report(self, msg: str, pct: Optional[float] = None) -> None:
        try:
            self._cb(msg, pct)
        except Exception:
            pass

    @staticmethod
    def decompose(
        y: "np.ndarray", period: int = 12,
    ) -> Tuple["np.ndarray", "np.ndarray", "np.ndarray"]:
        n      = len(y)
        y      = np.asarray(y, dtype=np.float64)
        period = max(2, int(period))
        if n < 2 * period:
            trend    = np.full(n, float(np.mean(y)), dtype=np.float64)
            seasonal = np.zeros(n, dtype=np.float64)
            return trend, seasonal, (y - trend)
        kernel     = np.ones(period, dtype=np.float64) / period
        trend_conv = np.convolve(y, kernel, mode="valid")
        n_valid    = len(trend_conv)
        pad_total  = n - n_valid
        pad_l      = max(0, pad_total // 2)
        pad_r      = max(0, pad_total - pad_l)
        trend = np.empty(n, dtype=np.float64)
        if pad_l > 0:
            trend[:pad_l] = trend_conv[0]
        trend[pad_l:pad_l + n_valid] = trend_conv
        if pad_r > 0:
            trend[pad_l + n_valid:] = trend_conv[-1]
        detrended = y - trend
        s_proto = np.array([float(np.mean(detrended[i::period])) for i in range(period)],
                            dtype=np.float64)
        s_proto  -= float(s_proto.mean())
        seasonal  = np.tile(s_proto, n // period + 2)[:n]
        residual  = y - trend - seasonal
        return trend, seasonal, residual

    @staticmethod
    def detect_anomalies(series: "pd.Series", threshold: float = 2.5) -> "pd.DataFrame":
        w          = min(max(3, len(series) // 8), 12)
        rm         = series.rolling(w, center=True, min_periods=1).mean()
        rs         = series.rolling(w, center=True, min_periods=1).std().fillna(0.0)
        global_std = float(series.std(ddof=1)) if len(series) > 1 else 1.0
        rs         = rs.where(rs > 1e-9, other=max(global_std, 1e-9))
        z          = np.abs((series - rm) / rs)
        idx        = z[z > threshold].index
        if len(idx) == 0:
            return pd.DataFrame(
                columns=["value", "z_score", "type"]
            ).astype({"value": "float64", "z_score": "float64", "type": "object"})
        anom  = series.loc[idx]
        types = np.where(anom.values > rm.loc[idx].values, "spike", "drop")
        return pd.DataFrame({
            "value":   anom.astype(np.float64),
            "z_score": z.loc[idx].astype(np.float64),
            "type":    pd.array(types, dtype=object),
        })

    @staticmethod
    def compute_statistics(df: "pd.DataFrame", roi_ci: Optional[Dict] = None) -> Dict:
        monthly = df.groupby("date")["sales"].sum()
        n_years = max(len(monthly) / 12.0, 1.0 / 12.0)
        total   = float(df["sales"].sum())
        first   = float(monthly.iloc[0])
        last    = float(monthly.iloc[-1])
        try:
            cagr = ((last / max(abs(first), 1e-10)) ** (1.0 / n_years) - 1.0) * 100.0
            if not math.isfinite(cagr):
                cagr = 0.0
        except Exception:
            cagr = 0.0

        def _best(col: str) -> str:
            if col not in df.columns or df[col].nunique() == 0:
                return "N/A"
            try:
                return str(df.groupby(col)["sales"].sum().idxmax())
            except Exception:
                return "N/A"

        returns_total = float(df["returns"].sum()) if "returns" in df.columns else 0.0
        avg_m  = float(monthly.mean())
        std_m  = float(monthly.std(ddof=1)) if len(monthly) > 1 else 0.0
        roi    = roi_ci or {"point": 0.0, "ci_lo": 0.0, "ci_hi": 0.0}
        peak_month   = monthly.idxmax()
        trough_month = monthly.idxmin()
        peak_str   = peak_month.isoformat()   if isinstance(peak_month,   pd.Timestamp) else str(peak_month)
        trough_str = trough_month.isoformat() if isinstance(trough_month, pd.Timestamp) else str(trough_month)
        return {
            "total_revenue":       total,
            "avg_monthly_revenue": avg_m,
            "peak_month":          peak_str,
            "peak_value":          float(monthly.max()),
            "trough_month":        trough_str,
            "trough_value":        float(monthly.min()),
            "std_dev":             std_m,
            "cv":                  (std_m / max(abs(avg_m), 1e-10) * 100.0),
            "cagr":                cagr,
            "best_product":        _best("product"),
            "best_region":         _best("region"),
            "total_returns":       returns_total,
            "return_rate":         (returns_total / max(abs(total), 1e-10) * 100.0),
            "marketing_roi":       roi["point"],
            "marketing_roi_ci_lo": roi["ci_lo"],
            "marketing_roi_ci_hi": roi["ci_hi"],
            "n_records":           len(df),
            "n_months":            len(monthly),
        }

    def analyze(self, df: "pd.DataFrame") -> SalesAnalysis:
        self._report("Aggregating monthly data …", 5)
        self._cancel.check()
        monthly = df.groupby("date")["sales"].sum().sort_index()
        y       = monthly.to_numpy(dtype=float)
        dates   = monthly.index

        cfg_period = self.config.seasonality_period
        if cfg_period <= 0:
            self._report("Detecting seasonality period …", 8)
            cfg_period = StatUtils.detect_seasonality(y)
        else:
            detected = StatUtils.detect_seasonality(y)
            if detected != cfg_period:
                self._report(f"Configured period={cfg_period}, auto-detected={detected}.", 8)

        self._report(f"Decomposing (period={cfg_period}) …", 15)
        self._cancel.check()
        trend, seasonal, residual = SalesAnalyzer.decompose(y, cfg_period)

        self._report("Detecting anomalies …", 20)
        anomalies = SalesAnalyzer.detect_anomalies(monthly)

        yoy      = monthly.pct_change(12) * 100.0
        mom      = monthly.pct_change()   * 100.0
        rolling3 = monthly.rolling(3, min_periods=1).mean()

        self._report("Computing statistics …", 25)
        roi_ci = None
        if "marketing_spend" in df.columns:
            mkt    = df.groupby("date")["marketing_spend"].sum().reindex(
                monthly.index, fill_value=0.0).to_numpy()
            roi_ci = StatUtils.marketing_roi_with_ci(
                y, mkt, n_boot=self.config.n_bootstrap)
        statistics = self.compute_statistics(df, roi_ci=roi_ci)

        self._report("Running forecasting models …", 30)
        engine = ForecastingEngine(self.config, self._cb, self._cancel)
        forecasts, ensemble, best_model = engine.forecast(y, cfg_period, dates=dates)

        self._report("Finalising …", 90)
        return SalesAnalysis(
            raw_data=df, trend=trend, seasonal=seasonal, residual=residual,
            anomalies=anomalies, yoy_growth=yoy, mom_growth=mom,
            rolling_avg=rolling3, statistics=statistics,
            forecasts=forecasts, best_model=best_model,
            ensemble_forecast=ensemble,
            detected_period=cfg_period,
        )


# ══════════════════════════════════════════════════════════════════════════════
#  CHART GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class ChartGenerator:
    PALETTE = Theme.CHART

    def __init__(self, config: ForecastConfig):
        self.config = config
        sns.set_theme(style="whitegrid", font_scale=0.95)
        plt.rcParams.update({
            "figure.facecolor":  "white",
            "axes.facecolor":    "white",
            "grid.alpha":        0.35,
            "axes.spines.top":   False,
            "axes.spines.right": False,
            "font.family":       "DejaVu Sans",
        })

    def _save(self, fig: "plt.Figure", name: str, dpi: int = 180) -> Path:
        path = self.config.charts_dir / f"{name}.png"
        try:
            self.config.charts_dir.mkdir(parents=True, exist_ok=True)
            fig.savefig(str(path), dpi=dpi, bbox_inches="tight",
                        facecolor="white", edgecolor="none")
        except Exception:
            logger.exception("Chart save failed for %s", name)
        finally:
            try:
                plt.close(fig)
            except Exception:
                pass
        return path

    @staticmethod
    def _money_fmt(x: float, _: Any) -> str:
        sym = "$"
        if abs(x) >= 1_000_000:
            return f"{sym}{x/1_000_000:.1f}M"
        if abs(x) >= 1_000:
            return f"{sym}{x/1_000:.0f}K"
        return f"{sym}{x:.0f}"

    def sales_overview(self, analysis: SalesAnalysis) -> Path:
        monthly = analysis.raw_data.groupby("date")["sales"].sum()
        fig, ax = plt.subplots(figsize=(12, 4.5))
        ax.fill_between(monthly.index, monthly.values, alpha=0.13, color=self.PALETTE[0])
        ax.plot(monthly.index, monthly.values, color=self.PALETTE[0], lw=2.5, label="Monthly Sales")
        ax.plot(monthly.index, analysis.rolling_avg.values,
                color=self.PALETTE[1], lw=2.0, ls="--", label="3-Month Rolling Avg")
        ax.plot(monthly.index, analysis.trend,
                color=self.PALETTE[2], lw=2.0, ls="-.", label="Trend")
        if not analysis.anomalies.empty:
            ax.scatter(analysis.anomalies.index, analysis.anomalies["value"],
                       c="red", zorder=5, s=80, marker="x", lw=2.5, label="Anomaly")
        ax.set_title("Historical Sales Overview", fontsize=14, fontweight="bold", pad=12)
        ax.set_ylabel(f"Revenue ({self.config.currency_symbol})", fontsize=11)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        ax.legend(frameon=True, fontsize=9, loc="upper left")
        ax.tick_params(axis="x", rotation=30)
        fig.tight_layout()
        return self._save(fig, "sales_overview")

    def forecast_chart(self, analysis: SalesAnalysis) -> Path:
        monthly      = analysis.raw_data.groupby("date")["sales"].sum()
        last_date    = monthly.index[-1]
        future_dates = pd.date_range(
            last_date + pd.DateOffset(months=1),
            periods=self.config.forecast_horizon, freq="MS")
        fig, ax = plt.subplots(figsize=(13, 5))
        ax.plot(monthly.index[-24:], monthly.values[-24:],
                color=self.PALETTE[0], lw=2.5, label="Historical", zorder=4)
        for i, (name, result) in enumerate(analysis.forecasts.items()):
            if result is None or not result.is_valid:
                continue
            n = min(len(result.predictions), len(future_dates))
            ax.plot(future_dates[:n], result.predictions[:n],
                    color=self.PALETTE[(i + 2) % len(self.PALETTE)],
                    lw=1.2, alpha=0.55, ls="--", label=name)
        ens = analysis.ensemble_forecast
        n   = min(len(ens.predictions), len(future_dates))
        ax.plot(future_dates[:n], ens.predictions[:n],
                color=self.PALETTE[1], lw=3, label="Ensemble Forecast", zorder=5)
        ax.fill_between(future_dates[:n], ens.lower_ci[:n], ens.upper_ci[:n],
                        alpha=0.20, color=self.PALETTE[1],
                        label=f"{int(self.config.confidence_level*100)}% CI")
        ax.axvline(x=last_date, color="gray", ls=":", lw=1.5, alpha=0.7)
        ax.set_title("Sales Forecast — All Models vs Ensemble", fontsize=14, fontweight="bold", pad=12)
        ax.set_ylabel(f"Revenue ({self.config.currency_symbol})", fontsize=11)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        ax.legend(frameon=True, fontsize=8, ncol=2, loc="upper left")
        ax.tick_params(axis="x", rotation=30)
        fig.tight_layout()
        return self._save(fig, "forecast")

    def decomposition_chart(self, analysis: SalesAnalysis) -> Path:
        monthly = analysis.raw_data.groupby("date")["sales"].sum()
        fig, axes = plt.subplots(4, 1, figsize=(12, 10), sharex=True)
        layers = [
            (monthly.values,    "Original",  self.PALETTE[0]),
            (analysis.trend,    "Trend",     self.PALETTE[1]),
            (analysis.seasonal, "Seasonal",  self.PALETTE[2]),
            (analysis.residual, "Residual",  self.PALETTE[3]),
        ]
        for ax, (data, title, color) in zip(axes, layers):
            ax.plot(monthly.index, data, color=color, lw=1.8)
            ax.fill_between(monthly.index, data, alpha=0.12, color=color)
            ax.set_ylabel(title, fontsize=10, fontweight="bold")
            ax.yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        axes[-1].tick_params(axis="x", rotation=30)
        fig.suptitle("Time Series Decomposition", fontsize=14, fontweight="bold", y=1.01)
        fig.tight_layout()
        return self._save(fig, "decomposition")

    def product_comparison(self, analysis: SalesAnalysis) -> Path:
        pm  = analysis.raw_data.groupby(["date", "product"])["sales"].sum().reset_index()
        fig, axes = plt.subplots(1, 2, figsize=(13, 5))
        for i, prod in enumerate(pm["product"].unique()):
            sub = pm[pm["product"] == prod]
            axes[0].plot(sub["date"], sub["sales"],
                         color=self.PALETTE[i % len(self.PALETTE)], lw=1.8, label=prod)
        axes[0].set_title("Revenue by Product Over Time", fontsize=12, fontweight="bold")
        axes[0].set_ylabel(f"Revenue ({self.config.currency_symbol})")
        axes[0].yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        axes[0].legend(fontsize=8)
        axes[0].tick_params(axis="x", rotation=30)
        totals = analysis.raw_data.groupby("product")["sales"].sum().sort_values(ascending=False)
        bars   = axes[1].bar(totals.index, totals.values, color=self.PALETTE[:len(totals)])
        axes[1].set_title("Total Revenue by Product", fontsize=12, fontweight="bold")
        axes[1].set_ylabel(f"Total Revenue ({self.config.currency_symbol})")
        axes[1].yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        for bar in bars:
            h = bar.get_height()
            axes[1].text(bar.get_x() + bar.get_width() / 2, h * 1.01,
                          self._money_fmt(h, None),
                          ha="center", va="bottom", fontsize=8.5, fontweight="bold")
        fig.tight_layout()
        return self._save(fig, "product_comparison")

    def model_comparison(self, analysis: SalesAnalysis) -> Path:
        valid = {k: v for k, v in analysis.forecasts.items() if v is not None and v.is_valid}
        valid["Ensemble"] = analysis.ensemble_forecast
        names = list(valid.keys())
        x     = np.arange(len(names))
        fig, axes = plt.subplots(1, 4, figsize=(17, 4.5))
        metrics = [
            ([v.smape for v in valid.values()], "sMAPE (%)\n(lower=better)",  self.PALETTE[2]),
            ([v.mase  for v in valid.values()], "MASE\n(lower=better)",       self.PALETTE[3]),
            ([v.rmse  for v in valid.values()], "RMSE ($)\n(lower=better)",   self.PALETTE[0]),
            ([v.r2    for v in valid.values()], "R² Score\n(higher=better)",  self.PALETTE[1]),
        ]
        for ax, (vals, title, color) in zip(axes, metrics):
            bars = ax.bar(x, vals, color=color, alpha=0.85, width=0.6)
            ax.set_xticks(x)
            ax.set_xticklabels(names, rotation=30, ha="right", fontsize=8.5)
            ax.set_title(title, fontsize=10, fontweight="bold")
            for bar in bars:
                h = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, h * 1.02,
                        f"{h:.2f}", ha="center", va="bottom", fontsize=7.5)
        fig.suptitle("Model Performance Comparison", fontsize=13, fontweight="bold", y=1.02)
        fig.tight_layout()
        return self._save(fig, "model_comparison")

    def yoy_growth_chart(self, analysis: SalesAnalysis) -> Path:
        yoy = analysis.yoy_growth.dropna()
        fig, ax = plt.subplots(figsize=(12, 4))
        bar_colors = [self.PALETTE[3] if v >= 0 else self.PALETTE[4] for v in yoy.values]
        ax.bar(yoy.index, yoy.values, color=bar_colors, width=20, alpha=0.85)
        ax.axhline(0, color="black", lw=0.8)
        ax.set_title("Year-over-Year Revenue Growth (%)", fontsize=13, fontweight="bold", pad=10)
        ax.set_ylabel("YoY Growth (%)")
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.1f}%"))
        ax.tick_params(axis="x", rotation=30)
        fig.tight_layout()
        return self._save(fig, "yoy_growth")

    def heatmap_chart(self, analysis: SalesAnalysis) -> Path:
        df          = analysis.raw_data.copy()
        df["month"] = df["date"].dt.month
        df["year"]  = df["date"].dt.year
        pivot       = df.pivot_table(values="sales", index="year", columns="month", aggfunc="sum")
        month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                       "Jul","Aug","Sep","Oct","Nov","Dec"]
        pivot.columns = [month_names[int(m) - 1] for m in pivot.columns]
        fig, ax = plt.subplots(figsize=(13, 4.5))
        sns.heatmap(pivot / 1e6, annot=True, fmt=".1f", cmap="Blues",
                    linewidths=0.5, ax=ax, cbar_kws={"label": "Revenue ($M)"})
        ax.set_title("Monthly Revenue Heatmap by Year ($M)", fontsize=13, fontweight="bold", pad=10)
        ax.set_xlabel("Month")
        ax.set_ylabel("Year")
        fig.tight_layout()
        return self._save(fig, "heatmap")

    def residual_diagnostics_chart(self, analysis: SalesAnalysis) -> Path:
        monthly = analysis.raw_data.groupby("date")["sales"].sum()
        resid   = analysis.residual.copy()
        n       = len(resid)
        fig, axes = plt.subplots(2, 2, figsize=(12, 8))
        fig.suptitle(f"Residual Diagnostics — {analysis.best_model}", fontsize=13, fontweight="bold")
        axes[0, 0].plot(monthly.index, resid, color=self.PALETTE[0], lw=1.5)
        axes[0, 0].axhline(0, ls="--", color="gray", lw=1)
        axes[0, 0].set_title("Residuals vs Time")
        axes[0, 0].yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        max_lag  = max(1, min(24, n // 2 - 2))
        acf_vals = []
        for k in range(1, max_lag + 1):
            if n - k < 2:
                break
            r = np.corrcoef(resid[:n - k], resid[k:])
            acf_vals.append(float(r[0, 1]) if not np.any(np.isnan(r)) else 0.0)
        if acf_vals:
            conf = 1.96 / math.sqrt(max(n, 2))
            axes[0, 1].bar(range(1, len(acf_vals) + 1), acf_vals, color=self.PALETTE[1], alpha=0.8)
            axes[0, 1].axhline( conf, ls="--", color="red", lw=1, label="95% band")
            axes[0, 1].axhline(-conf, ls="--", color="red", lw=1)
            axes[0, 1].set_title("ACF of Residuals")
            axes[0, 1].set_xlabel("Lag")
            axes[0, 1].legend(fontsize=8)
        try:
            stats.probplot(resid, dist="norm", plot=axes[1, 0])
            axes[1, 0].set_title("Normal Q-Q Plot")
        except Exception:
            axes[1, 0].set_title("Normal Q-Q Plot (unavailable)")
        axes[1, 1].hist(resid, bins=min(20, max(5, n // 3)),
                         color=self.PALETTE[0], alpha=0.7, edgecolor="white")
        mu, sigma = float(np.mean(resid)), float(np.std(resid))
        if sigma > 1e-10 and n > 5:
            r_range = float(resid.max() - resid.min())
            bin_w   = r_range / max(min(20, n // 3), 1)
            x       = np.linspace(mu - 4 * sigma, mu + 4 * sigma, 300)
            axes[1, 1].plot(x, stats.norm.pdf(x, mu, sigma) * n * bin_w,
                             color=self.PALETTE[2], lw=2, label="Normal fit")
            axes[1, 1].legend(fontsize=8)
        axes[1, 1].set_title("Residual Distribution")
        fig.tight_layout()
        return self._save(fig, "residual_diagnostics")

    def generate_all(
        self, analysis: SalesAnalysis,
        progress_cb: Optional[Callable] = None,
    ) -> Dict[str, Path]:
        cb    = progress_cb or (lambda msg, pct: None)
        steps = [
            ("overview",         self.sales_overview),
            ("forecast",         self.forecast_chart),
            ("decomposition",    self.decomposition_chart),
            ("products",         self.product_comparison),
            ("model_comparison", self.model_comparison),
            ("yoy_growth",       self.yoy_growth_chart),
            ("heatmap",          self.heatmap_chart),
            ("diagnostics",      self.residual_diagnostics_chart),
        ]
        charts: Dict[str, Path] = {}
        for idx, (name, fn) in enumerate(steps):
            cb(f"Rendering chart: {name.replace('_', ' ').title()} …",
               91 + int(idx / len(steps) * 5))
            try:
                charts[name] = fn(analysis)
            except Exception as exc:
                cb(f"Chart '{name}' skipped: {exc}", None)
                logger.exception("Chart %s failed", name)
        return charts


# ══════════════════════════════════════════════════════════════════════════════
#  PDF REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class PDFReportGenerator:
    def __init__(self, config: ForecastConfig):
        self.config = config
        self.styles = getSampleStyleSheet()
        self._setup_styles()

    @staticmethod
    def _rl_color(rgb: Tuple) -> "colors.Color":
        return colors.Color(*[float(c) for c in rgb[:3]])

    def _setup_styles(self) -> None:
        primary = self._rl_color(self.config.primary_color)
        accent  = self._rl_color(self.config.accent_color)

        def _add(name: str, **kw: Any) -> None:
            if name not in self.styles:
                self.styles.add(ParagraphStyle(name, **kw))

        _add("ReportTitle", parent=self.styles["Title"],    fontSize=24, textColor=primary,
             spaceAfter=6, fontName="Helvetica-Bold", alignment=TA_CENTER)
        _add("ReportSub",   parent=self.styles["Normal"],   fontSize=11, textColor=accent,
             spaceAfter=18, alignment=TA_CENTER)
        _add("SecHead",     parent=self.styles["Heading1"], fontSize=14, textColor=primary,
             spaceBefore=18, spaceAfter=8, fontName="Helvetica-Bold")
        _add("SubHead",     parent=self.styles["Heading2"], fontSize=11, textColor=accent,
             spaceBefore=10, spaceAfter=6, fontName="Helvetica-Bold")
        _add("Body",        parent=self.styles["Normal"],   fontSize=9,  leading=13,
             spaceAfter=6, alignment=TA_JUSTIFY)
        _add("Bullet",      parent=self.styles["Normal"],   fontSize=9,  leading=13,
             leftIndent=14, bulletIndent=0, spaceBefore=2)
        _add("Small",       parent=self.styles["Normal"],   fontSize=8,  leading=11,
             textColor=colors.HexColor("#555555"))
        _add("Footer",      parent=self.styles["Normal"],   fontSize=7,  alignment=TA_CENTER,
             textColor=colors.HexColor("#888888"))

    def _hdr_footer(self, canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColorRGB(0.5, 0.5, 0.5)
        canvas.drawString(inch, 0.5 * inch,
                           f"{self.config.company_name} — Confidential")
        canvas.drawRightString(
            A4[0] - inch, 0.5 * inch,
            f"Page {doc.page}  ·  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.setStrokeColorRGB(0.12, 0.29, 0.49)
        canvas.setLineWidth(0.5)
        canvas.line(inch, 0.65 * inch, A4[0] - inch, 0.65 * inch)
        canvas.restoreState()

    @staticmethod
    def _tbl_style(
        header_color: Tuple = (0.12, 0.29, 0.49),
        alt_color:    Tuple = (0.94, 0.97, 0.99),
        n_rows: int = 0,
    ) -> TableStyle:
        cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0), colors.Color(*header_color)),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 9),
            ("FONTSIZE",      (0, 1), (-1, -1), 8),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUND", (0, 1), (-1, -1), [colors.white, colors.Color(*alt_color)]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ]
        return TableStyle(cmds)

    def _chart_img(self, path: Path, width: float = 6.2 * inch) -> Optional["RLImage"]:
        if path and path.exists():
            try:
                img    = RLImage(str(path), width=width)
                img.hAlign = "CENTER"
                return img
            except Exception:
                pass
        return None

    def _metrics_table(self, analysis: SalesAnalysis) -> "Table":
        sym   = self.config.currency_symbol
        valid = {k: v for k, v in analysis.forecasts.items() if v is not None and v.is_valid}
        valid["Ensemble ★"] = analysis.ensemble_forecast
        headers = ["Model", "sMAPE (%)", "MASE", f"RMSE ({sym})", "R²", "CV sMAPE", "Time (s)"]
        rows    = [headers]
        for name, res in valid.items():
            cv_str = (f"{res.cv_smape:.2f}"
                      if res.cv_smape is not None and math.isfinite(res.cv_smape) else "—")
            rows.append([
                name,
                f"{res.smape:.2f}", f"{res.mase:.3f}",
                f"{sym}{res.rmse:,.0f}", f"{res.r2:.3f}",
                cv_str, f"{res.train_time_s:.2f}",
            ])
        col_widths = [1.9*inch, 0.8*inch, 0.7*inch, 1.0*inch, 0.7*inch, 0.8*inch, 0.7*inch]
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(self._tbl_style(n_rows=len(rows)))
        return tbl

    def generate(
        self, analysis: SalesAnalysis, charts: Dict[str, Path], output_path: Path,
    ) -> Path:
        pdf_path = output_path.with_suffix(".pdf")
        doc      = SimpleDocTemplate(
            str(pdf_path), pagesize=A4,
            leftMargin=inch, rightMargin=inch,
            topMargin=0.8*inch, bottomMargin=0.8*inch,
        )
        cfg = self.config
        st  = analysis.statistics
        ens = analysis.ensemble_forecast
        S   = self.styles
        story: List = []

        # ── Cover ────────────────────────────────────────────────────────────
        story.extend([
            Spacer(1, 1.2 * inch),
            Paragraph(cfg.report_title, S["ReportTitle"]),
            Paragraph(cfg.company_name, S["ReportSub"]),
            Spacer(1, 0.3 * inch),
            HRFlowable(width="100%", thickness=1.5,
                       color=self._rl_color(cfg.primary_color)),
            Spacer(1, 0.2 * inch),
            Paragraph(
                f"Prepared by: {cfg.analyst_name}<br/>"
                f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}<br/>"
                f"Data records: {st['n_records']:,}  ·  Months: {st['n_months']}  ·  "
                f"Forecast horizon: {cfg.forecast_horizon} months",
                S["Small"]),
            PageBreak(),
        ])

        # ── Executive Summary ─────────────────────────────────────────────────
        story.append(Paragraph("1. Executive Summary", S["SecHead"]))
        sym  = cfg.currency_symbol
        roi_str = (f"{st['marketing_roi']:.2f}x  "
                   f"[{st['marketing_roi_ci_lo']:.2f}x – {st['marketing_roi_ci_hi']:.2f}x]")
        kpis = [
            ("Total Revenue",              f"{sym}{st['total_revenue']:,.2f}"),
            ("Average Monthly Revenue",    f"{sym}{st['avg_monthly_revenue']:,.2f}"),
            ("Peak Month / Revenue",       f"{_fmt_month(st['peak_month'])}  /  {sym}{st['peak_value']:,.2f}"),
            ("Trough Month / Revenue",     f"{_fmt_month(st['trough_month'])}  /  {sym}{st['trough_value']:,.2f}"),
            ("Compound Annual Growth",     f"{st['cagr']:.2f}%"),
            ("Revenue Volatility (CV)",    f"{st['cv']:.2f}%"),
            ("Best Product",               str(st["best_product"])),
            ("Best Region",                str(st["best_region"])),
            ("Marketing ROI (95% CI)",     roi_str),
            ("Anomalies Detected",         str(len(analysis.anomalies))),
            ("Best Forecast Model",        analysis.best_model),
            ("Ensemble sMAPE / MASE",      f"{ens.smape:.2f}%  /  {ens.mase:.3f}"),
            ("Total Forecast Revenue",     f"{sym}{ens.predictions.sum():,.0f}"),
        ]
        kpi_rows = [[Paragraph(k, S["Body"]), Paragraph(v, S["Body"])] for k, v in kpis]
        kpi_tbl  = Table(kpi_rows, colWidths=[2.5*inch, 4.0*inch])
        kpi_tbl.setStyle(self._tbl_style(n_rows=len(kpi_rows)))
        story.extend([kpi_tbl, Spacer(1, 0.3*inch)])

        # ── Optional model availability callout ──────────────────────────────
        optionals = []
        if _XGBOOST_AVAILABLE:
            optionals.append("XGBoost ✓")
        if _LIGHTGBM_AVAILABLE:
            optionals.append("LightGBM ✓")
        if _PROPHET_AVAILABLE:
            optionals.append("Prophet ✓")
        if optionals:
            story.append(Paragraph(
                f"<b>v2 Advanced Models available:</b> {', '.join(optionals)}", S["Small"]))
            story.append(Spacer(1, 0.15*inch))

        # ── Sales Overview ────────────────────────────────────────────────────
        story.append(Paragraph("2. Historical Sales Overview", S["SecHead"]))
        img = self._chart_img(charts.get("overview"))
        if img:
            story.extend([img, Spacer(1, 0.2*inch)])

        # ── Time Series Decomposition ─────────────────────────────────────────
        story.append(Paragraph("3. Time Series Decomposition", S["SecHead"]))
        img = self._chart_img(charts.get("decomposition"))
        if img:
            story.extend([img, Spacer(1, 0.2*inch)])

        # ── Forecasting ───────────────────────────────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("4. Sales Forecast", S["SecHead"]))
        img = self._chart_img(charts.get("forecast"))
        if img:
            story.extend([img, Spacer(1, 0.2*inch)])

        # ── Model Performance Table ───────────────────────────────────────────
        story.append(Paragraph("5. Model Performance", S["SecHead"]))
        story.extend([self._metrics_table(analysis), Spacer(1, 0.2*inch)])

        img = self._chart_img(charts.get("model_comparison"), width=6.5*inch)
        if img:
            story.extend([img, Spacer(1, 0.2*inch)])

        # ── Product Analysis ─────────────────────────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("6. Product Analysis", S["SecHead"]))
        img = self._chart_img(charts.get("products"))
        if img:
            story.extend([img, Spacer(1, 0.2*inch)])

        # ── Growth & Seasonality ─────────────────────────────────────────────
        story.append(Paragraph("7. Growth & Seasonality", S["SecHead"]))
        for key in ("yoy_growth", "heatmap"):
            img = self._chart_img(charts.get(key))
            if img:
                story.extend([img, Spacer(1, 0.2*inch)])

        # ── Residual Diagnostics ─────────────────────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("8. Residual Diagnostics", S["SecHead"]))
        img = self._chart_img(charts.get("diagnostics"))
        if img:
            story.extend([img, Spacer(1, 0.2*inch)])

        # ── Feature Importance ───────────────────────────────────────────────
        feat_models = {
            k: v for k, v in analysis.forecasts.items()
            if v and v.feature_importance
        }
        if feat_models:
            story.append(Paragraph("9. Feature Importance (ML Models)", S["SecHead"]))
            for name, res in feat_models.items():
                story.append(Paragraph(name, S["SubHead"]))
                fi_rows = [["Feature", "Importance"]]
                for feat, imp in list(res.feature_importance.items())[:10]:
                    fi_rows.append([feat, f"{imp:.4f}"])
                fi_tbl = Table(fi_rows, colWidths=[3.5*inch, 1.5*inch])
                fi_tbl.setStyle(self._tbl_style(n_rows=len(fi_rows)))
                story.extend([fi_tbl, Spacer(1, 0.15*inch)])

        doc.build(story, onFirstPage=self._hdr_footer, onLaterPages=self._hdr_footer)
        return pdf_path


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class ExcelReportGenerator:
    _HEADER_FILL   = "1E4D7B"
    _ALT_FILL      = "F0F5FC"
    _BEST_FILL     = "D4EFDF"
    _ENS_FILL      = "EBF5FB"
    _TITLE_FONT_SZ = 15
    _HEAD_FONT_SZ  = 10
    _BODY_FONT_SZ  = 9

    def __init__(self, config: ForecastConfig):
        self.config = config

    @staticmethod
    def _fill(hex_: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_.lstrip("#"))

    @staticmethod
    def _font(bold: bool = False, size: int = 10,
              color: str = "000000", italic: bool = False) -> Font:
        return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

    @staticmethod
    def _border(color: str = "D0D8E0") -> Border:
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _align(h: str = "center", wrap: bool = True) -> Alignment:
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def _header_cell(self, ws: Any, row: int, col: int,
                     value: Any, fill: str = "") -> None:
        c           = ws.cell(row=row, column=col, value=value)
        c.font      = self._font(bold=True, size=self._HEAD_FONT_SZ, color="FFFFFF")
        c.fill      = self._fill(fill or self._HEADER_FILL)
        c.alignment = self._align()
        c.border    = self._border("4A90D9")

    def _data_cell(
        self, ws: Any, row: int, col: int,
        value: Any, shade: bool = False, bold: bool = False,
        color: str = "000000", h_align: str = "center",
        number_format: str = "General",
    ) -> None:
        c               = ws.cell(row=row, column=col, value=value)
        c.font          = self._font(bold=bold, size=self._BODY_FONT_SZ, color=color)
        c.fill          = self._fill(self._ALT_FILL if shade else "FFFFFF")
        c.alignment     = self._align(h_align)
        c.border        = self._border()
        c.number_format = number_format

    def _write_header_row(self, ws: Any, row: int, headers: List[str], col_start: int = 1) -> None:
        for j, h in enumerate(headers, col_start):
            self._header_cell(ws, row, j, h)

    def _write_data_row(
        self, ws: Any, row: int, values: List[Any],
        col_start: int = 1, shade: bool = False,
        bold: bool = False, color: str = "000000",
    ) -> None:
        for j, v in enumerate(values, col_start):
            self._data_cell(ws, row, j, v, shade=shade, bold=bold, color=color)

    def _write_title_block(self, ws: Any, title: str, subtitle: str = "") -> int:
        last_col = get_column_letter(max(ws.max_column or 1, 10))
        ws.merge_cells(f"A1:{last_col}1")
        c           = ws["A1"]
        c.value     = title
        c.font      = Font(name="Calibri", bold=True, size=self._TITLE_FONT_SZ, color="FFFFFF")
        c.fill      = self._fill(self._HEADER_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36
        if subtitle:
            ws.merge_cells(f"A2:{last_col}2")
            c2           = ws["A2"]
            c2.value     = subtitle
            c2.font      = Font(name="Calibri", italic=True, size=9, color="444444")
            c2.fill      = self._fill("EEF3FA")
            c2.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 18
            return 4
        return 3

    def _autofit_columns(self, ws: Any, min_w: int = 8, max_w: int = 35) -> None:
        for col in ws.columns:
            try:
                best   = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                letter = get_column_letter(col[0].column)
                ws.column_dimensions[letter].width = min(max(best + 2, min_w), max_w)
            except Exception:
                pass

    def _add_xl_table(self, ws: Any, start_row: int, end_row: int,
                      start_col: int, end_col: int, name: str) -> None:
        try:
            ref   = (f"{get_column_letter(start_col)}{start_row}:"
                     f"{get_column_letter(end_col)}{end_row}")
            tbl   = XLTable(displayName=name, ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False)
            tbl.tableStyleInfo = style
            ws.add_table(tbl)
        except Exception as exc:
            logger.warning("Could not add Excel table '%s': %s", name, exc)

    def _sheet_summary(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws    = wb.create_sheet("📊 Summary")
        start = self._write_title_block(
            ws, "Executive Summary",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {self.config.company_name}")
        st  = analysis.statistics
        ens = analysis.ensemble_forecast
        sym = self.config.currency_symbol
        roi_ci = (f"{st['marketing_roi']:.2f}x  "
                  f"[{st['marketing_roi_ci_lo']:.2f}x – {st['marketing_roi_ci_hi']:.2f}x]")
        optionals = []
        if _XGBOOST_AVAILABLE:  optionals.append("XGBoost")
        if _LIGHTGBM_AVAILABLE: optionals.append("LightGBM")
        if _PROPHET_AVAILABLE:  optionals.append("Prophet")
        kpis = [
            ("Total Revenue",              f"{sym}{st['total_revenue']:,.2f}"),
            ("Avg Monthly Revenue",        f"{sym}{st['avg_monthly_revenue']:,.2f}"),
            ("Peak Month",                 _fmt_month(st["peak_month"])),
            ("Peak Revenue",               f"{sym}{st['peak_value']:,.2f}"),
            ("Trough Month",               _fmt_month(st["trough_month"])),
            ("Trough Revenue",             f"{sym}{st['trough_value']:,.2f}"),
            ("CAGR",                       f"{st['cagr']:.2f}%"),
            ("Revenue Volatility (CV)",    f"{st['cv']:.2f}%"),
            ("Best Product",               str(st["best_product"])),
            ("Best Region",                str(st["best_region"])),
            ("Total Returns",              f"{sym}{st['total_returns']:,.2f}"),
            ("Return Rate",                f"{st['return_rate']:.3f}%"),
            ("Marketing ROI (95% CI)",     roi_ci),
            ("Records in Dataset",         f"{st['n_records']:,}"),
            ("Months of History",          f"{st['n_months']}"),
            ("Detected Seasonal Period",   f"{analysis.detected_period} months"),
            ("Forecast Horizon",           f"{self.config.forecast_horizon} months"),
            ("Ensemble sMAPE",             f"{ens.smape:.2f}%"),
            ("Ensemble MASE",              f"{ens.mase:.3f}"),
            ("Best Single Model",          analysis.best_model),
            ("Total Forecast Revenue",     f"{sym}{ens.predictions.sum():,.0f}"),
            ("v2 Advanced Models",         ", ".join(optionals) if optionals else "None installed"),
            ("Config File",                str(_LOCAL_CONFIG if _LOCAL_CONFIG.exists() else _CONFIG_FILE)),
        ]
        self._write_header_row(ws, start, ["KPI Metric", "Value"], col_start=1)
        r = start + 1
        for i, (k, v) in enumerate(kpis):
            self._write_data_row(ws, r, [k, v], shade=(i % 2 == 0))
            ws.cell(row=r, column=1).alignment = self._align("left")
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, 2, "SummaryKPIs")
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 52
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_forecast(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws           = wb.create_sheet("🔮 Forecast")
        monthly      = analysis.raw_data.groupby("date")["sales"].sum()
        last_date    = monthly.index[-1]
        future_dates = pd.date_range(
            last_date + pd.DateOffset(months=1),
            periods=self.config.forecast_horizon, freq="MS")
        ens   = analysis.ensemble_forecast
        start = self._write_title_block(
            ws, "Ensemble Sales Forecast",
            f"Horizon: {self.config.forecast_horizon} months  |  "
            f"Best model: {analysis.best_model}  |  "
            f"Ensemble sMAPE: {ens.smape:.2f}%  |  MASE: {ens.mase:.3f}")
        headers = ["Month", "Point Forecast ($)", "Lower CI ($)", "Upper CI ($)",
                   "CI Width ($)", "MoM Change (%)", "Cumulative ($)"]
        self._write_header_row(ws, start, headers)
        r, prev, cum = start + 1, float(monthly.iloc[-1]), 0.0
        for i, (dt, fc, lo, hi) in enumerate(
            zip(future_dates, ens.predictions, ens.lower_ci, ens.upper_ci)
        ):
            fc_f, lo_f, hi_f = float(fc), float(lo), float(hi)
            mom_pct = (fc_f - prev) / max(abs(prev), 1e-10) * 100.0
            cum    += fc_f
            shade   = (i % 2 == 0)
            vals    = [dt.strftime("%b %Y"),
                       round(fc_f, 2), round(lo_f, 2), round(hi_f, 2),
                       round(hi_f - lo_f, 2), round(mom_pct, 2), round(cum, 2)]
            self._write_data_row(ws, r, vals, shade=shade)
            mom_cell = ws.cell(row=r, column=6)
            mom_cell.number_format = '+0.00;-0.00;0.00'
            mom_cell.font = self._font(
                bold=True, size=self._BODY_FONT_SZ,
                color="1A6B1A" if mom_pct >= 0 else "C51A1A")
            for col_idx in [2, 3, 4, 5, 7]:
                ws.cell(row=r, column=col_idx).number_format = '#,##0.00'
            prev = fc_f
            r   += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "ForecastTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_models(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws    = wb.create_sheet("🤖 Models")
        valid = {k: v for k, v in analysis.forecasts.items() if v is not None}
        valid["Ensemble"] = analysis.ensemble_forecast
        start = self._write_title_block(
            ws, "Forecasting Model Performance",
            "sMAPE = symmetric MAPE  |  MASE = Mean Absolute Scaled Error  |  "
            "CV sMAPE = walk-forward cross-validation")
        headers = ["Model", "sMAPE (%)", "MASE", "RMSE ($)", "R²",
                   "MAE ($)", "Accuracy (%)", "CV sMAPE (%)", "Train (s)", "AIC", "Diagnostics"]
        self._write_header_row(ws, start, headers)
        r          = start + 1
        best_smape = min(v.smape for v in valid.values())
        for i, (name, res) in enumerate(valid.items()):
            is_best = abs(res.smape - best_smape) < 1e-9
            is_ens  = "Ensemble" in name
            shade   = (i % 2 == 0)
            cv_str  = (f"{res.cv_smape:.2f}"
                       if res.cv_smape is not None and math.isfinite(res.cv_smape) else "—")
            aic_str  = f"{res.aic:.1f}" if res.aic is not None else "—"
            diag_str = res.diagnostics.summary() if res.diagnostics else "—"
            vals = [name, round(res.smape, 3), round(res.mase, 4),
                    round(res.rmse, 2), round(res.r2, 4), round(res.mae, 2),
                    round(res.accuracy_score, 2), cv_str,
                    round(res.train_time_s, 2), aic_str, diag_str]
            self._write_data_row(ws, r, vals, shade=shade)
            fill_hex = (self._BEST_FILL if is_best else self._ENS_FILL if is_ens else None)
            if fill_hex:
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c_idx).fill = self._fill(fill_hex)
                    if is_best:
                        ws.cell(row=r, column=c_idx).font = self._font(
                            bold=True, size=self._BODY_FONT_SZ, color="1A6B1A")
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "ModelMetrics")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_history(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws      = wb.create_sheet("📈 History")
        monthly = analysis.raw_data.groupby("date")["sales"].sum().reset_index()
        monthly.columns = ["Date", "Total Sales ($)"]
        monthly["3M Rolling Avg ($)"]  = monthly["Total Sales ($)"].rolling(3).mean().round(2)
        monthly["12M Rolling Avg ($)"] = monthly["Total Sales ($)"].rolling(12).mean().round(2)
        monthly["YoY Growth (%)"]      = monthly["Total Sales ($)"].pct_change(12).mul(100).round(2)
        monthly["MoM Growth (%)"]      = monthly["Total Sales ($)"].pct_change().mul(100).round(2)
        monthly["Date"]                = monthly["Date"].dt.strftime("%Y-%m")
        start   = self._write_title_block(ws, "Historical Monthly Sales", f"{len(monthly)} months")
        headers = monthly.columns.tolist()
        self._write_header_row(ws, start, headers)
        r = start + 1
        for i, row in enumerate(monthly.itertuples(index=False)):
            shade = (i % 2 == 0)
            vals  = list(row)
            self._write_data_row(ws, r, vals, shade=shade)
            for col_idx, v in [(5, vals[4]), (6, vals[5])]:
                if v is not None and not (isinstance(v, float) and math.isnan(v)):
                    cell      = ws.cell(row=r, column=col_idx)
                    cell.font = self._font(size=self._BODY_FONT_SZ,
                                           color="1A6B1A" if float(v) >= 0 else "C51A1A")
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "HistoryTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"
        for row in ws.iter_rows(min_row=start + 1, max_row=r - 1, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = "#,##0.00"

    def _sheet_products(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws   = wb.create_sheet("🏷️ Products")
        prod = (analysis.raw_data.groupby("product")
                .agg(total=("sales", "sum"), avg=("sales", "mean"),
                     units=("units", "sum"), returns=("returns", "sum"))
                .reset_index().sort_values("total", ascending=False))
        prod["return_rate_%"]   = (prod["returns"] / prod["total"].replace(0, 1) * 100).round(3)
        prod["revenue_share_%"] = (prod["total"] / prod["total"].sum() * 100).round(2)
        start   = self._write_title_block(ws, "Product-Level Analysis", "Sorted by total revenue")
        headers = ["Product", "Total Revenue ($)", "Avg Monthly ($)", "Units Sold",
                   "Returns ($)", "Return Rate (%)", "Revenue Share (%)"]
        self._write_header_row(ws, start, headers)
        r = start + 1
        for i, row in prod.iterrows():
            vals = [str(row["product"]), round(float(row["total"]), 2),
                    round(float(row["avg"]), 2), int(row["units"]),
                    round(float(row["returns"]), 2),
                    round(float(row["return_rate_%"]), 3),
                    round(float(row["revenue_share_%"]), 2)]
            self._write_data_row(ws, r, vals, shade=(i % 2 == 0))
            ws.cell(row=r, column=1).alignment = self._align("left")
            for col_idx in [2, 3, 5]:
                ws.cell(row=r, column=col_idx).number_format = "#,##0.00"
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "ProductTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_decomp(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws      = wb.create_sheet("🔬 Decomposition")
        monthly = analysis.raw_data.groupby("date")["sales"].sum().reset_index()
        start   = self._write_title_block(
            ws, "Time Series Decomposition",
            f"Seasonal period: {analysis.detected_period} months  |  "
            "Additive model: Original = Trend + Seasonal + Residual")
        headers = ["Date", "Original ($)", "Trend ($)", "Seasonal ($)", "Residual ($)"]
        self._write_header_row(ws, start, headers)
        dates = monthly["date"].dt.strftime("%Y-%m").tolist()
        orig  = monthly["sales"].tolist()
        r     = start + 1
        for i, (d, o, tr, se, re) in enumerate(
            zip(dates, orig, analysis.trend, analysis.seasonal, analysis.residual)
        ):
            shade = (i % 2 == 0)
            self._write_data_row(ws, r, [d, round(o, 2), round(float(tr), 2),
                                         round(float(se), 2), round(float(re), 2)], shade=shade)
            for col_idx in [2, 3, 4, 5]:
                ws.cell(row=r, column=col_idx).number_format = "#,##0.00"
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "DecompTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_anomalies(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws    = wb.create_sheet("⚠️ Anomalies")
        start = self._write_title_block(
            ws, "Detected Sales Anomalies",
            "Rolling Z-score method  |  threshold = 2.5σ")
        if analysis.anomalies.empty:
            ws.cell(row=start, column=1, value="No anomalies detected in the dataset.")
            return
        headers = ["Date", "Sales Value ($)", "Z-Score", "Type"]
        self._write_header_row(ws, start, headers)
        r = start + 1
        for i, (dt, row) in enumerate(analysis.anomalies.iterrows()):
            shade = (i % 2 == 0)
            self._write_data_row(ws, r, [
                str(dt)[:10] if hasattr(dt, '__str__') else str(dt),
                round(float(row["value"]), 2),
                round(float(row["z_score"]), 3),
                str(row["type"]),
            ], shade=shade)
            ws.cell(row=r, column=2).number_format = "#,##0.00"
            type_cell = ws.cell(row=r, column=4)
            type_cell.font = self._font(
                size=self._BODY_FONT_SZ,
                color="C51A1A" if str(row["type"]) == "spike" else "1A1AC5")
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "AnomalyTable")
        self._autofit_columns(ws)

    def _sheet_config(self, wb: Any) -> None:
        ws    = wb.create_sheet("⚙️ Config")
        start = self._write_title_block(
            ws, "Run Configuration",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        cfg = self.config
        rows: List[Tuple[str, str]] = [
            ("Company Name",             cfg.company_name),
            ("Analyst Name",             cfg.analyst_name),
            ("Currency Symbol",          cfg.currency_symbol),
            ("Forecast Horizon",         f"{cfg.forecast_horizon} months"),
            ("Confidence Level",         f"{cfg.confidence_level:.0%}"),
            ("Seasonality Period",       f"{cfg.seasonality_period} (0=auto)"),
            ("Bootstrap Samples",        str(cfg.n_bootstrap)),
            ("CV Folds",                 str(cfg.cv_folds)),
            ("ARIMA",                    "✓" if cfg.use_arima else "✗"),
            ("Holt-Winters",             "✓" if cfg.use_exp_smoothing else "✗"),
            ("Random Forest",            "✓" if cfg.use_random_forest else "✗"),
            ("Gradient Boosting",        "✓" if cfg.use_gradient_boosting else "✗"),
            ("Ridge Regression",         "✓" if cfg.use_ridge else "✗"),
            ("XGBoost",                  f"{'✓' if cfg.use_xgboost else '✗'} (installed: {_XGBOOST_AVAILABLE})"),
            ("LightGBM",                 f"{'✓' if cfg.use_lightgbm else '✗'} (installed: {_LIGHTGBM_AVAILABLE})"),
            ("Prophet",                  f"{'✓' if cfg.use_prophet else '✗'} (installed: {_PROPHET_AVAILABLE})"),
            ("Prophet Changepoint Scale",str(cfg.prophet_changepoint_scale)),
            ("Prophet Seasonality Mode", cfg.prophet_seasonality_mode),
            ("Report Formats",           ", ".join(cfg.report_format)),
            ("Output Path",              str(cfg.output_path)),
        ]
        self._write_header_row(ws, start, ["Setting", "Value"])
        r = start + 1
        for i, (k, v) in enumerate(rows):
            self._write_data_row(ws, r, [k, v], shade=(i % 2 == 0))
            ws.cell(row=r, column=1).alignment = self._align("left")
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, 2, "ConfigTable")
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 45

    def _sheet_feature_importance(self, wb: Any, analysis: SalesAnalysis) -> None:
        feat_models = {k: v for k, v in analysis.forecasts.items()
                       if v and v.feature_importance}
        if not feat_models:
            return
        ws    = wb.create_sheet("🎯 Feature Importance")
        start = self._write_title_block(ws, "Feature Importance — ML Models")
        r     = start
        for name, res in feat_models.items():
            ws.cell(row=r, column=1, value=name).font = self._font(bold=True, size=11)
            r += 1
            self._write_header_row(ws, r, ["Feature", "Importance Score"])
            r += 1
            for i, (feat, imp) in enumerate(res.feature_importance.items()):
                self._write_data_row(ws, r, [feat, round(imp, 5)], shade=(i % 2 == 0))
                r += 1
            r += 1
        self._autofit_columns(ws)

    def generate(
        self, analysis: SalesAnalysis, output_path: Path,
    ) -> Path:
        xl_path = output_path.with_suffix(".xlsx")
        wb      = openpyxl.Workbook()
        wb.remove(wb.active)
        try:
            self._sheet_summary(wb, analysis)
            self._sheet_forecast(wb, analysis)
            self._sheet_models(wb, analysis)
            self._sheet_history(wb, analysis)
            self._sheet_products(wb, analysis)
            self._sheet_decomp(wb, analysis)
            self._sheet_anomalies(wb, analysis)
            self._sheet_config(wb)
            self._sheet_feature_importance(wb, analysis)
            wb.save(str(xl_path))
        except Exception as exc:
            raise ReportGenerationError(f"Excel generation failed: {exc}") from exc
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return xl_path


# ══════════════════════════════════════════════════════════════════════════════
#  TOOLTIP WIDGET
# ══════════════════════════════════════════════════════════════════════════════

class ToolTip:
    def __init__(self, widget: "tk.Widget", text: str, delay: int = 600):
        self._widget = widget
        self._text   = text
        self._delay  = delay
        self._tip:   Optional["tk.Toplevel"] = None
        self._job:   Optional[str] = None
        widget.bind("<Enter>",    self._schedule)
        widget.bind("<Leave>",    self._cancel)
        widget.bind("<Button-1>", self._cancel)

    def _schedule(self, _: Any) -> None:
        self._cancel()
        self._job = self._widget.after(self._delay, self._show)

    def _cancel(self, _: Any = None) -> None:
        if self._job:
            self._widget.after_cancel(self._job)
            self._job = None
        if self._tip:
            try:
                self._tip.destroy()
            except Exception:
                pass
            self._tip = None

    def _show(self) -> None:
        try:
            x = self._widget.winfo_rootx() + 20
            y = self._widget.winfo_rooty() + self._widget.winfo_height() + 4
            self._tip = tk.Toplevel(self._widget)
            self._tip.wm_overrideredirect(True)
            self._tip.wm_geometry(f"+{x}+{y}")
            tk.Label(
                self._tip, text=self._text,
                background="#2D333B", foreground="#E6EDF3",
                relief="flat", bd=0,
                font=("Segoe UI", 8), padx=8, pady=4,
                wraplength=280,
            ).pack()
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
#  SORTABLE TREEVIEW
# ══════════════════════════════════════════════════════════════════════════════

class SortableTree(ttk.Treeview):
    def __init__(self, parent: "tk.Widget",
                 columns: List[Tuple[str, str, int]],
                 height: int = 8):
        col_ids = [c[0] for c in columns]
        super().__init__(parent, columns=col_ids, show="headings", height=height)
        self._sort_col:  str  = col_ids[0]
        self._sort_desc: bool = False
        for col_id, col_label, col_w in columns:
            self.heading(col_id, text=col_label,
                         command=lambda c=col_id: self._sort(c))
            self.column(col_id, width=col_w, anchor="center", minwidth=50)
        style = ttk.Style()
        style.configure("Treeview",
                         background=Theme.SURFACE, foreground=Theme.TEXT2,
                         rowheight=26, fieldbackground=Theme.SURFACE,
                         bordercolor=Theme.BORDER, borderwidth=0)
        style.configure("Treeview.Heading",
                         background=Theme.SURFACE2, foreground=Theme.TEXT,
                         relief="flat", font=("Segoe UI", 9, "bold"))
        style.map("Treeview",
                  background=[("selected", Theme.BLUE)],
                  foreground=[("selected", "#FFFFFF")])

    def _sort(self, col: str) -> None:
        if col == self._sort_col:
            self._sort_desc = not self._sort_desc
        else:
            self._sort_col  = col
            self._sort_desc = False
        rows = [(self.set(k, col), k) for k in self.get_children("")]
        try:
            rows.sort(key=lambda t: float(t[0].replace(",", "").replace("—", "9999")),
                      reverse=self._sort_desc)
        except ValueError:
            rows.sort(key=lambda t: t[0], reverse=self._sort_desc)
        for i, (_, k) in enumerate(rows):
            self.move(k, "", i)
        arrow = " ▼" if self._sort_desc else " ▲"
        for c in self["columns"]:
            self.heading(c, text=self.heading(c, "text").rstrip(" ▼▲"))
        self.heading(col, text=self.heading(col, "text") + arrow)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ══════════════════════════════════════════════════════════════════════════════

class SalesForecastApp:
    T = Theme

    def __init__(self, root: "tk.Tk"):
        self.root           = root
        self._running       = False
        self._cancel_token: Optional[CancelToken] = None
        self._last_analysis: Optional[SalesAnalysis] = None
        self._queue: queue.Queue = queue.Queue()

        self._init_vars()
        self._build_ui()
        self._load_config_into_ui()
        self._poll()

    def _init_vars(self) -> None:
        self.v_use_demo   = tk.BooleanVar(value=True)
        self.v_data_path  = tk.StringVar()
        self.v_company    = tk.StringVar(value="My Company")
        self.v_currency   = tk.StringVar(value="$")
        self.v_horizon    = tk.IntVar(value=12)
        self.v_use_arima  = tk.BooleanVar(value=True)
        self.v_use_es     = tk.BooleanVar(value=True)
        self.v_use_rf     = tk.BooleanVar(value=True)
        self.v_use_gb     = tk.BooleanVar(value=True)
        self.v_use_ridge  = tk.BooleanVar(value=True)
        self.v_use_xgb    = tk.BooleanVar(value=True)
        self.v_use_lgbm   = tk.BooleanVar(value=True)
        self.v_use_prophet= tk.BooleanVar(value=True)
        self.v_make_pdf   = tk.BooleanVar(value=True)
        self.v_make_excel = tk.BooleanVar(value=True)

    def _load_config_into_ui(self) -> None:
        cfg = ForecastConfig.load()
        self.v_company.set(cfg.company_name)
        self.v_currency.set(cfg.currency_symbol)
        self.v_horizon.set(cfg.forecast_horizon)
        self.v_use_arima.set(cfg.use_arima)
        self.v_use_es.set(cfg.use_exp_smoothing)
        self.v_use_rf.set(cfg.use_random_forest)
        self.v_use_gb.set(cfg.use_gradient_boosting)
        self.v_use_ridge.set(cfg.use_ridge)
        self.v_use_xgb.set(cfg.use_xgboost)
        self.v_use_lgbm.set(cfg.use_lightgbm)
        self.v_use_prophet.set(cfg.use_prophet)
        self.v_make_pdf.set("pdf" in cfg.report_format)
        self.v_make_excel.set("excel" in cfg.report_format)
        self._update_hz(None)
        self._log_msg(f"Config loaded — {cfg.company_name}", "accent")

    def _build_config_object(self) -> ForecastConfig:
        cfg = ForecastConfig.load()
        cfg.company_name          = self.v_company.get().strip() or "My Company"
        cfg.currency_symbol       = self.v_currency.get().strip() or "$"
        cfg.forecast_horizon      = self.v_horizon.get()
        cfg.use_arima             = self.v_use_arima.get()
        cfg.use_exp_smoothing     = self.v_use_es.get()
        cfg.use_random_forest     = self.v_use_rf.get()
        cfg.use_gradient_boosting = self.v_use_gb.get()
        cfg.use_ridge             = self.v_use_ridge.get()
        cfg.use_xgboost           = self.v_use_xgb.get()
        cfg.use_lightgbm          = self.v_use_lgbm.get()
        cfg.use_prophet           = self.v_use_prophet.get()
        cfg.report_format = (
            (["pdf"]    if self.v_make_pdf.get()   else []) +
            (["excel"]  if self.v_make_excel.get() else []) +
            ["json"]
        )
        return cfg

    def _build_ui(self) -> None:
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        outer = tk.Frame(self.root, bg=self.T.BG)
        outer.grid(row=0, column=0, sticky="nsew")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(1, weight=1)

        # ── Header ─────────────────────────────────────────────────────────────
        hdr = tk.Frame(outer, bg=self.T.SURFACE,
                        highlightbackground=self.T.BORDER, highlightthickness=1)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text=f"  {APP_TITLE}",
                  font=("Segoe UI", 14, "bold"),
                  bg=self.T.SURFACE, fg=self.T.TEXT).pack(side="left", pady=12)
        tk.Label(hdr, text=f"v{APP_VERSION}",
                  font=("Segoe UI", 9),
                  bg=self.T.SURFACE, fg=self.T.MUTED).pack(side="left", padx=6)

        avail = []
        if _XGBOOST_AVAILABLE:  avail.append("XGBoost")
        if _LIGHTGBM_AVAILABLE: avail.append("LightGBM")
        if _PROPHET_AVAILABLE:  avail.append("Prophet")
        if avail:
            tk.Label(hdr, text="  +  " + "  ·  ".join(avail),
                      font=("Segoe UI", 8), bg=self.T.SURFACE,
                      fg=self.T.SUCCESS).pack(side="left")

        self._status_lbl = tk.Label(
            hdr, text="Ready", font=("Segoe UI", 9),
            bg=self.T.SURFACE, fg=self.T.MUTED)
        self._status_lbl.pack(side="right", padx=14)

        # ── Notebook ───────────────────────────────────────────────────────────
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TNotebook",           background=self.T.BG, borderwidth=0)
        style.configure("TNotebook.Tab",       background=self.T.SURFACE2,
                         foreground=self.T.MUTED, padding=[14, 8],
                         font=("Segoe UI", 9))
        style.map("TNotebook.Tab",
                  background=[("selected", self.T.SURFACE)],
                  foreground=[("selected", self.T.TEXT)])

        self._nb = ttk.Notebook(outer)
        self._nb.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)

        tab_setup   = tk.Frame(self._nb, bg=self.T.BG)
        tab_results = tk.Frame(self._nb, bg=self.T.BG)
        tab_log     = tk.Frame(self._nb, bg=self.T.BG)
        self._nb.add(tab_setup,   text="⚙  Setup")
        self._nb.add(tab_results, text="📊  Results")
        self._nb.add(tab_log,     text="📋  Log")
        self._tab_results = tab_results

        self._build_setup_tab(tab_setup)
        self._build_results_tab(tab_results)
        self._build_log_tab(tab_log)

        # ── Bottom toolbar ─────────────────────────────────────────────────────
        toolbar = tk.Frame(outer, bg=self.T.SURFACE,
                            highlightbackground=self.T.BORDER, highlightthickness=1)
        toolbar.grid(row=2, column=0, sticky="ew")
        toolbar.columnconfigure(3, weight=1)

        self._run_btn = tk.Button(
            toolbar,
            text="  🚀  Run Analysis & Generate Reports",
            font=("Segoe UI", 10, "bold"),
            bg=self.T.ACCENT, fg="#FFFFFF",
            activebackground=self.T.ACCENT_HV, activeforeground="#FFFFFF",
            relief="flat", bd=0, padx=20, pady=10, cursor="hand2",
            command=self._run)
        self._run_btn.grid(row=0, column=0, padx=(12, 6), pady=8)

        self._cancel_btn = tk.Button(
            toolbar, text="  ✕  Cancel",
            font=("Segoe UI", 9),
            bg=self.T.SURFACE2, fg=self.T.MUTED,
            activebackground=self.T.DANGER, activeforeground="#FFFFFF",
            relief="flat", bd=0, padx=12, pady=10, cursor="hand2",
            state="disabled", command=self._cancel)
        self._cancel_btn.grid(row=0, column=1, padx=(0, 6), pady=8)

        save_cfg_btn = tk.Button(
            toolbar, text="  💾  Save Config",
            font=("Segoe UI", 9),
            bg=self.T.SURFACE2, fg=self.T.TEXT2,
            activebackground=self.T.BORDER, activeforeground=self.T.TEXT,
            relief="flat", bd=0, padx=12, pady=10, cursor="hand2",
            command=self._save_config)
        save_cfg_btn.grid(row=0, column=2, padx=(0, 6), pady=8)
        ToolTip(save_cfg_btn, "Save current settings to TOML config file")

        pbar_frame = tk.Frame(toolbar, bg=self.T.SURFACE)
        pbar_frame.grid(row=0, column=3, sticky="ew", padx=(0, 12))
        pbar_frame.columnconfigure(0, weight=1)

        self._pbar = ttk.Progressbar(pbar_frame, mode="determinate", length=280)
        self._pbar.grid(row=0, column=0, sticky="ew", pady=(12, 4))
        self._pbar.grid_remove()

        self._pct_lbl = tk.Label(pbar_frame, text="", font=("Segoe UI", 8),
                                   bg=self.T.SURFACE, fg=self.T.MUTED)
        self._pct_lbl.grid(row=1, column=0)
        self._pct_lbl.grid_remove()

    # ── Setup tab ──────────────────────────────────────────────────────────────

    def _build_setup_tab(self, parent: "tk.Frame") -> None:
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(0, weight=0)

        scroll_canvas = tk.Canvas(parent, bg=self.T.BG, highlightthickness=0)
        scroll_canvas.grid(row=0, column=0, columnspan=2, sticky="nsew")
        parent.rowconfigure(0, weight=1)

        vsb = tk.Scrollbar(parent, orient="vertical", command=scroll_canvas.yview,
                            bg=self.T.SURFACE, troughcolor=self.T.SURFACE2, relief="flat")
        vsb.grid(row=0, column=2, sticky="ns")
        scroll_canvas.configure(yscrollcommand=vsb.set)

        inner = tk.Frame(scroll_canvas, bg=self.T.BG)
        win_id = scroll_canvas.create_window((0, 0), window=inner, anchor="nw")

        def _resize(e: Any) -> None:
            scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all"))
            scroll_canvas.itemconfig(win_id, width=scroll_canvas.winfo_width())

        inner.bind("<Configure>", _resize)
        scroll_canvas.bind("<Configure>", lambda e: scroll_canvas.itemconfig(
            win_id, width=e.width))

        def _on_mousewheel(e: Any) -> None:
            scroll_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        scroll_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        col_left  = tk.Frame(inner, bg=self.T.BG)
        col_right = tk.Frame(inner, bg=self.T.BG)
        col_left.grid(row=0, column=0, sticky="nsew", padx=(24, 12), pady=12)
        col_right.grid(row=0, column=1, sticky="nsew", padx=(0, 24), pady=12)
        inner.columnconfigure(0, weight=1)
        inner.columnconfigure(1, weight=1)

        self._build_data_card(col_left)
        self._build_config_card(col_left)
        self._build_models_card(col_right)
        self._build_output_card(col_right)

    def _card(self, parent: "tk.Frame") -> "tk.Frame":
        f = tk.Frame(parent, bg=self.T.SURFACE,
                      highlightbackground=self.T.BORDER, highlightthickness=1)
        f.pack(fill="x", expand=False, pady=(0, 14))
        return f

    def _card_header(self, card: "tk.Frame", icon: str, title: str, subtitle: str = "") -> None:
        hdr = tk.Frame(card, bg=self.T.SURFACE2)
        hdr.pack(fill="x")
        row = tk.Frame(hdr, bg=self.T.SURFACE2)
        row.pack(fill="x", padx=14, pady=10)
        tk.Label(row, text=f"{icon}  {title}",
                  font=("Segoe UI", 10, "bold"),
                  bg=self.T.SURFACE2, fg=self.T.TEXT).pack(anchor="w")
        if subtitle:
            tk.Label(row, text=subtitle, font=("Segoe UI", 8),
                      bg=self.T.SURFACE2, fg=self.T.MUTED).pack(anchor="w")
        tk.Frame(card, bg=self.T.BORDER, height=1).pack(fill="x")

    def _labeled_entry(self, parent: "tk.Widget", label: str, var: "tk.Variable",
                       width: int = 20, tooltip: str = "") -> "tk.Entry":
        row = tk.Frame(parent, bg=self.T.SURFACE)
        row.pack(fill="x", padx=14, pady=4)
        row.columnconfigure(1, weight=1)
        tk.Label(row, text=label, font=("Segoe UI", 9),
                  bg=self.T.SURFACE, fg=self.T.MUTED, width=width, anchor="w"
                  ).grid(row=0, column=0, sticky="w")
        e = tk.Entry(row, textvariable=var, font=("Segoe UI", 9),
                      bg=self.T.SURFACE2, fg=self.T.TEXT,
                      insertbackground=self.T.TEXT,
                      relief="flat", bd=5,
                      highlightbackground=self.T.BORDER, highlightthickness=1)
        e.grid(row=0, column=1, sticky="ew", padx=(8, 0), ipady=5)
        if tooltip:
            ToolTip(e, tooltip)
        return e

    def _checkbutton(self, parent: "tk.Widget", text: str,
                     var: "tk.BooleanVar", tooltip: str = "") -> "tk.Checkbutton":
        cb = tk.Checkbutton(
            parent, text=f"  {text}", variable=var,
            font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.TEXT,
            selectcolor=self.T.BLUE, activebackground=self.T.SURFACE,
            activeforeground=self.T.TEXT, cursor="hand2", bd=0)
        cb.pack(anchor="w", padx=14, pady=2)
        if tooltip:
            ToolTip(cb, tooltip)
        return cb

    def _build_data_card(self, parent: "tk.Frame") -> None:
        card = self._card(parent)
        self._card_header(card, "📂", "Data Source",
                           "Load your own file or use built-in demo data")
        body = tk.Frame(card, bg=self.T.SURFACE, pady=8)
        body.pack(fill="x")
        tk.Checkbutton(
            body, text="  Use built-in demo data  (5 products · 60 months)",
            variable=self.v_use_demo, font=("Segoe UI", 10),
            bg=self.T.SURFACE, fg=self.T.TEXT, selectcolor=self.T.ACCENT,
            activebackground=self.T.SURFACE, activeforeground=self.T.TEXT,
            cursor="hand2", bd=0, command=self._toggle_demo).pack(anchor="w", padx=14, pady=(0, 6))
        self._file_row = tk.Frame(body, bg=self.T.SURFACE)
        self._file_row.pack(fill="x", padx=14, pady=(0, 4))
        self._file_row.columnconfigure(0, weight=1)
        tk.Label(self._file_row, text="File  (CSV / TSV / Excel / JSON / Parquet):",
                  font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.MUTED
                  ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 4))
        self._path_entry = tk.Entry(
            self._file_row, textvariable=self.v_data_path,
            font=("Segoe UI", 9), bg=self.T.SURFACE2, fg=self.T.TEXT,
            insertbackground=self.T.TEXT, relief="flat", bd=6,
            highlightbackground=self.T.BORDER, highlightthickness=1)
        self._path_entry.grid(row=1, column=0, sticky="ew", ipady=5)
        btn_frame = tk.Frame(self._file_row, bg=self.T.SURFACE)
        btn_frame.grid(row=1, column=1, padx=(8, 0))
        tk.Button(btn_frame, text=" Browse… ", font=("Segoe UI", 9),
                   bg=self.T.BLUE, fg="#FFFFFF",
                   activebackground=self.T.BLUE_HV, activeforeground="#FFFFFF",
                   relief="flat", bd=0, padx=10, pady=5, cursor="hand2",
                   command=self._pick_file).pack(side="left")
        tk.Button(btn_frame, text=" 👁 Preview ", font=("Segoe UI", 9),
                   bg=self.T.SURFACE2, fg=self.T.TEXT2,
                   activebackground=self.T.BORDER, activeforeground=self.T.TEXT,
                   relief="flat", bd=0, padx=8, pady=5, cursor="hand2",
                   command=self._preview_data).pack(side="left", padx=(6, 0))
        tk.Label(
            body,
            text=("  Required: 'date', 'sales'   "
                  "Optional: product · region · units · returns · marketing_spend\n"
                  "  Supported: .csv  .tsv  .xlsx  .xlsm  .xls  .json  .parquet"),
            font=("Segoe UI", 8), bg=self.T.SURFACE, fg=self.T.MUTED2,
        ).pack(anchor="w", padx=14, pady=(4, 8))
        self._toggle_demo()

    def _build_config_card(self, parent: "tk.Frame") -> None:
        card = self._card(parent)
        self._card_header(card, "🔧", "Configuration")
        body = tk.Frame(card, bg=self.T.SURFACE, pady=6)
        body.pack(fill="x")
        self._labeled_entry(body, "Company Name",    self.v_company,
                             tooltip="Appears in report headers")
        self._labeled_entry(body, "Currency Symbol", self.v_currency,
                             tooltip="e.g.  $  €  £  ₹")
        hz_row = tk.Frame(body, bg=self.T.SURFACE)
        hz_row.pack(fill="x", padx=14, pady=5)
        hz_row.columnconfigure(1, weight=1)
        tk.Label(hz_row, text="Forecast Horizon", font=("Segoe UI", 9),
                  bg=self.T.SURFACE, fg=self.T.MUTED, width=20, anchor="w"
                  ).grid(row=0, column=0, sticky="w")
        hz_inner = tk.Frame(hz_row, bg=self.T.SURFACE)
        hz_inner.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        hz_inner.columnconfigure(0, weight=1)
        tk.Scale(hz_inner, variable=self.v_horizon, from_=1, to=36,
                  orient="horizontal", bg=self.T.SURFACE, fg=self.T.TEXT,
                  troughcolor=self.T.BORDER2, activebackground=self.T.ACCENT,
                  highlightthickness=0, bd=0, font=("Segoe UI", 8),
                  showvalue=False, command=self._update_hz
                  ).grid(row=0, column=0, sticky="ew")
        self._hz_lbl = tk.Label(hz_inner, text="12 months",
                                 font=("Segoe UI", 9, "bold"),
                                 bg=self.T.SURFACE, fg=self.T.ACCENT, width=10)
        self._hz_lbl.grid(row=0, column=1)
        tk.Frame(card, bg=self.T.SURFACE, height=6).pack()

    def _build_models_card(self, parent: "tk.Frame") -> None:
        card = self._card(parent)
        self._card_header(card, "🤖", "Forecasting Models",
                           "Enabled models are combined via inverse CV-sMAPE stacking")
        body = tk.Frame(card, bg=self.T.SURFACE, pady=8)
        body.pack(fill="x")

        core_models = [
            (self.v_use_arima,  "ARIMA  (auto order, exact MLE)",
             "Grid-search by AIC; CI via ψ-weight variance propagation"),
            (self.v_use_es,     "Holt-Winters  (auto add/mul)",
             "SSE-minimised; AIC selects additive vs multiplicative"),
            (self.v_use_rf,     "Random Forest  (200 trees, PACF lags)",
             "Feature matrix with lag/rolling/Fourier features; bootstrap CI"),
            (self.v_use_gb,     "Gradient Boosting  (200 rounds)",
             "Shared feature matrix; walk-forward cross-validation"),
            (self.v_use_ridge,  "Ridge Regression  (L2 baseline)",
             "Fast regularised linear baseline with bootstrap CI"),
        ]
        for var, label, tip in core_models:
            self._checkbutton(body, label, var, tooltip=tip)

        tk.Frame(body, bg=self.T.BORDER, height=1).pack(fill="x", padx=14, pady=6)
        tk.Label(body, text="  🆕  v2 Advanced Models (optional — install separately)",
                  font=("Segoe UI", 8, "bold"), bg=self.T.SURFACE,
                  fg=self.T.TEAL).pack(anchor="w", padx=14)

        advanced_models = [
            (self.v_use_xgb,    "XGBoost  (300 rounds, L1+L2 reg)",
             f"{'✓ installed' if _XGBOOST_AVAILABLE else '✗ pip install xgboost'}  — "
             "gradient-boosted trees with native regularisation"),
            (self.v_use_lgbm,   "LightGBM  (300 rounds, leaf-wise)",
             f"{'✓ installed' if _LIGHTGBM_AVAILABLE else '✗ pip install lightgbm'}  — "
             "fast histogram-based gradient boosting"),
            (self.v_use_prophet,"Prophet  (trend + seasonality + holidays)",
             f"{'✓ installed' if _PROPHET_AVAILABLE else '✗ pip install prophet'}  — "
             "Meta's additive model with built-in uncertainty intervals"),
        ]
        for var, label, tip in advanced_models:
            cb = self._checkbutton(body, label, var, tooltip=tip)
        tk.Frame(body, bg=self.T.BORDER, height=1).pack(fill="x", padx=14, pady=8)
        tk.Label(body,
                  text="  ⚡ Ensemble: inverse CV-sMAPE weights (walk-forward, non-leaky)",
                  font=("Segoe UI", 8), bg=self.T.SURFACE, fg=self.T.MUTED,
                  ).pack(anchor="w", padx=14, pady=(0, 6))

    def _build_output_card(self, parent: "tk.Frame") -> None:
        card = self._card(parent)
        self._card_header(card, "📁", "Output",
                           "Reports + TOML config saved automatically")
        body = tk.Frame(card, bg=self.T.SURFACE, pady=10)
        body.pack(fill="x")
        tk.Label(body, text="Report formats:", font=("Segoe UI", 9),
                  bg=self.T.SURFACE, fg=self.T.MUTED).pack(anchor="w", padx=14)
        fmt_row = tk.Frame(body, bg=self.T.SURFACE)
        fmt_row.pack(anchor="w", padx=14, pady=4)
        for var, label, icon, tip in [
            (self.v_make_pdf,   "PDF Report",     "📄",
             "Multi-page PDF with charts, diagnostics, and feature importance"),
            (self.v_make_excel, "Excel Workbook", "📊",
             "10-sheet workbook including Config and Feature Importance sheets"),
        ]:
            cb = tk.Checkbutton(
                fmt_row, text=f"  {icon}  {label}", variable=var,
                font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.TEXT,
                selectcolor=self.T.BLUE, activebackground=self.T.SURFACE,
                activeforeground=self.T.TEXT, cursor="hand2", bd=0)
            cb.pack(side="left", padx=(0, 20))
            ToolTip(cb, tip)
        tk.Frame(body, bg=self.T.BORDER, height=1).pack(fill="x", padx=14, pady=8)
        tk.Label(body,
                  text=("  ℹ  JSON summary always saved.\n"
                        "  💾  TOML config auto-saved on each run.\n"
                        f"  📂  Config: {_LOCAL_CONFIG if _LOCAL_CONFIG.parent.is_dir() else _CONFIG_FILE}"),
                  font=("Segoe UI", 8), bg=self.T.SURFACE, fg=self.T.MUTED,
                  justify="left").pack(anchor="w", padx=14, pady=(0, 8))

    def _build_results_tab(self, parent: "tk.Frame") -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)
        hdr = tk.Frame(parent, bg=self.T.BG)
        hdr.grid(row=0, column=0, sticky="ew", padx=24, pady=(18, 10))
        tk.Label(hdr, text="Analysis Results", font=("Segoe UI", 16, "bold"),
                  bg=self.T.BG, fg=self.T.TEXT).pack(anchor="w")
        self._res_sub = tk.Label(
            hdr, text="No results yet — run an analysis first.",
            font=("Segoe UI", 9), bg=self.T.BG, fg=self.T.MUTED)
        self._res_sub.pack(anchor="w")
        self._kpi_bar = tk.Frame(parent, bg=self.T.BG)
        self._kpi_bar.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 10))
        tbl = tk.Frame(parent, bg=self.T.SURFACE,
                        highlightbackground=self.T.BORDER, highlightthickness=1)
        tbl.grid(row=2, column=0, sticky="nsew", padx=24, pady=(0, 10))
        tbl.columnconfigure(0, weight=1)
        tbl.rowconfigure(1, weight=1)
        th = tk.Frame(tbl, bg=self.T.SURFACE)
        th.grid(row=0, column=0, columnspan=2, sticky="ew", padx=16, pady=8)
        tk.Label(th, text="Model Performance",
                  font=("Segoe UI", 11, "bold"),
                  bg=self.T.SURFACE, fg=self.T.TEXT).pack(side="left")
        tk.Label(th,
                  text="Click column headers to sort  ·  sMAPE (lower=better)  ·  CV = walk-forward",
                  font=("Segoe UI", 8), bg=self.T.SURFACE, fg=self.T.MUTED).pack(side="left", padx=(12, 0))
        cols = [
            ("model",    "Model",      160),
            ("smape",    "sMAPE (%)",   80),
            ("mase",     "MASE",        70),
            ("rmse",     "RMSE ($)",    95),
            ("r2",       "R² Score",    75),
            ("cv_smape", "CV sMAPE",    80),
            ("time",     "Time (s)",    70),
        ]
        self._model_tree = SortableTree(tbl, cols, height=9)
        self._model_tree.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 10))
        self._model_tree.tag_configure("ens",  background="#0D2A20", foreground="#3FB950")
        self._model_tree.tag_configure("best", background="#0D1E38", foreground="#58A6FF")
        sb = tk.Scrollbar(tbl, command=self._model_tree.yview,
                           bg=self.T.SURFACE, troughcolor=self.T.SURFACE2, relief="flat")
        sb.grid(row=1, column=1, sticky="ns", pady=(0, 10))
        self._model_tree.configure(yscrollcommand=sb.set)
        self._files_row = tk.Frame(parent, bg=self.T.BG)
        self._files_row.grid(row=3, column=0, sticky="ew", padx=24, pady=(0, 18))

    def _build_log_tab(self, parent: "tk.Frame") -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)
        hdr = tk.Frame(parent, bg=self.T.BG)
        hdr.grid(row=0, column=0, sticky="ew", padx=24, pady=(18, 8))
        tk.Label(hdr, text="Activity Log", font=("Segoe UI", 16, "bold"),
                  bg=self.T.BG, fg=self.T.TEXT).pack(anchor="w")
        tk.Label(hdr, text="Real-time pipeline events — colour-coded by severity",
                  font=("Segoe UI", 9), bg=self.T.BG, fg=self.T.MUTED).pack(anchor="w")
        log_card = tk.Frame(parent, bg=self.T.SURFACE,
                             highlightbackground=self.T.BORDER, highlightthickness=1)
        log_card.grid(row=1, column=0, sticky="nsew", padx=24, pady=(0, 24))
        log_card.columnconfigure(0, weight=1)
        log_card.rowconfigure(0, weight=1)
        self._log = tk.Text(
            log_card, bg=self.T.SURFACE2, fg=self.T.TEXT2,
            font=("Consolas", 9), insertbackground=self.T.TEXT,
            selectbackground=self.T.BLUE, relief="flat", bd=0,
            wrap="word", state="disabled", padx=12, pady=10)
        self._log.grid(row=0, column=0, sticky="nsew")
        sb2 = tk.Scrollbar(log_card, command=self._log.yview,
                            bg=self.T.SURFACE, troughcolor=self.T.SURFACE2, relief="flat")
        sb2.grid(row=0, column=1, sticky="ns")
        self._log.configure(yscrollcommand=sb2.set)
        for tag, fg in [
            ("ok",     self.T.SUCCESS), ("warn",   self.T.WARN),
            ("error",  self.T.DANGER),  ("accent", "#58A6FF"),
            ("muted",  self.T.MUTED),   ("info",   self.T.TEXT2),
            ("sep",    self.T.BORDER),
        ]:
            self._log.tag_configure(tag, foreground=fg)
        ctrl = tk.Frame(log_card, bg=self.T.SURFACE)
        ctrl.grid(row=1, column=0, columnspan=2, sticky="ew", padx=12, pady=6)
        for lbl, cmd in [("Copy All", self._copy_log), ("Clear Log", self._clear_log)]:
            tk.Button(ctrl, text=lbl, font=("Segoe UI", 9),
                       bg=self.T.SURFACE2, fg=self.T.MUTED,
                       activebackground=self.T.BORDER, activeforeground=self.T.TEXT,
                       relief="flat", bd=0, padx=10, pady=4, cursor="hand2",
                       command=cmd).pack(side="right", padx=(0, 6))

    # ── Callbacks ──────────────────────────────────────────────────────────────

    def _toggle_demo(self) -> None:
        if self.v_use_demo.get():
            self._file_row.pack_forget()
        else:
            self._file_row.pack(fill="x", padx=14, pady=(0, 4))

    def _pick_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Sales Data File",
            filetypes=[
                ("All supported", "*.csv *.tsv *.xlsx *.xlsm *.xls *.json *.parquet"),
                ("CSV/TSV",       "*.csv *.tsv"),
                ("Excel",         "*.xlsx *.xlsm *.xls"),
                ("JSON",          "*.json"),
                ("Parquet",       "*.parquet"),
                ("All files",     "*.*"),
            ])
        if path:
            self.v_data_path.set(path)
            self.v_use_demo.set(False)
            self._toggle_demo()
            self._log_msg(f"Data file selected: {Path(path).name}", "accent")

    def _preview_data(self) -> None:
        path_str = self.v_data_path.get().strip()
        if not path_str:
            messagebox.showinfo("Preview", "Please select a data file first.")
            return
        try:
            df = DataLoader.load(path_str)
        except Exception as exc:
            messagebox.showerror("Preview Error", str(exc))
            return
        top = tk.Toplevel(self.root)
        top.title(f"Data Preview — {Path(path_str).name}")
        top.configure(bg=self.T.BG)
        top.geometry("900x450")
        cols = list(df.columns)
        tree = ttk.Treeview(top, columns=cols, show="headings", height=18)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(80, min(160, len(c) * 9)), anchor="center")
        for _, row in df.head(20).iterrows():
            tree.insert("", "end", values=[str(v)[:30] for v in row])
        vsb = tk.Scrollbar(top, orient="vertical",   command=tree.yview)
        hsb = tk.Scrollbar(top, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        top.columnconfigure(0, weight=1)
        top.rowconfigure(0, weight=1)
        tk.Label(top,
                  text=f"Showing first 20 of {len(df):,} rows  ·  "
                       f"{len(cols)} columns  ·  "
                       f"Date range: {df['date'].min().strftime('%Y-%m')} – "
                       f"{df['date'].max().strftime('%Y-%m')}",
                  font=("Segoe UI", 8), bg=self.T.BG, fg=self.T.MUTED
                  ).grid(row=2, column=0, columnspan=2, padx=8, pady=6, sticky="w")

    def _update_hz(self, _: Any) -> None:
        h = self.v_horizon.get()
        self._hz_lbl.configure(text=f"{h} month{'s' if h != 1 else ''}")

    def _set_status(self, text: str, color: str = Theme.MUTED) -> None:
        try:
            self._status_lbl.configure(text=text, fg=color)
        except Exception:
            pass

    def _log_msg(self, msg: str, tag: str = "info") -> None:
        try:
            self._log.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            self._log.insert("end", f"[{ts}]  {msg}\n", tag)
            self._log.see("end")
            self._log.configure(state="disabled")
        except Exception:
            pass

    def _copy_log(self) -> None:
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(self._log.get("1.0", "end"))
        except Exception:
            pass

    def _clear_log(self) -> None:
        try:
            self._log.configure(state="normal")
            self._log.delete("1.0", "end")
            self._log.configure(state="disabled")
        except Exception:
            pass

    def _save_config(self) -> None:
        cfg = self._build_config_object()
        if cfg.save():
            target = _LOCAL_CONFIG if _LOCAL_CONFIG.parent.is_dir() else _CONFIG_FILE
            self._log_msg(f"Config saved → {target}", "ok")
            messagebox.showinfo("Config Saved", f"Settings saved to:\n{target}")
        else:
            messagebox.showwarning("Save Failed",
                                   "Could not save config. Check log for details.")

    def _poll(self) -> None:
        try:
            while True:
                item = self._queue.get_nowait()
                kind = item[0]
                if kind == "log":
                    self._log_msg(item[1], item[2] if len(item) > 2 else "info")
                elif kind == "progress":
                    msg, pct = item[1], item[2]
                    self._log_msg(msg, "info")
                    self._set_status(msg, self.T.WARN)
                    if pct is not None:
                        try:
                            self._pbar["value"] = pct
                            self._pct_lbl.configure(text=f"{int(pct)}%")
                        except Exception:
                            pass
                elif kind == "done":
                    self._on_done(item[1])
                elif kind == "error":
                    self._on_error(item[1])
        except queue.Empty:
            pass
        self.root.after(80, self._poll)

    def _run(self) -> None:
        if self._running:
            return
        cfg = self._build_config_object()
        errs = cfg.validate()
        if errs:
            messagebox.showerror("Validation Errors", "\n".join(f"• {e}" for e in errs))
            return
        out_path = filedialog.asksaveasfilename(
            title="Save Reports As",
            defaultextension="",
            initialdir=str(_LOG_DIR),
            initialfile="sales_report",
            filetypes=[("Base name (no extension)", "*")],
        )
        if not out_path:
            return
        cfg.output_path = Path(out_path)
        cfg.save()

        self._running      = True
        self._cancel_token = CancelToken()
        self._run_btn.configure(state="disabled", text="  ⏳  Running …")
        self._cancel_btn.configure(state="normal")
        self._pbar.grid()
        self._pbar["value"] = 0
        self._pct_lbl.grid()
        self._pct_lbl.configure(text="0%")
        self._set_status("Running …", self.T.WARN)
        self._log_msg("━" * 56, "sep")
        self._log_msg(f"Starting analysis → {cfg.output_path.name}", "accent")

        t = threading.Thread(
            target=self._pipeline,
            args=(cfg, self._queue, self._cancel_token),
            daemon=True)
        t.start()

    @staticmethod
    def _pipeline(
        cfg: ForecastConfig,
        q: queue.Queue,
        cancel: CancelToken,
    ) -> None:
        def report(msg: str, pct: Optional[float] = None) -> None:
            q.put(("progress", msg, pct))

        def log(msg: str, tag: str = "info") -> None:
            q.put(("log", msg, tag))

        try:
            cfg.ensure_dirs()
            log(f"Output directory: {cfg.output_path.parent}", "muted")

            log("Loading data …", "info")
            source = None if not Path(cfg.output_path.parent / "dummy").exists() else None
            source_path = None
            try:
                data_path_var = getattr(cfg, "_data_path_override", None)
            except Exception:
                data_path_var = None

            df = DataLoader.load(data_path_var)
            log(f"Loaded {len(df):,} rows, {len(df.columns)} columns", "ok")
            log(f"Date range: {df['date'].min().strftime('%Y-%m')} → {df['date'].max().strftime('%Y-%m')}", "muted")

            analyzer = SalesAnalyzer(cfg, report, cancel)
            analysis = analyzer.analyze(df)

            report("Generating charts …", 91)
            charts = ChartGenerator(cfg).generate_all(analysis, report)
            log(f"Charts: {len(charts)} rendered", "ok")

            output_files: Dict[str, Path] = {}
            if "pdf" in cfg.report_format:
                report("Building PDF report …", 95)
                try:
                    pdf_gen = PDFReportGenerator(cfg)
                    pdf_p   = pdf_gen.generate(analysis, charts, cfg.output_path)
                    output_files["pdf"] = pdf_p
                    sz = pdf_p.stat().st_size // 1024
                    log(f"PDF saved: {pdf_p.name} ({sz:,} KB)", "ok")
                except Exception as exc:
                    log(f"PDF failed: {exc}", "warn")

            if "excel" in cfg.report_format:
                report("Building Excel workbook …", 97)
                try:
                    xl_p = ExcelReportGenerator(cfg).generate(analysis, cfg.output_path)
                    output_files["excel"] = xl_p
                    sz = xl_p.stat().st_size // 1024
                    log(f"Excel saved: {xl_p.name} ({sz:,} KB)", "ok")
                except Exception as exc:
                    log(f"Excel failed: {exc}", "warn")

            json_path = cfg.output_path.with_suffix(".json")
            st        = analysis.statistics
            ens       = analysis.ensemble_forecast
            summary   = {
                "generated_at":  datetime.now().isoformat(),
                "app_version":   APP_VERSION,
                "company":       cfg.company_name,
                "statistics":    st,
                "best_model":    analysis.best_model,
                "ensemble": {
                    "smape":    ens.smape,
                    "mase":     ens.mase,
                    "rmse":     ens.rmse,
                    "r2":       ens.r2,
                    "forecast": ens.predictions.tolist(),
                    "lower_ci": ens.lower_ci.tolist(),
                    "upper_ci": ens.upper_ci.tolist(),
                    "weights":  ens.metadata.get("weights", {}),
                },
                "models": {
                    name: {
                        "smape":    res.smape,
                        "mase":     res.mase,
                        "rmse":     res.rmse,
                        "r2":       res.r2,
                        "cv_smape": res.cv_smape,
                        "aic":      res.aic,
                        "metadata": res.metadata,
                    }
                    for name, res in analysis.forecasts.items() if res is not None
                },
                "optional_models": {
                    "xgboost_available":  _XGBOOST_AVAILABLE,
                    "lightgbm_available": _LIGHTGBM_AVAILABLE,
                    "prophet_available":  _PROPHET_AVAILABLE,
                },
                "platform_info": {
                    "python":   sys.version,
                    "platform": sys.platform,
                    "pandas":   pd.__version__,
                    "numpy":    np.__version__,
                },
            }
            json_path.write_text(
                json.dumps(summary, indent=2, cls=_NpEncoder), encoding="utf-8")
            output_files["json"] = json_path
            log(f"JSON summary saved: {json_path.name}", "ok")

            for _attempt in range(3):
                try:
                    shutil.rmtree(cfg.charts_dir, ignore_errors=False)
                    break
                except Exception:
                    time.sleep(0.3)

            report("Complete! ✓", 100)
            q.put(("done", {"analysis": analysis, "files": output_files, "cfg": cfg}))

        except InterruptedError:
            q.put(("log", "Analysis cancelled by user.", "warn"))
            q.put(("error", "cancelled"))
        except (DataLoadError, ForecastError, ReportGenerationError) as exc:
            q.put(("log", f"ERROR: {exc}", "error"))
            q.put(("error", str(exc)))
        except Exception as exc:
            q.put(("log", f"UNEXPECTED ERROR: {exc}", "error"))
            q.put(("log", traceback.format_exc(), "error"))
            q.put(("error", str(exc)))

    def _on_done(self, result: Dict) -> None:
        self._running = False
        self._run_btn.configure(state="normal", text="  🚀  Run Analysis & Generate Reports")
        self._cancel_btn.configure(state="disabled")
        self._pbar.grid_remove()
        self._pct_lbl.grid_remove()
        self._set_status("✓  Analysis complete", self.T.SUCCESS)
        self._log_msg("━" * 56, "sep")
        self._log_msg("ANALYSIS COMPLETE ✓", "ok")
        self._log_msg("━" * 56, "sep")

        analysis = result["analysis"]
        files    = result["files"]
        cfg      = result["cfg"]
        self._last_analysis = analysis
        self._populate_results(analysis, files, cfg)
        self._nb.select(self._tab_results)

        file_lines = "\n".join(
            f"• {fmt.upper()}: {Path(p).name}" for fmt, p in files.items())
        messagebox.showinfo(
            "Analysis Complete 🎉",
            f"All reports saved to:\n{cfg.output_path.parent}\n\n"
            f"{file_lines}\n\n"
            f"Best Model      : {analysis.best_model}\n"
            f"Ensemble sMAPE  : {analysis.ensemble_forecast.smape:.2f}%\n"
            f"Ensemble MASE   : {analysis.ensemble_forecast.mase:.3f}\n"
            f"Anomalies found : {len(analysis.anomalies)}")

    def _on_error(self, err: str) -> None:
        self._running = False
        self._run_btn.configure(state="normal", text="  🚀  Run Analysis & Generate Reports")
        self._cancel_btn.configure(state="disabled")
        self._pbar.grid_remove()
        self._pct_lbl.grid_remove()
        if err == "cancelled":
            self._set_status("Cancelled", self.T.MUTED)
            return
        self._set_status("✗  Analysis failed", self.T.DANGER)
        messagebox.showerror("Analysis Failed",
                              f"An error occurred:\n\n{err}\n\n"
                              "See the Log tab for full traceback.")

    def _cancel(self) -> None:
        if self._cancel_token:
            self._cancel_token.cancel()
        self._cancel_btn.configure(state="disabled")
        self._set_status("Cancelling …", self.T.WARN)
        self._log_msg("Cancel requested — stopping after current step …", "warn")

    def _populate_results(
        self, analysis: SalesAnalysis, files: Dict, cfg: ForecastConfig,
    ) -> None:
        st  = analysis.statistics
        ens = analysis.ensemble_forecast
        self._res_sub.configure(
            text=(f"Completed {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  "
                  f"Best model: {analysis.best_model}  ·  "
                  f"sMAPE: {ens.smape:.2f}%  ·  MASE: {ens.mase:.3f}"))

        for w in self._kpi_bar.winfo_children():
            w.destroy()
        kpis = [
            ("💰", "Total Revenue",  f"{cfg.currency_symbol}{st['total_revenue']/1e6:.2f}M", self.T.BLUE),
            ("📅", "Avg / Month",    f"{cfg.currency_symbol}{st['avg_monthly_revenue']/1e3:.0f}K", self.T.BLUE),
            ("📈", "CAGR",           f"{st['cagr']:.1f}%",        self.T.SUCCESS),
            ("🔮", "Forecast Total", f"{cfg.currency_symbol}{ens.predictions.sum()/1e6:.2f}M", self.T.PURPLE),
            ("🎯", "sMAPE",          f"{ens.smape:.1f}%",          self.T.WARN),
            ("📣", "Mktg ROI",       f"{st['marketing_roi']:.1f}x", self.T.TEAL),
            ("⚠️", "Anomalies",      f"{len(analysis.anomalies)}",  self.T.DANGER),
        ]
        for i, (icon, label, val, color) in enumerate(kpis):
            kc = tk.Frame(self._kpi_bar, bg=self.T.SURFACE,
                           highlightbackground=self.T.BORDER, highlightthickness=1)
            kc.pack(side="left", expand=True, fill="x", padx=(0 if i == 0 else 6, 0), pady=2)
            tk.Label(kc, text=icon,  font=("Segoe UI", 13),
                      bg=self.T.SURFACE, fg=color).pack(pady=(10, 1))
            tk.Label(kc, text=val,   font=("Segoe UI", 13, "bold"),
                      bg=self.T.SURFACE, fg=color).pack()
            tk.Label(kc, text=label, font=("Segoe UI", 8),
                      bg=self.T.SURFACE, fg=self.T.MUTED).pack(pady=(1, 10))

        for row in self._model_tree.get_children():
            self._model_tree.delete(row)
        valid = {k: v for k, v in analysis.forecasts.items() if v is not None}
        valid["Ensemble ★"] = ens
        for name, res in valid.items():
            tag    = ("ens"  if "Ensemble" in name else
                      "best" if name == analysis.best_model else "")
            cv_str = (f"{res.cv_smape:.2f}"
                      if res.cv_smape is not None and math.isfinite(res.cv_smape)
                      else "—")
            self._model_tree.insert("", "end",
                values=(name,
                        f"{res.smape:.2f}", f"{res.mase:.3f}",
                        f"{cfg.currency_symbol}{res.rmse:,.0f}",
                        f"{res.r2:.3f}", cv_str,
                        f"{res.train_time_s:.2f}"),
                tags=(tag,))

        for w in self._files_row.winfo_children():
            w.destroy()
        tk.Label(self._files_row, text="Generated files:",
                  font=("Segoe UI", 10, "bold"),
                  bg=self.T.BG, fg=self.T.TEXT).pack(anchor="w", pady=(0, 6))
        btn_row = tk.Frame(self._files_row, bg=self.T.BG)
        btn_row.pack(anchor="w")
        icons = {"pdf": "📄", "excel": "📊", "json": "🗒️"}
        for fmt, path in files.items():
            p = Path(path)
            if not p.exists():
                continue
            sz = p.stat().st_size // 1024
            b  = tk.Button(
                btn_row,
                text=f"  {icons.get(fmt, '📁')}  {p.name}  ({sz:,} KB)",
                font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.TEXT2,
                activebackground=self.T.BORDER, activeforeground=self.T.TEXT,
                relief="flat", bd=0, padx=12, pady=7, cursor="hand2",
                command=lambda _p=p: self._open_file(_p))
            b.pack(side="left", padx=(0, 8))
            ToolTip(b, f"Click to open {p.name}")
        tk.Button(
            btn_row, text="  📂  Open Folder",
            font=("Segoe UI", 9), bg=self.T.BLUE, fg="#FFFFFF",
            activebackground=self.T.BLUE_HV, activeforeground="#FFFFFF",
            relief="flat", bd=0, padx=12, pady=7, cursor="hand2",
            command=lambda: self._open_folder(cfg.output_path.parent),
        ).pack(side="left")

    @staticmethod
    def _open_file(path: Path) -> None:
        try:
            if sys.platform.startswith("darwin"):
                os.system(f'open "{path}"')
            elif sys.platform.startswith("win"):
                os.startfile(str(path))   # type: ignore[attr-defined]
            else:
                os.system(f'xdg-open "{path}" &')
        except Exception as exc:
            messagebox.showwarning("Open Failed", str(exc))

    @staticmethod
    def _open_folder(folder: Path) -> None:
        try:
            if sys.platform.startswith("darwin"):
                os.system(f'open "{folder}"')
            elif sys.platform.startswith("win"):
                os.startfile(str(folder)) # type: ignore[attr-defined]
            else:
                os.system(f'xdg-open "{folder}" &')
        except Exception as exc:
            messagebox.showwarning("Open Folder Failed", str(exc))


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    if _IMPORT_ERROR:
        _err_root = tk.Tk()
        _err_root.withdraw()
        messagebox.showerror(
            "Missing Dependencies",
            f"A required library is missing:\n\n{_IMPORT_ERROR}\n\n"
            "Install core dependencies:\n\n"
            "  pip install numpy pandas matplotlib seaborn scipy \\\n"
            "              scikit-learn openpyxl reportlab pillow\n\n"
            "Install optional v2 models (any or all):\n\n"
            "  pip install xgboost lightgbm prophet\n\n"
            "Install TOML support (Python < 3.11):\n\n"
            "  pip install tomli tomli-w\n\n"
            "Then restart the application.")
        _err_root.destroy()
        sys.exit(1)

    root = tk.Tk()
    root.title(f"{APP_TITLE}  v{APP_VERSION}")
    root.configure(bg=Theme.BG)

    w, h = 1220, 820
    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    root.geometry(f"{w}x{h}+{max(0, (sw - w) // 2)}+{max(0, (sh - h) // 2)}")
    root.minsize(900, 680)

    if sys.platform.startswith("win"):
        try:
            import ctypes
            hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 20,
                ctypes.byref(ctypes.c_int(2)), ctypes.sizeof(ctypes.c_int(2)))
        except Exception:
            pass

    app = SalesForecastApp(root)

    def _on_close() -> None:
        try:
            if app._cancel_token is not None:
                app._cancel_token.cancel()
        except Exception:
            pass
        try:
            root.destroy()
        except Exception:
            pass

    root.protocol("WM_DELETE_WINDOW", _on_close)

    app._log_msg(f"Sales Forecasting System {APP_VERSION} ready.", "ok")
    app._log_msg("━" * 56, "sep")
    app._log_msg("⚙  Setup tab → configure → Run Analysis", "accent")
    app._log_msg("💾  Config auto-saved to TOML on each run", "accent")
    app._log_msg("━" * 56, "sep")
    app._log_msg("Core models: ARIMA · Holt-Winters · RF · GBM · Ridge", "muted")
    app._log_msg(f"XGBoost   : {'✓ ready' if _XGBOOST_AVAILABLE  else '✗ pip install xgboost'}", "muted")
    app._log_msg(f"LightGBM  : {'✓ ready' if _LIGHTGBM_AVAILABLE else '✗ pip install lightgbm'}", "muted")
    app._log_msg(f"Prophet   : {'✓ ready' if _PROPHET_AVAILABLE  else '✗ pip install prophet'}", "muted")
    cfg_src = _LOCAL_CONFIG if _LOCAL_CONFIG.exists() else (_CONFIG_FILE if _CONFIG_FILE.exists() else None)
    app._log_msg(f"Config    : {cfg_src or 'defaults (no config file found)'}", "muted")
    app._log_msg("━" * 56, "sep")

    try:
        root.mainloop()
    except KeyboardInterrupt:
        logger.info("Interrupted by user.")
    finally:
        try:
            root.destroy()
        except Exception:
            pass


if __name__ == "__main__":
    main()
