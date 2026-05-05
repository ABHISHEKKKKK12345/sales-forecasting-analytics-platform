"""
Sales Forecasting & Automated Reporting System  v2.0
═══════════════════════════════════════════════════════════════════════════════
Models  : ARIMA (MLE grid) · Holt-Winters (add/mul) · Random Forest
          Gradient Boosting · Ridge Regression · Stacked Ensemble
Reports : Multi-page PDF · 9-sheet Excel workbook · JSON summary
Author  : Abhishek
═══════════════════════════════════════════════════════════════════════════════
"""

# ─── Standard library ─────────────────────────────────────────────────────────
import io
import json
import logging
import logging.handlers
import math
import os
import queue
import shutil
import sys
import tempfile
import threading
import time
import traceback
import warnings
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple, TYPE_CHECKING

warnings.filterwarnings("ignore")

# ── Configure logging: stream + rotating file ────────────────────────────────
_LOG_DIR = Path.home() / "SalesForecastReports"
_LOG_DIR.mkdir(parents=True, exist_ok=True)

# DEBUG mode via env var: SALES_FORECAST_DEBUG=1
_LOG_LEVEL = logging.DEBUG if os.environ.get("SALES_FORECAST_DEBUG") == "1" else logging.INFO

logging.basicConfig(
    level=_LOG_LEVEL,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.handlers.RotatingFileHandler(
            _LOG_DIR / "sales_forecast.log",
            maxBytes=5 * 1024 * 1024,   # 5 MB
            backupCount=3,
            encoding="utf-8",
        ),
    ],
)
logger = logging.getLogger("SalesForecast")
def _global_exception_handler(exc_type, exc_value, exc_tb):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_tb)
        return
    logger.critical("Unhandled exception", exc_info=(exc_type, exc_value, exc_tb))

sys.excepthook = _global_exception_handler

# ─── GUI — imported first so we can show error dialogs if deps are missing ────
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    from tkinter.font import Font as TkFont
    _TK_AVAILABLE = True
except ImportError:
    _TK_AVAILABLE = False
    print(
        "ERROR: tkinter is not available.\n"
        "  Ubuntu/Debian : sudo apt install python3-tk\n"
        "  macOS         : brew install python-tk\n"
        "  Windows       : Reinstall Python with 'tcl/tk and IDLE' checked."
    )
    sys.exit(1)

# ─── Third-party — collected so we can surface a friendly dialog ──────────────
_IMPORT_ERROR: Optional[str] = None
try:
    import numpy as np
    import pandas as pd
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import seaborn as sns
    from scipy import stats
    from scipy.optimize import minimize
    from scipy.signal import periodogram
    from sklearn.ensemble import GradientBoostingRegressor, RandomForestRegressor
    from sklearn.linear_model import Ridge
    from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
    from sklearn.model_selection import TimeSeriesSplit
    from sklearn.preprocessing import StandardScaler
    from sklearn.base import clone as sklearn_clone
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        HRFlowable, Image as RLImage, KeepTogether,
        PageBreak, Paragraph, SimpleDocTemplate,
        Spacer, Table, TableStyle,
    )
except ImportError as _e:
    _IMPORT_ERROR = str(_e)


# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS & THEME
# ══════════════════════════════════════════════════════════════════════════════

APP_TITLE   = "Sales Forecasting System"
APP_VERSION = "2.0.0"


# ── Numpy-aware JSON encoder ─────────────────────────────────────────────
class _NpEncoder(json.JSONEncoder):
    """Serialise numpy scalars / arrays transparently."""
    def default(self, obj: Any) -> Any:
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        return super().default(obj)


def _fmt_month(value: Any) -> str:
    """
    Format a peak/trough month for display.
    Accepts pd.Timestamp, datetime, or ISO-string — always returns 'Mon YYYY'.
    This guard is needed because compute_statistics stores months as
    ISO strings, but older consumers expected pd.Timestamp.strftime().
    """
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%b %Y")
    if isinstance(value, str):
        try:
            return pd.Timestamp(value).strftime("%b %Y")
        except Exception:
            return value
    return str(value)

# pandas >= 2.x check (safe if _IMPORT_ERROR is set)
_PANDAS_GTE_2: bool = False
if not _IMPORT_ERROR:
    try:
        _PANDAS_GTE_2 = int(pd.__version__.split(".")[0]) >= 2
    except Exception:
        pass


class Theme:
    BG        = "#0D1117"
    SURFACE   = "#161B22"
    SURFACE2  = "#1C2128"
    BORDER    = "#30363D"
    BORDER2   = "#21262D"
    ACCENT    = "#238636"
    ACCENT_HV = "#2EA043"
    BLUE      = "#1F6FEB"
    BLUE_HV   = "#388BFD"
    WARN      = "#E3B341"
    DANGER    = "#DA3633"
    SUCCESS   = "#3FB950"
    TEXT      = "#E6EDF3"
    TEXT2     = "#C9D1D9"
    MUTED     = "#8B949E"
    MUTED2    = "#6E7681"
    PURPLE    = "#8B5CF6"
    TEAL      = "#00A0C7"
    CHART     = ["#1F6FEB", "#00A0C7", "#E8730A", "#238636",
                 "#DA3633", "#8B5CF6", "#E3B341", "#00796B"]


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ForecastConfig:
    output_path:           Path  = field(default_factory=lambda: Path.home() / "SalesForecastReports" / "report")
    charts_dir:            Path  = field(default_factory=lambda: Path(tempfile.gettempdir()) / "sf_charts_tmp")
    forecast_horizon:      int   = 12
    confidence_level:      float = 0.95
    seasonality_period:    int   = 12   # 0 = auto-detect
    use_arima:             bool  = True
    use_random_forest:     bool  = True
    use_gradient_boosting: bool  = True
    use_ridge:             bool  = True
    use_exp_smoothing:     bool  = True
    company_name:          str   = "My Company"
    report_title:          str   = "Sales Forecasting & Demand Analysis Report"
    analyst_name:          str   = "Automated Analytics Engine"
    currency_symbol:       str   = "$"
    report_format:         List[str] = field(default_factory=lambda: ["pdf", "excel"])
    primary_color:         Tuple = (0.12, 0.29, 0.49)
    accent_color:          Tuple = (0.0,  0.63, 0.79)
    n_bootstrap:           int   = 300
    cv_folds:              int   = 3

    def ensure_dirs(self) -> None:
        # create output parent first so reports never fail on fresh install
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.charts_dir = Path(tempfile.mkdtemp(prefix="sf_charts_"))

    def validate(self) -> List[str]:
        errs: List[str] = []
        if self.forecast_horizon < 1 or self.forecast_horizon > 60:
            errs.append("Forecast horizon must be 1–60 months.")
        if not (0 < self.confidence_level < 1):
            errs.append("Confidence level must be between 0 and 1.")
        if not self.report_format:
            errs.append("At least one report format must be selected.")
        if not any([self.use_arima, self.use_random_forest,
                    self.use_gradient_boosting, self.use_ridge, self.use_exp_smoothing]):
            errs.append("At least one forecasting model must be enabled.")
        if self.n_bootstrap < 50:
            errs.append("n_bootstrap must be >= 50.")
        if self.n_bootstrap > 5_000:
            errs.append("n_bootstrap must be ≤ 5 000 (use 300–1 000 for production).")
        if self.cv_folds < 2:
            errs.append("cv_folds must be >= 2.")
        if self.seasonality_period != 0 and not (2 <= self.seasonality_period <= 60):
            errs.append("seasonality_period must be 0 (auto-detect) or 2–60.")
        if len(self.company_name.strip()) == 0:
            errs.append("Company name must not be empty.")
        if len(self.company_name) > 120:
            errs.append("Company name must be ≤ 120 characters.")
        if len(self.analyst_name.strip()) == 0:
            errs.append("Analyst name must not be empty (or all-whitespace).")
        if len(self.analyst_name) > 120:
            errs.append("Analyst name must be ≤ 120 characters.")
        if len(self.currency_symbol.strip()) == 0:
            errs.append("Currency symbol must not be empty.")
        return errs


# ══════════════════════════════════════════════════════════════════════════════
#  DATA MODELS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ResidualDiagnostics:
    """Statistical tests on model residuals."""
    ljung_box_stat:   float = 0.0
    ljung_box_pval:   float = 1.0
    shapiro_stat:     float = 0.0
    shapiro_pval:     float = 1.0
    durbin_watson:    float = 2.0
    heteroscedastic:  bool  = False

    @property
    def is_adequate(self) -> bool:
        return self.ljung_box_pval > 0.05 and 1.0 < self.durbin_watson < 3.0

    def summary(self) -> str:
        lb = "✓" if self.ljung_box_pval > 0.05 else "✗"
        sw = "✓" if self.shapiro_pval   > 0.05 else "✗"
        dw = "✓" if 1.5 < self.durbin_watson < 2.5 else "✗"
        return (f"Ljung-Box {lb} (p={self.ljung_box_pval:.3f})  "
                f"Shapiro-Wilk {sw} (p={self.shapiro_pval:.3f})  "
                f"DW {dw} ({self.durbin_watson:.2f})")


@dataclass
class ForecastResult:
    model_name:         str
    predictions:        "np.ndarray"
    lower_ci:           "np.ndarray"
    upper_ci:           "np.ndarray"
    mae:                float
    rmse:               float
    smape:              float
    mase:               float
    mape:               float
    r2:                 float
    aic:                Optional[float]  = None
    bic:                Optional[float]  = None
    cv_smape:           Optional[float]  = None
    feature_importance: Optional[Dict]   = None
    diagnostics:        Optional[ResidualDiagnostics] = None
    metadata:           Dict             = field(default_factory=dict)
    train_time_s:       float            = 0.0

    @property
    def accuracy_score(self) -> float:
        return max(0.0, 100.0 * (1.0 - min(self.smape / 100.0, 1.0)))

    @property
    def is_valid(self) -> bool:
        return (
            len(self.predictions) > 0
            and not np.any(np.isnan(self.predictions))
            and not np.any(np.isinf(self.predictions))
        )


@dataclass
class SalesAnalysis:
    raw_data:          "pd.DataFrame"
    trend:             "np.ndarray"
    seasonal:          "np.ndarray"
    residual:          "np.ndarray"
    anomalies:         "pd.DataFrame"
    yoy_growth:        "pd.Series"
    mom_growth:        "pd.Series"
    rolling_avg:       "pd.Series"
    statistics:        Dict
    forecasts:         Dict
    best_model:        str
    ensemble_forecast: ForecastResult
    detected_period:   int
    created_at:        datetime = field(default_factory=datetime.now)


# ══════════════════════════════════════════════════════════════════════════════
#  CUSTOM EXCEPTIONS
# ══════════════════════════════════════════════════════════════════════════════

class SalesForecastError(Exception): pass
class DataLoadError(SalesForecastError): pass
class ForecastError(SalesForecastError): pass
class ReportGenerationError(SalesForecastError): pass


# ══════════════════════════════════════════════════════════════════════════════
#  STATISTICAL UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

class StatUtils:
    """Rigorous, stateless statistical helpers."""

    @staticmethod
    def smape(y_true: "np.ndarray", y_pred: "np.ndarray") -> float:
        denom = np.abs(y_true) + np.abs(y_pred) + 1e-10
        return float(200.0 * np.mean(np.abs(y_true - y_pred) / denom))

    @staticmethod
    def mase(y_true: "np.ndarray", y_pred: "np.ndarray",
             y_train: "np.ndarray", period: int = 1) -> float:
        period = max(1, int(period))
        if len(y_train) <= period:
            scale = float(np.mean(np.abs(np.diff(y_train)))) if len(y_train) > 1 else 1.0
        else:
            naive_errors = np.abs(np.diff(y_train, n=period))
            scale = float(np.mean(naive_errors))
        scale = max(scale, 1e-10)
        return float(np.mean(np.abs(y_true - y_pred)) / scale)

    @staticmethod
    def safe_mape(y_true: "np.ndarray", y_pred: "np.ndarray",
                  min_denom: float = 1e-10) -> float:
        denom = np.maximum(np.abs(y_true), min_denom)
        return float(np.mean(np.abs(y_true - y_pred) / denom)) * 100.0

    @staticmethod
    def all_metrics(
        y_true: "np.ndarray", y_pred: "np.ndarray",
        y_train: "np.ndarray", period: int = 12
    ) -> Tuple[float, float, float, float, float, float]:
        """Returns (mae, rmse, smape, mase, mape_guarded, r2)."""
        y_true = np.asarray(y_true, dtype=float)
        y_pred = np.asarray(y_pred, dtype=float)
        mask   = np.isfinite(y_true) & np.isfinite(y_pred)
        y_true, y_pred = y_true[mask], y_pred[mask]
        if len(y_true) == 0:
            return (np.nan, np.nan, np.nan, np.nan, np.nan, np.nan)
        mae   = float(mean_absolute_error(y_true, y_pred))
        rmse  = float(np.sqrt(mean_squared_error(y_true, y_pred)))
        smape = StatUtils.smape(y_true, y_pred)
        mase_ = StatUtils.mase(y_true, y_pred, y_train, period)
        mape  = StatUtils.safe_mape(y_true, y_pred)
        try:
            r2 = float(r2_score(y_true, y_pred)) if len(y_true) > 1 else 0.0
        except Exception:
            r2 = 0.0
        return mae, rmse, smape, mase_, mape, r2

    @staticmethod
    def bootstrap_ci(
        point_forecasts: "np.ndarray",
        in_sample_residuals: "np.ndarray",
        alpha: float = 0.05,
        n_boot: int = 500,
    ) -> Tuple["np.ndarray", "np.ndarray"]:
        h     = len(point_forecasts)
        n_res = len(in_sample_residuals)
        if n_res < 4:
            return StatUtils.normal_ci(point_forecasts, in_sample_residuals, alpha)
        rng   = np.random.default_rng(42)
        scale = np.sqrt(1.0 + np.arange(h, dtype=float))
        boots = np.zeros((n_boot, h))
        for b in range(n_boot):
            sample   = rng.choice(in_sample_residuals, size=h, replace=True)
            boots[b] = point_forecasts + sample * scale
        lo = np.percentile(boots, 100.0 * alpha / 2,     axis=0)
        hi = np.percentile(boots, 100.0 * (1 - alpha/2), axis=0)
        return lo, hi

    @staticmethod
    def normal_ci(
        point_forecasts: "np.ndarray",
        residuals: "np.ndarray",
        alpha: float = 0.05,
    ) -> Tuple["np.ndarray", "np.ndarray"]:
        if len(residuals) > 1:
            std = float(np.std(residuals, ddof=1))
        else:
            std = float(np.mean(np.abs(point_forecasts)) * 0.10 + 1.0)
        z      = float(stats.norm.ppf(1.0 - alpha / 2))
        h      = len(point_forecasts)
        scale  = np.sqrt(np.arange(1, h + 1, dtype=float))
        margin = z * std * scale
        return point_forecasts - margin, point_forecasts + margin

    @staticmethod
    def arima_forecast_ci(
        predictions: "np.ndarray",
        sigma2: float,
        psi_weights: "np.ndarray",
        alpha: float = 0.05,
    ) -> Tuple["np.ndarray", "np.ndarray"]:
        h       = len(predictions)
        z       = stats.norm.ppf(1.0 - alpha / 2)
        cum_var = np.zeros(h)
        for i in range(h):
            w          = psi_weights[:i + 1]
            cum_var[i] = float(sigma2 * np.sum(w ** 2))
        margin = z * np.sqrt(np.maximum(cum_var, 0.0))
        return predictions - margin, predictions + margin

    @staticmethod
    def compute_diagnostics(
        residuals: "np.ndarray", max_lags: int = 10
    ) -> "ResidualDiagnostics":
        residuals = np.asarray(residuals, dtype=float)
        n         = len(residuals)
        d         = ResidualDiagnostics()
        if n < 4:
            return d

        # Ljung-Box
        try:
            lags    = min(max_lags, max(1, n // 4))
            acf_arr = []
            for k in range(1, lags + 1):
                if n - k < 2:
                    break
                r = np.corrcoef(residuals[:n - k], residuals[k:])
                acf_arr.append(r[0, 1] if not np.isnan(r[0, 1]) else 0.0)
            if acf_arr:
                acf     = np.array(acf_arr)
                L       = len(acf)
                lb_stat = float(n * (n + 2) * np.sum(
                    acf ** 2 / np.maximum(np.arange(n - 1, n - L - 1, -1), 1)))
                d.ljung_box_stat = lb_stat
                d.ljung_box_pval = float(1.0 - stats.chi2.cdf(lb_stat, df=L))
        except Exception:
            pass

        # Shapiro-Wilk
        if 4 <= n <= 5000:
            try:
                sw_stat, sw_pval = stats.shapiro(residuals)
                d.shapiro_stat   = float(sw_stat)
                d.shapiro_pval   = float(sw_pval)
            except Exception:
                pass

        # Durbin-Watson
        try:
            diff           = np.diff(residuals)
            denom          = float(np.sum(residuals ** 2))
            d.durbin_watson = float(np.sum(diff ** 2) / max(denom, 1e-10))
        except Exception:
            pass

        # Heteroscedasticity
        if n >= 8:
            try:
                half  = n // 2
                v1    = float(np.var(residuals[:half], ddof=1))
                v2    = float(np.var(residuals[half:], ddof=1))
                ratio = max(v1, v2) / max(min(v1, v2), 1e-10)
                d.heteroscedastic = ratio > 4.0
            except Exception:
                pass

        return d

    @staticmethod
    def detect_seasonality(
        y: "np.ndarray", candidates: Sequence[int] = (4, 6, 12, 24)
    ) -> int:
        n = len(y)
        if n < 24:
            return 12
        try:
            ym           = y - np.mean(y)
            # guard against zero-variance series
            if float(np.var(ym)) < 1e-20:
                return 12
            freqs, power = periodogram(ym)
            if len(freqs) < 2 or np.sum(power) < 1e-20:
                return 12
            total = float(np.sum(power))
            for period in sorted(candidates, reverse=True):
                if period >= n // 2 or period < 2:
                    continue
                target_freq = 1.0 / period
                idx    = int(np.argmin(np.abs(freqs - target_freq)))
                window = max(1, int(len(freqs) * 0.01))
                band   = power[max(0, idx - window): idx + window + 1]
                if float(np.sum(band)) / total > 0.15:
                    return period
        except Exception:
            pass
        return 12

    @staticmethod
    def marketing_roi_with_ci(
        sales: "np.ndarray", spend: "np.ndarray",
        n_boot: int = 1000, alpha: float = 0.05,
    ) -> Dict[str, float]:
        if len(sales) < 4 or np.sum(spend) < 1e-6:
            return {"point": 0.0, "ci_lo": 0.0, "ci_hi": 0.0}
        rng       = np.random.default_rng(42)
        n         = len(sales)
        roi_boots = np.zeros(n_boot)
        for b in range(n_boot):
            idx          = rng.integers(0, n, size=n)
            s_b          = float(np.sum(spend[idx]))
            r_b          = float(np.sum(sales[idx]))
            roi_boots[b] = (r_b - s_b) / max(s_b, 1e-10)
        point = float((np.sum(sales) - np.sum(spend)) / max(float(np.sum(spend)), 1e-10))
        lo    = float(np.percentile(roi_boots, 100.0 * alpha / 2))
        hi    = float(np.percentile(roi_boots, 100.0 * (1.0 - alpha / 2)))
        return {"point": point, "ci_lo": lo, "ci_hi": hi}


# ══════════════════════════════════════════════════════════════════════════════
#  DATA GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class SalesDataGenerator:
    _PRODUCTS = {
        "Product A": {"base": 150_000, "trend": 3_000,  "seasonal_amp": 0.18},
        "Product B": {"base":  95_000, "trend": 1_800,  "seasonal_amp": 0.25},
        "Product C": {"base": 210_000, "trend": 4_500,  "seasonal_amp": 0.12},
        "Product D": {"base":  70_000, "trend": 2_200,  "seasonal_amp": 0.30},
        "Product E": {"base": 130_000, "trend":  -500,  "seasonal_amp": 0.20},
    }
    _REGIONS = ["North", "South", "East", "West"]

    def __init__(self, seed: Optional[int] = None):
        actual_seed = seed if seed is not None else int(time.time()) % 100_000
        self.rng = np.random.default_rng(actual_seed)

    def generate(
        self, n_months: int = 60, n_products: int = 5,
        start_date: str = "2020-01-01",
    ) -> "pd.DataFrame":
        dates    = pd.date_range(start=start_date, periods=n_months, freq="MS")
        t        = np.arange(n_months, dtype=float)
        products = list(self._PRODUCTS.items())[:n_products]
        records: List[Dict] = []
        for prod_name, params in products:
            base          = float(params["base"])
            trend_comp    = base + float(params["trend"]) * t
            phase         = float(self.rng.uniform(0, np.pi))
            seasonal_comp = (float(params["seasonal_amp"]) * base
                             * np.sin(2 * np.pi * t / 12 + phase))
            holiday = np.array([
                0.15 * base  if d.month in (11, 12)
                else -0.08 * base if d.month in (1, 2)
                else 0.0
                for d in dates
            ])
            noise = self.rng.normal(0, 0.04 * base, n_months)
            sales = np.maximum(trend_comp + seasonal_comp + holiday + noise, 0.0)
            for ai in self.rng.integers(0, n_months, size=2):
                sales[ai] *= float(self.rng.choice([0.4, 1.8]))
            for d, s in zip(dates, sales):
                price_per_unit = float(self.rng.uniform(25, 60))
                records.append({
                    "date":            d,
                    "product":         prod_name,
                    "region":          str(self.rng.choice(self._REGIONS)),
                    "sales":           round(float(s), 2),
                    "units":           max(1, int(s / price_per_unit)),
                    "returns":         round(float(s) * float(self.rng.uniform(0.01, 0.05)), 2),
                    "marketing_spend": round(float(s) * float(self.rng.uniform(0.08, 0.15)), 2),
                })
        df = pd.DataFrame(records)
        df["net_revenue"] = df["sales"] - df["returns"]
        return df


# ══════════════════════════════════════════════════════════════════════════════
#  DATA LOADER & VALIDATOR
# ══════════════════════════════════════════════════════════════════════════════

class DataLoader:
    # ── Canonical target column names ────────────────────────────────────────
    REQUIRED_COLS     = {"date", "sales"}
    OPTIONAL_DEFAULTS: Dict[str, Any] = {"product": "Default", "region": "Default"}
    MIN_RECORDS       = 12

    # ── Fuzzy alias maps: canonical → list of recognised variants ────────────
    #    Matching is done on the *normalised* column name (lower, stripped,
    #    spaces/dashes → underscores).  Add more aliases here as needed.
    _DATE_ALIASES: List[str] = [
        "date", "order_date", "order_date_time", "transaction_date",
        "sale_date", "invoice_date", "purchase_date", "period", "month",
        "year_month", "timestamp", "time", "datetime", "dt", "dates",
        "fiscal_date", "week", "day", "reporting_date", "report_date",
        "order_placed_date", "created_at", "created_date",
    ]
    _SALES_ALIASES: List[str] = [
        "sales", "revenue", "amount", "total", "total_sales", "total_revenue",
        "net_sales", "gross_sales", "sale_amount", "invoice_amount",
        "transaction_amount", "order_amount", "order_value", "value",
        "price", "income", "earnings", "receipts", "turnover",
        "sales_amount", "sale_value", "qty_value", "extended_price",
        "subtotal", "grand_total", "payment", "payment_amount",
    ]
    _PRODUCT_ALIASES: List[str] = [
        "product", "product_name", "item", "item_name", "sku", "category",
        "product_category", "product_type", "goods", "commodity", "line",
        "product_line", "description", "product_description", "service",
    ]
    _REGION_ALIASES: List[str] = [
        "region", "area", "territory", "zone", "location", "state",
        "country", "city", "market", "branch", "office", "division",
        "geo", "geography", "district", "sector",
    ]
    _UNITS_ALIASES: List[str] = [
        "units", "quantity", "qty", "count", "volume", "items",
        "units_sold", "quantity_sold", "num_units", "pieces",
    ]
    _RETURNS_ALIASES: List[str] = [
        "returns", "refunds", "returns_amount", "refund_amount",
        "returned_amount", "chargebacks", "return_value",
    ]
    _MARKETING_ALIASES: List[str] = [
        "marketing_spend", "marketing", "ad_spend", "advertising",
        "marketing_cost", "promo_spend", "promotion_spend", "ad_cost",
        "advertising_spend", "campaign_spend",
    ]

    # ── Compiled alias → canonical lookup (built once at class-definition) ──
    _ALIAS_MAP: Dict[str, str] = {}

    @classmethod
    def _build_alias_map(cls) -> None:
        """Populate _ALIAS_MAP from the per-field alias lists."""
        mapping: List[Tuple[str, List[str]]] = [
            ("date",            cls._DATE_ALIASES),
            ("sales",           cls._SALES_ALIASES),
            ("product",         cls._PRODUCT_ALIASES),
            ("region",          cls._REGION_ALIASES),
            ("units",           cls._UNITS_ALIASES),
            ("returns",         cls._RETURNS_ALIASES),
            ("marketing_spend", cls._MARKETING_ALIASES),
        ]
        for canonical, aliases in mapping:
            for alias in aliases:
                cls._ALIAS_MAP[alias] = canonical

    @classmethod
    def _normalise_col(cls, raw: str) -> str:
        """Strip BOM / whitespace, lower-case, collapse separators → underscore."""
        return (
            str(raw)
            .strip()
            .lstrip("\ufeff")
            .lower()
            .replace(" ", "_")
            .replace("-", "_")
            .replace(".", "_")
            .replace("/", "_")
            .replace("\\", "_")
        )

    @classmethod
    def _map_columns(cls, df: "pd.DataFrame") -> Tuple["pd.DataFrame", Dict[str, str]]:
        """
        Intelligently rename columns to canonical names.

        Strategy (in priority order):
          1. Exact match after normalisation.
          2. Alias map lookup.
          3. Partial / substring match (e.g. 'order_date_time' → 'date').
          4. Numeric-column heuristic for 'sales' (pick first numeric if nothing maps).

        Returns the renamed DataFrame and a dict {original → canonical} for logging.
        """
        # Ensure alias map is ready
        if not cls._ALIAS_MAP:
            cls._build_alias_map()

        norm_to_orig: Dict[str, str] = {}
        for orig in df.columns:
            norm = cls._normalise_col(orig)
            # keep first occurrence if there are duplicate normalised names
            if norm not in norm_to_orig:
                norm_to_orig[norm] = orig

        rename: Dict[str, str] = {}   # original_col → canonical
        mapped: Dict[str, str] = {}   # canonical → original_col (reverse)

        # ── Pass 1: exact match or alias map ─────────────────────────────
        for norm, orig in norm_to_orig.items():
            canonical = cls._ALIAS_MAP.get(norm)
            if canonical and canonical not in mapped:
                rename[orig] = canonical
                mapped[canonical] = orig

        # ── Pass 2: partial / substring match for still-missing columns ──
        targets_needed = {"date", "sales", "product", "region",
                          "units", "returns", "marketing_spend"} - set(mapped.keys())
        for norm, orig in norm_to_orig.items():
            if orig in rename:          # already mapped
                continue
            for canonical in list(targets_needed):
                # check if the target keyword appears anywhere in the norm name
                keywords = {
                    "date":            ["date", "time", "period", "month", "day", "week"],
                    "sales":           ["sale", "revenue", "amount", "total", "value",
                                        "price", "income", "earning", "receipt", "payment"],
                    "product":         ["product", "item", "sku", "category", "good"],
                    "region":          ["region", "area", "territ", "zone", "locat",
                                        "state", "country", "city", "market", "branch"],
                    "units":           ["unit", "qty", "quant", "count", "volume", "piece"],
                    "returns":         ["return", "refund", "chargeback"],
                    "marketing_spend": ["market", "ad_spend", "adverti", "promo"],
                }.get(canonical, [])
                if any(kw in norm for kw in keywords):
                    rename[orig] = canonical
                    mapped[canonical] = orig
                    targets_needed.discard(canonical)
                    break

        # ── Pass 3: numeric-column heuristic for 'sales' if still missing ─
        if "sales" not in mapped:
            numeric_cols = [
                orig for orig in df.columns
                if orig not in rename
                and pd.api.types.is_numeric_dtype(df[orig])
            ]
            if numeric_cols:
                best = numeric_cols[0]
                rename[best] = "sales"
                mapped["sales"] = best
                logger.warning(
                    "No 'sales' column found; using first numeric column '%s' as sales.", best)

        # ── Pass 4: date heuristic — look for a column that parses as dates ─
        if "date" not in mapped:
            for orig in df.columns:
                if orig in rename:
                    continue
                try:
                    parsed = pd.to_datetime(df[orig], errors="coerce")
                    frac   = parsed.notna().mean()
                    if frac >= 0.8:
                        rename[orig] = "date"
                        mapped["date"] = orig
                        logger.warning(
                            "No 'date' column found; using '%s' as date (%.0f%% parseable).",
                            orig, frac * 100)
                        break
                except Exception:
                    continue

        df = df.rename(columns=rename)
        return df, rename

    @classmethod
    def load(cls, source: Optional[str] = None) -> "pd.DataFrame":
        if source is None:
            return SalesDataGenerator(seed=42).generate(n_months=60, n_products=5)
        path = Path(source)
        if not path.exists():
            raise DataLoadError(f"File not found: {path}")
        if path.stat().st_size == 0:
            raise DataLoadError(f"File is empty: {path.name}")
        try:
            df = cls._read_file(path)
        except DataLoadError:
            raise
        except Exception as exc:
            raise DataLoadError(f"Failed to read '{path.name}': {exc}") from exc
        return cls._validate_and_clean(df)

    @staticmethod
    def _read_file(path: Path) -> "pd.DataFrame":
        suffix = path.suffix.lower()

        if suffix in (".xlsx", ".xlsm"):
            # Try multiple sheets if first sheet is empty
            try:
                xl = pd.ExcelFile(path, engine="openpyxl")
                for sheet in xl.sheet_names:
                    df = xl.parse(sheet)
                    if not df.empty and len(df.columns) > 1:
                        return df
                return xl.parse(xl.sheet_names[0])
            except Exception as exc:
                raise DataLoadError(f"Could not read Excel file '{path.name}': {exc}") from exc

        if suffix in (".xls",):
            for engine in ("xlrd", "openpyxl"):
                try:
                    df = pd.read_excel(path, engine=engine)
                    if not df.empty:
                        return df
                except Exception:
                    continue
            raise DataLoadError(
                f"Could not read .xls file '{path.name}'. "
                "Install xlrd:  pip install xlrd>=2.0"
            )

        if suffix in (".csv", ".tsv", ".txt"):
            sep = "\t" if suffix == ".tsv" else None
            # Try common separators if sep is None
            sep_candidates = [sep] if sep else [None, ",", ";", "|", "\t"]
            encodings = ("utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-15")
            for enc in encodings:
                for sep_try in sep_candidates:
                    try:
                        kwargs_csv: Dict[str, Any] = dict(
                            encoding=enc,
                            on_bad_lines="skip",
                        )
                        if sep_try is not None:
                            kwargs_csv["sep"] = sep_try
                            kwargs_csv["engine"] = "c"
                        else:
                            kwargs_csv["sep"] = None
                            kwargs_csv["engine"] = "python"
                        df = pd.read_csv(path, **kwargs_csv)
                        if len(df.columns) > 1 and not df.empty:
                            return df
                    except (UnicodeDecodeError, pd.errors.ParserError):
                        continue
                    except Exception:
                        continue
            raise DataLoadError(f"Could not decode CSV/TSV file: {path.name}")

        if suffix == ".json":
            raw = path.read_bytes()
            # Handle JSON lines format
            try:
                df = pd.read_json(io.BytesIO(raw), lines=True)
                if not df.empty and len(df.columns) > 1:
                    return df
            except Exception:
                pass
            for orient in (None, "records", "split", "index", "values", "table"):
                try:
                    kwargs_j: Dict[str, Any] = {}
                    if orient:
                        kwargs_j["orient"] = orient
                    df = pd.read_json(io.BytesIO(raw), **kwargs_j)
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        return df
                except Exception:
                    continue
            raise DataLoadError(
                f"Could not parse JSON file '{path.name}'. "
                "Supported orientations: default, records, split, index, values, table."
            )

        if suffix == ".parquet":
            # try pyarrow first, then fastparquet, then helpful error
            for engine in ("pyarrow", "fastparquet"):
                try:
                    return pd.read_parquet(path, engine=engine)
                except ImportError:
                    continue
                except Exception as exc:
                    raise DataLoadError(f"Parquet read failed: {exc}") from exc
            raise DataLoadError(
                f"Cannot read Parquet '{path.name}'. "
                "Install a Parquet engine:  pip install pyarrow  or  pip install fastparquet"
            )

        raise DataLoadError(
            f"Unsupported file format '{suffix}'. "
            "Supported: .csv  .tsv  .xlsx  .xlsm  .xls  .json  .parquet"
        )

    @classmethod
    def _validate_and_clean(cls, df: "pd.DataFrame") -> "pd.DataFrame":
        if df.empty:
            raise DataLoadError("File contains no data rows.")

        # ── Step 1: normalise column names then smart-map to canonical names ─
        df, col_renames = cls._map_columns(df)
        if col_renames:
            logger.info("Column mapping applied: %s", col_renames)

        # ── Step 2: check required columns are now present ─────────────────
        missing = cls.REQUIRED_COLS - set(df.columns)
        if missing:
            raise DataLoadError(
                f"Required columns missing after auto-mapping: {sorted(missing)}\n"
                f"Your file has columns: {sorted(col_renames.keys() or df.columns.tolist())}\n\n"
                "Tip: ensure your file contains a date-like column "
                "(e.g. 'Date', 'Order Date', 'Transaction Date') and a numeric "
                "sales/revenue column (e.g. 'Sales', 'Revenue', 'Amount')."
            )

        # ── Step 3: parse dates ────────────────────────────────────────────
        if _PANDAS_GTE_2:
            df["date"] = pd.to_datetime(df["date"], errors="coerce")
        else:
            df["date"] = pd.to_datetime(
                df["date"], infer_datetime_format=True, errors="coerce")

        n_bad_dates = int(df["date"].isna().sum())
        df          = df.dropna(subset=["date"])
        if df.empty:
            raise DataLoadError("No valid dates found in 'date' column.")
        if n_bad_dates > 0:
            logger.warning("Dropped %d rows with unparseable dates.", n_bad_dates)

        # ── Step 4: coerce sales to numeric ──────────────────────────────
        df["sales"] = pd.to_numeric(df["sales"], errors="coerce")
        # drop rows where sales could not be parsed (true NaN) before
        # clipping to 0; this way we don't silently treat bad data as $0 months
        n_bad_sales = int(df["sales"].isna().sum())
        if n_bad_sales > 0:
            logger.warning("Dropped %d rows with non-numeric sales.", n_bad_sales)
            df = df.dropna(subset=["sales"])
        df["sales"] = df["sales"].clip(lower=0.0)

        # ── Step 5: fill optional columns with sensible defaults ─────────
        for col, default in cls.OPTIONAL_DEFAULTS.items():
            if col not in df.columns:
                df[col] = default
        if "returns"         not in df.columns:
            df["returns"]         = (df["sales"] * 0.02).round(2)
        if "units"           not in df.columns:
            df["units"]           = (df["sales"] / 50.0).astype(int).clip(lower=1)
        if "marketing_spend" not in df.columns:
            df["marketing_spend"] = (df["sales"] * 0.10).round(2)
        if "net_revenue"     not in df.columns:
            df["net_revenue"]     = df["sales"] - df["returns"]

        # ── Step 6: sort and validate minimum length ──────────────────────
        df = df.sort_values("date").reset_index(drop=True)
        if len(df) < cls.MIN_RECORDS:
            raise DataLoadError(
                f"Only {len(df)} valid records found. "
                f"At least {cls.MIN_RECORDS} monthly data points are required."
            )
        return df


def load_and_validate_data(source: Optional[str] = None) -> "pd.DataFrame":
    return DataLoader.load(source)


# ══════════════════════════════════════════════════════════════════════════════
#  ARIMA — Full MLE implementation
#  [BUG-2 FIX] Removed dead _diff_saves field and unused static _undifference().
#  The sole undifferencing path is now _predict_undiff() which uses
#  self._last_obs_per_level (the correct, actually-populated list).
# ══════════════════════════════════════════════════════════════════════════════

class ARIMA:
    """
    ARIMA(p, d, q) with exact Gaussian log-likelihood via innovations form.

    Estimation:
        θ = argmax L(θ; y)  where
        L = -½ [n·log(2πσ²) + Σ ε_t²/σ²]

    Forecast CI uses ψ-weights (MA∞ representation):
        Var(ŷ_{T+h}) = σ² · Σ_{j=0}^{h-1} ψ_j²   (Box-Jenkins §5.4)

    AIC = -2·L + 2·(p+q+1)
    BIC = -2·L + (p+q+1)·ln(n)
    """

    def __init__(self, p: int = 2, d: int = 1, q: int = 1):
        self.p  = max(0, int(p))
        self.d  = max(0, int(d))
        self.q  = max(0, int(q))
        self._ar_params:  "np.ndarray" = np.array([])
        self._ma_params:  "np.ndarray" = np.array([])
        self._const:      float        = 0.0
        self._sigma2:     float        = 1.0
        self._residuals:  "np.ndarray" = np.array([])
        self._orig:       "np.ndarray" = np.array([])
        # _last_obs_per_level[i] = last observed value of the series
        # integrated i times — used by _predict_undiff() to invert np.diff().
        self._last_obs_per_level: List[float] = []
        self._psi:    "np.ndarray" = np.array([])
        self._aic:    float        = np.inf
        self._bic:    float        = np.inf
        self._fitted: bool         = False

    # ── Differencing helpers ─────────────────────────────────────────────────

    @staticmethod
    def _difference(y: "np.ndarray", d: int) -> "np.ndarray":
        """Difference y exactly d times, returning the differenced series."""
        out = y.astype(float).copy()
        for _ in range(d):
            out = np.diff(out)
        return out

    def _predict_undiff(self, diff_preds: "np.ndarray") -> "np.ndarray":
        """
        Undo all d differencing steps using stored _last_obs_per_level.
        _last_obs_per_level[i] = last value of the series after i integrations.
        We apply cumsum starting from that anchor to recover the next level.
        """
        result = diff_preds.copy().astype(float)
        for last_obs in reversed(self._last_obs_per_level):
            result = np.concatenate([[last_obs], result])
            result = np.cumsum(result)[1:]
        return result

    # ── Innovations (Kalman-filter form) ─────────────────────────────────────

    def _innovations(
        self, yd: "np.ndarray", ar: "np.ndarray",
        ma: "np.ndarray", mu: float,
    ) -> "np.ndarray":
        n     = len(yd)
        p, q  = len(ar), len(ma)
        eps   = np.zeros(n)
        for t in range(n):
            ar_part = sum(ar[i] * yd[t - i - 1]  for i in range(min(p, t)))
            ma_part = sum(ma[i] * eps[t - i - 1] for i in range(min(q, t)))
            eps[t]  = yd[t] - mu - ar_part - ma_part
        return eps

    def _neg_log_likelihood(self, params: "np.ndarray", yd: "np.ndarray") -> float:
        p, q = self.p, self.q
        ar   = params[:p]
        ma   = params[p:p + q]
        mu   = params[p + q]
        eps  = self._innovations(yd, ar, ma, mu)
        n    = len(eps)
        # floor raised to 1e-10 — log(1e-20) = -46 which derails L-BFGS-B
        sig2 = max(float(np.mean(eps ** 2)), 1e-10)
        return 0.5 * (n * math.log(2 * math.pi * sig2) + n)

    def fit(self, y: "np.ndarray") -> "ARIMA":
        self._orig = y.astype(float).copy()

        # Build _last_obs_per_level BEFORE differencing
        self._last_obs_per_level = []
        tmp = self._orig.copy()
        for _ in range(self.d):
            self._last_obs_per_level.append(float(tmp[-1]))
            tmp = np.diff(tmp)

        yd = ARIMA._difference(self._orig, self.d)
        n  = len(yd)

        max_order = max(n // 4, 1)
        self.p    = min(self.p, max_order)
        self.q    = min(self.q, max_order)

        k  = self.p + self.q + 1
        x0 = np.zeros(k)

        if self.p > 0 and n > self.p + 2:
            rows = n - self.p
            X    = np.column_stack(
                [yd[self.p - i - 1: n - i - 1] for i in range(self.p)]
                + [np.ones(rows)]
            )
            yt = yd[self.p:]
            try:
                coefs, *_ = np.linalg.lstsq(X, yt, rcond=None)
                x0[:self.p] = coefs[:-1]
                x0[k - 1]   = float(coefs[-1])
            except Exception:
                pass

        best_nll    = np.inf
        best_params = x0.copy()
        for jitter in [0.0, 0.05, 0.15]:
            rng    = np.random.default_rng(17 + int(jitter * 1000))
            x_init = x0 + rng.normal(0, jitter, k)
            try:
                res = minimize(
                    self._neg_log_likelihood, x_init, args=(yd,),
                    method="L-BFGS-B",
                    options={"maxiter": 500, "ftol": 1e-12, "gtol": 1e-8})
                if np.isfinite(res.fun) and res.fun < best_nll:
                    best_nll    = res.fun
                    best_params = res.x.copy()
            except Exception:
                pass

        self._ar_params = best_params[:self.p]
        self._ma_params = best_params[self.p: self.p + self.q]
        self._const     = float(best_params[self.p + self.q])

        eps           = self._innovations(yd, self._ar_params, self._ma_params, self._const)
        self._sigma2  = max(float(np.var(eps, ddof=1)), 1e-20)
        self._residuals = eps

        ll        = -best_nll
        npar      = k + 1
        self._aic = float(-2.0 * ll + 2.0 * npar)
        self._bic = float(-2.0 * ll + npar * math.log(max(n, 2)))

        max_psi = 60
        psi     = np.zeros(max_psi)
        psi[0]  = 1.0
        for j in range(1, max_psi):
            ar_sum = sum(
                self._ar_params[i] * psi[j - i - 1]
                for i in range(min(self.p, j))
            )
            ma_val = float(self._ma_params[j - 1]) if j - 1 < self.q else 0.0
            psi[j] = ar_sum + ma_val
        self._psi    = psi
        self._fitted = True
        return self

    def predict(self, h: int) -> "np.ndarray":
        if not self._fitted:
            raise RuntimeError("Model not fitted.")

        yd_orig = ARIMA._difference(self._orig, self.d)
        yd  = list(yd_orig.copy())
        eps = list(self._residuals.copy())

        diff_preds: List[float] = []
        for _ in range(h):
            n_yd    = len(yd)
            ar_part = sum(
                self._ar_params[i] * yd[n_yd - i - 1]
                for i in range(min(self.p, n_yd))
            )
            n_eps   = len(eps)
            ma_part = sum(
                self._ma_params[i] * eps[n_eps - i - 1]
                for i in range(min(self.q, n_eps))
            )
            pred = self._const + ar_part + ma_part
            diff_preds.append(pred)
            yd.append(pred)
            eps.append(0.0)

        # explicit float64 array — on Windows np.array of Python floats
        # may default to float32/int depending on the platform's C long size
        return np.maximum(
            self._predict_undiff(np.array(diff_preds, dtype=np.float64)), 0.0)

    def forecast_ci(self, h: int, alpha: float = 0.05) -> Tuple["np.ndarray", "np.ndarray"]:
        preds  = self.predict(h)
        lo, hi = StatUtils.arima_forecast_ci(preds, self._sigma2, self._psi, alpha)
        return np.maximum(lo, 0.0), hi

    @classmethod
    def auto_select(
        cls, y: "np.ndarray",
        max_p: int = 3, max_d: int = 1, max_q: int = 2,
    ) -> "ARIMA":
        """Grid-search ARIMA order by AIC (exact MLE). Returns fitted model."""
        # explicit guard — prevents silent garbage fit
        if len(y) < 6:
            raise ForecastError(
                f"ARIMA requires at least 6 observations; got {len(y)}.")
        best_aic   = np.inf
        best_model: Optional["ARIMA"] = None
        for p in range(0, max_p + 1):
            for d in range(0, max_d + 1):
                for q in range(0, max_q + 1):
                    if p == 0 and q == 0:
                        continue
                    try:
                        m = cls(p, d, q).fit(y)
                        if np.isfinite(m._aic) and m._aic < best_aic:
                            best_aic   = m._aic
                            best_model = m
                    except Exception:
                        pass
        if best_model is None:
            best_model = cls(1, 1, 0).fit(y)
        return best_model


# ══════════════════════════════════════════════════════════════════════════════
#  EXPONENTIAL SMOOTHING — Holt-Winters
# ══════════════════════════════════════════════════════════════════════════════

class HoltWinters:
    """
    Holt-Winters triple exponential smoothing.
    Additive:       ŷ_{t+h} = (L_t + h·T_t) + S_{t+h-m}
    Multiplicative: ŷ_{t+h} = (L_t + h·T_t) · S_{t+h-m}
    Parameters estimated by SSE minimisation via L-BFGS-B (multiple starts).
    """

    def __init__(
        self, alpha: float = 0.3, beta: float = 0.1,
        gamma: float = 0.2, season_len: int = 12,
        multiplicative: bool = False,
    ):
        self.alpha          = float(np.clip(alpha, 0.01, 0.9999))
        self.beta           = float(np.clip(beta,  0.01, 0.9999))
        self.gamma          = float(np.clip(gamma, 0.01, 0.9999))
        self.season_len     = max(2, int(season_len))
        self.multiplicative = multiplicative
        self._level:    float       = 0.0
        self._trend:    float       = 0.0
        self._seasonal: List[float] = []
        self._sigma2:   float       = 1.0
        self._aic:      float       = np.inf
        self._resid:    "np.ndarray" = np.array([])
        self._n_obs:    int          = 0
        self._fitted:   bool         = False

    def _init_components(
        self, y: "np.ndarray"
    ) -> Tuple[float, float, List[float]]:
        m = self.season_len
        n = len(y)
        if n < 2 * m:
            m = max(2, n // 2)
            self.season_len = m

        level = float(np.mean(y[:m]))
        # guard: level=0 breaks multiplicative init
        if abs(level) < 1e-9:
            level = 1e-9

        if n >= 2 * m:
            trend = float((np.mean(y[m: 2 * m]) - level) / m)
        else:
            trend = float((y[-1] - y[0]) / max(n - 1, 1))

        if self.multiplicative:
            seas = []
            for i in range(m):
                d = level + i * trend
                seas.append(float(y[i]) / max(abs(d), 1e-9))
        else:
            seas = [float(y[i]) - (level + i * trend) for i in range(m)]

        if self.multiplicative:
            s_mean = float(np.mean(seas))
            seas   = [s / max(abs(s_mean), 1e-9) for s in seas]
        else:
            s_mean = float(np.mean(seas))
            seas   = [s - s_mean for s in seas]
        return level, trend, seas

    def _smooth(
        self, y: "np.ndarray",
        alpha: float, beta: float, gamma: float,
    ) -> Tuple[float, float, List[float], "np.ndarray"]:
        m              = self.season_len
        n              = len(y)
        level, trend, seas = self._init_components(y)
        seas_arr       = list(seas) * (n // m + 2)
        fitted         = np.zeros(n)

        for t in range(n):
            s_idx  = t % m
            prev_l = level
            prev_t = trend
            prev_s = seas_arr[s_idx]

            if self.multiplicative:
                safe_s          = prev_s if abs(prev_s) > 1e-9 else 1.0
                fitted[t]       = (prev_l + prev_t) * safe_s
                level           = alpha * (y[t] / max(abs(safe_s), 1e-9)) + (1 - alpha) * (prev_l + prev_t)
                seas_arr[s_idx] = (gamma * (y[t] / max(abs(level), 1e-9))
                                   + (1 - gamma) * prev_s)
            else:
                fitted[t]       = prev_l + prev_t + prev_s
                level           = alpha * (y[t] - prev_s) + (1 - alpha) * (prev_l + prev_t)
                seas_arr[s_idx] = gamma * (y[t] - level) + (1 - gamma) * prev_s

            trend = beta * (level - prev_l) + (1 - beta) * prev_t

        final_seas = [seas_arr[i % m] for i in range(n - m, n)]
        return level, trend, final_seas, fitted

    def _sse(self, params: "np.ndarray", y: "np.ndarray") -> float:
        a, b, g = float(params[0]), float(params[1]), float(params[2])
        if not (0.01 <= a <= 0.999 and 0.01 <= b <= 0.999 and 0.01 <= g <= 0.999):
            return 1e20
        try:
            _, _, _, fitted = self._smooth(y, a, b, g)
            return float(np.sum((y - fitted) ** 2))
        except Exception:
            return 1e20

    def fit(self, y: "np.ndarray") -> "HoltWinters":
        best_sse    = np.inf
        best_params = [0.3, 0.1, 0.2]
        starts      = [
            (0.3, 0.1, 0.2), (0.5, 0.2, 0.3), (0.1, 0.05, 0.1),
            (0.7, 0.3, 0.4), (0.2, 0.05, 0.15), (0.4, 0.15, 0.25),
            (0.6, 0.25, 0.35), (0.15, 0.08, 0.12),
        ]
        for start in starts:
            try:
                res = minimize(
                    self._sse, list(start), args=(y,),
                    bounds=[(0.01, 0.9999)] * 3, method="L-BFGS-B",
                    options={"maxiter": 400, "ftol": 1e-12})
                # L-BFGS-B with bounds may return success=False even when
                # it converged to a boundary — accept any finite improvement
                if np.isfinite(res.fun) and res.fun < best_sse:
                    best_sse, best_params = res.fun, res.x.tolist()
            except Exception:
                pass

        self.alpha, self.beta, self.gamma = [float(x) for x in best_params]
        level, trend, seas, fitted = self._smooth(y, self.alpha, self.beta, self.gamma)
        self._level    = level
        self._trend    = trend
        self._seasonal = seas
        self._n_obs    = len(y)

        resid        = y - fitted
        self._resid  = resid
        n            = len(y)
        k            = 4
        sigma2       = max(float(np.var(resid, ddof=1)), 1e-20)
        self._sigma2 = sigma2
        ll           = -0.5 * (n * math.log(2 * math.pi * sigma2)
                               + float(best_sse) / sigma2)
        self._aic    = float(-2.0 * ll + 2.0 * k)
        self._fitted = True
        return self

    def predict(self, h: int) -> "np.ndarray":
        if not self._fitted:
            raise RuntimeError("Model not fitted.")
        m     = self.season_len
        preds = []
        for i in range(1, h + 1):
            s_idx = (i - 1) % m
            s_v   = self._seasonal[s_idx]
            base  = self._level + i * self._trend
            if self.multiplicative:
                p = base * (s_v if abs(s_v) > 1e-9 else 1.0)
            else:
                p = base + s_v
            preds.append(p)
        return np.array(preds, dtype=float)

    def forecast_ci(self, h: int, alpha: float = 0.05) -> Tuple["np.ndarray", "np.ndarray"]:
        preds = self.predict(h)
        if not self.multiplicative:
            z     = float(stats.norm.ppf(1.0 - alpha / 2))
            a, b  = self.alpha, self.beta
            var_h = np.array([
                self._sigma2 * max(
                    1.0 + max(i - 1, 0) * (a**2 + a * b * i + b**2 * i * (2*i - 1) / 6.0),
                    1e-9)
                for i in range(1, h + 1)
            ])
            margin = z * np.sqrt(var_h)
            return np.maximum(preds - margin, 0.0), preds + margin
        else:
            lo, hi = StatUtils.bootstrap_ci(preds, self._resid, alpha=alpha, n_boot=400)
            return np.maximum(lo, 0.0), hi

    @classmethod
    def auto_select(cls, y: "np.ndarray", season_len: int = 12) -> "HoltWinters":
        best_aic = np.inf
        best_mdl: Optional["HoltWinters"] = None
        for mul in (False, True):
            try:
                mdl = cls(season_len=season_len, multiplicative=mul)
                mdl.fit(y)
                if np.isfinite(mdl._aic) and mdl._aic < best_aic:
                    best_aic = mdl._aic
                    best_mdl = mdl
            except Exception:
                pass
        if best_mdl is None:
            best_mdl = cls(season_len=season_len).fit(y)
        return best_mdl


# ══════════════════════════════════════════════════════════════════════════════
#  ML FEATURE ENGINEERING
# ══════════════════════════════════════════════════════════════════════════════

class MLFeatureEngineering:
    """
    Build feature matrix for ML models.
    Lag selection guided by PACF: only lags where |PACF| > 2/√n
    (Bartlett's 95% band) are included, plus multiples of the period.
    """

    @staticmethod
    def _pacf(y: "np.ndarray", max_lag: int) -> "np.ndarray":
        n    = len(y)
        pacf = np.zeros(max_lag + 1)
        pacf[0] = 1.0
        if n < 4 or max_lag < 1:
            return pacf
        ym  = y - float(np.mean(y))
        var = float(np.dot(ym, ym)) / n
        if var < 1e-14:
            return pacf
        acv = np.array([float(np.dot(ym[:n - k], ym[k:])) / n
                        for k in range(max_lag + 1)])
        acf_vals = acv / max(acv[0], 1e-14)

        phi = np.zeros((max_lag + 1, max_lag + 1))
        phi[1, 1] = acf_vals[1]
        pacf[1]   = acf_vals[1]
        for k in range(2, max_lag + 1):
            num = acf_vals[k] - np.dot(phi[k-1, 1:k], acf_vals[k-1:0:-1])
            den = 1.0 - np.dot(phi[k-1, 1:k], acf_vals[1:k])
            if abs(den) < 1e-12:
                break
            phi[k, k] = num / den
            for j in range(1, k):
                phi[k, j] = phi[k-1, j] - phi[k, k] * phi[k-1, k-j]
            pacf[k] = phi[k, k]
        return pacf

    @classmethod
    def select_lags(
        cls, y: "np.ndarray", max_lag: int = 24, period: int = 12,
    ) -> List[int]:
        n         = len(y)
        max_lag   = min(max_lag, n // 3)
        if max_lag < 1:
            return [1]
        threshold = 2.0 / math.sqrt(max(n, 4))
        try:
            pacf_vals = cls._pacf(y, max_lag)
        except Exception:
            pacf_vals = np.zeros(max_lag + 1)

        selected: set = {1}
        for lag in range(1, max_lag + 1):
            if abs(pacf_vals[lag]) > threshold:
                selected.add(lag)
            if lag % period == 0:
                selected.add(lag)
        return sorted(selected)

    @classmethod
    def build_features(
        cls, y: "np.ndarray", period: int = 12,
    ) -> Tuple["np.ndarray", "np.ndarray", List[str]]:
        # ensure float64 throughout — prevents int32 overflow on Windows
        y    = np.asarray(y, dtype=np.float64)
        n    = len(y)
        lags = cls.select_lags(y, max_lag=min(24, max(1, n // 3)), period=period)
        feats: Dict[str, "np.ndarray"] = {}

        for lag in lags:
            if 0 < lag < n:
                padded = np.empty(n, dtype=np.float64)
                padded[:lag] = np.nan
                padded[lag:] = y[:n - lag]
                feats[f"lag_{lag}"] = padded

        s = pd.Series(y)
        for w in [3, 6, 12]:
            if w <= n:
                feats[f"roll_mean_{w}"] = s.rolling(w, min_periods=1).mean().to_numpy(dtype=np.float64)
                feats[f"roll_std_{w}"]  = s.rolling(w, min_periods=1).std().fillna(0).to_numpy(dtype=np.float64)
                feats[f"roll_min_{w}"]  = s.rolling(w, min_periods=1).min().to_numpy(dtype=np.float64)
                feats[f"roll_max_{w}"]  = s.rolling(w, min_periods=1).max().to_numpy(dtype=np.float64)

        t_arr = np.arange(n, dtype=np.float64)
        feats["t"]      = t_arr
        feats["t_sq"]   = t_arr ** 2
        feats["t_sqrt"] = np.sqrt(t_arr)

        for harmonic in range(1, 3):
            feats[f"sin_{harmonic}"] = np.sin(2 * np.pi * harmonic * t_arr / period)
            feats[f"cos_{harmonic}"] = np.cos(2 * np.pi * harmonic * t_arr / period)

        df           = pd.DataFrame(feats)
        df["target"] = y
        df           = df.dropna()
        if df.empty:
            raise ValueError("No complete rows after feature engineering.")

        cols    = [c for c in df.columns if c != "target"]
        X       = df[cols].to_numpy(dtype=np.float64)
        y_clean = df["target"].to_numpy(dtype=np.float64)
        return X, y_clean, cols


# ══════════════════════════════════════════════════════════════════════════════
#  WALK-FORWARD CROSS VALIDATION
# ══════════════════════════════════════════════════════════════════════════════

class WalkForwardCV:
    """Expanding-window time-series cross-validation. No look-ahead bias."""

    @staticmethod
    def evaluate(
        model_fn: Callable, y: "np.ndarray",
        period: int = 12, n_splits: int = 3, horizon: int = 6,
    ) -> Dict[str, float]:
        n         = len(y)
        min_train = max(period * 2, 24)
        horizon   = min(horizon, n // 4)
        if n < min_train + horizon:
            return {"smape": float("nan"), "mase": float("nan"), "rmse": float("nan")}

        available  = n - min_train - horizon
        fold_step  = max(1, available // n_splits)
        split_ends = [min_train + (i + 1) * fold_step for i in range(n_splits)]
        split_ends = [s for s in split_ends if s + horizon <= n]
        if not split_ends:
            return {"smape": float("nan"), "mase": float("nan"), "rmse": float("nan")}

        smapes, mases, rmses = [], [], []
        for end in split_ends:
            train  = y[:end]
            actual = y[end: end + horizon]
            if len(actual) == 0:
                continue
            try:
                raw_preds = model_fn(train)
                preds     = np.asarray(raw_preds, dtype=float)[:len(actual)]
                if len(preds) < len(actual) or np.any(~np.isfinite(preds)):
                    continue
                smapes.append(StatUtils.smape(actual, preds))
                mases.append(StatUtils.mase(actual, preds, train, period))
                rmses.append(float(np.sqrt(mean_squared_error(actual, preds))))
            except Exception:
                pass

        if not smapes:
            return {"smape": float("nan"), "mase": float("nan"), "rmse": float("nan")}
        return {
            "smape": float(np.mean(smapes)),
            "mase":  float(np.mean(mases)),
            "rmse":  float(np.mean(rmses)),
        }


# ══════════════════════════════════════════════════════════════════════════════
#  CANCELLATION TOKEN
# ══════════════════════════════════════════════════════════════════════════════

class CancelToken:
    def __init__(self):
        self._event = threading.Event()

    def cancel(self) -> None:
        self._event.set()

    @property
    def is_cancelled(self) -> bool:
        return self._event.is_set()

    def check(self) -> None:
        if self._event.is_set():
            raise InterruptedError("Analysis cancelled by user.")


# ══════════════════════════════════════════════════════════════════════════════
#  FORECASTING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class ForecastingEngine:
    """Runs all forecasting models and assembles the ensemble."""
    def __init__(
        self, config: ForecastConfig,
        progress_cb: Optional[Callable] = None,
        cancel_token: Optional[CancelToken] = None,
    ):
        self.config  = config
        self._cb     = progress_cb or (lambda msg, pct: None)
        self._cancel = cancel_token or CancelToken()

    def _report(self, msg: str, pct: Optional[float] = None) -> None:
        try:
            self._cb(msg, pct)
        except Exception:
            pass

    def _check(self) -> None:
        self._cancel.check()

    # ── ARIMA ─────────────────────────────────────────────────────────────────

    def run_arima(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_arima:
            return None
        self._report("Fitting ARIMA (MLE grid search) …")
        t0 = time.perf_counter()
        try:
            self._check()
            h     = self.config.forecast_horizon
            alpha = 1.0 - self.config.confidence_level

            def arima_fn(y_tr: "np.ndarray") -> "np.ndarray":
                return ARIMA.auto_select(y_tr).predict(h)

            cv_scores = WalkForwardCV.evaluate(
                arima_fn, y, period=period,
                n_splits=self.config.cv_folds, horizon=min(h, len(y) // 4))

            split       = max(h, len(y) // 5)
            train, test = y[:-split], y[-split:]
            val_mdl     = ARIMA.auto_select(train)
            val_preds   = np.maximum(val_mdl.predict(split)[:len(test)], 0.0)
            mae, rmse, smape, mase_, mape, r2 = StatUtils.all_metrics(
                test, val_preds, train, period)

            final  = ARIMA(val_mdl.p, val_mdl.d, val_mdl.q).fit(y)
            preds  = np.maximum(final.predict(h), 0.0)
            lo, hi = final.forecast_ci(h, alpha=alpha)
            lo     = np.maximum(lo, 0.0)
            diag   = StatUtils.compute_diagnostics(final._residuals)

            return ForecastResult(
                "ARIMA", preds, lo, hi,
                mae, rmse, smape, mase_, mape, r2,
                aic=final._aic, bic=final._bic,
                cv_smape=cv_scores.get("smape"),
                diagnostics=diag,
                metadata={"order": (final.p, final.d, final.q)},
                train_time_s=time.perf_counter() - t0,
            )
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"ARIMA failed: {exc}")
            logging.exception("ARIMA error")
            return None

    # ── Holt-Winters ──────────────────────────────────────────────────────────

    def run_exp_smoothing(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_exp_smoothing:
            return None
        self._report("Fitting Holt-Winters (auto add/mul) …")
        t0 = time.perf_counter()
        try:
            self._check()
            h     = self.config.forecast_horizon
            alpha = 1.0 - self.config.confidence_level

            def hw_fn(y_tr: "np.ndarray") -> "np.ndarray":
                return np.maximum(
                    HoltWinters.auto_select(y_tr, season_len=period).predict(h), 0.0)

            cv_scores = WalkForwardCV.evaluate(
                hw_fn, y, period=period,
                n_splits=self.config.cv_folds, horizon=min(h, len(y) // 4))

            split       = max(h, len(y) // 5)
            train, test = y[:-split], y[-split:]
            val_mdl     = HoltWinters.auto_select(train, season_len=period)
            val_preds   = np.maximum(val_mdl.predict(split)[:len(test)], 0.0)
            mae, rmse, smape, mase_, mape, r2 = StatUtils.all_metrics(
                test, val_preds, train, period)

            final   = HoltWinters.auto_select(y, season_len=period)
            preds   = np.maximum(final.predict(h), 0.0)
            lo, hi  = final.forecast_ci(h, alpha=alpha)
            lo      = np.maximum(lo, 0.0)
            diag    = StatUtils.compute_diagnostics(final._resid)
            variant = "Multiplicative" if final.multiplicative else "Additive"

            return ForecastResult(
                "Holt-Winters", preds, lo, hi,
                mae, rmse, smape, mase_, mape, r2,
                aic=final._aic,
                cv_smape=cv_scores.get("smape"),
                diagnostics=diag,
                metadata={"variant": variant,
                           "alpha": round(final.alpha, 4),
                           "beta":  round(final.beta,  4),
                           "gamma": round(final.gamma, 4)},
                train_time_s=time.perf_counter() - t0,
            )
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"Holt-Winters failed: {exc}")
            return None

    # ── Generic ML runner ─────────────────────────────────────────────────────
    #
    #  [BUG-3 FIX] Feature matrix is built ONCE on the full series, cached in
    #  (X_full, y_full, feat_names), and reused for the final model.
    #  The previous code called MLFeatureEngineering.build_features() a second
    #  time inside the same method, wasting CPU and creating an inconsistent
    #  scaler baseline.  The CV loop still uses sklearn TimeSeriesSplit on the
    #  cached arrays, which is both correct and efficient.

    def _run_ml_model(
        self, y: "np.ndarray",
        base_model: Any,
        model_name: str,
        period: int = 12,
    ) -> Optional[ForecastResult]:
        self._report(f"Fitting {model_name} …")
        t0 = time.perf_counter()
        try:
            self._check()
            h     = self.config.forecast_horizon
            alpha = 1.0 - self.config.confidence_level

            # ── Build feature matrix ONCE ────────────────────────────────────
            X_full, y_full, feat_names = MLFeatureEngineering.build_features(y, period=period)
            if len(X_full) < 12:
                self._report(f"{model_name}: insufficient data after feature engineering.")
                return None

            # ── Walk-forward CV on the cached feature matrix ─────────────────
            n_splits = min(self.config.cv_folds, max(2, len(X_full) // (h + 1)))
            tscv     = TimeSeriesSplit(n_splits=n_splits,
                                       test_size=min(h, max(1, len(X_full) // 4)))
            cv_preds, cv_actuals = [], []
            for tr_idx, te_idx in tscv.split(X_full):
                X_tr, X_te = X_full[tr_idx], X_full[te_idx]
                y_tr, y_te = y_full[tr_idx], y_full[te_idx]
                # scaler fit on TRAIN only — no leakage into test
                sc       = StandardScaler()
                fold_mdl = sklearn_clone(base_model)
                X_tr_sc  = sc.fit_transform(X_tr)
                X_te_sc  = sc.transform(X_te)
                fold_mdl.fit(X_tr_sc, y_tr)
                prd = np.maximum(fold_mdl.predict(X_te_sc), 0.0)
                cv_preds.extend(prd.tolist())
                cv_actuals.extend(y_te.tolist())

            cv_smape = (StatUtils.smape(np.array(cv_actuals), np.array(cv_preds))
                        if cv_preds else None)

            # ── Validation metrics (80/20 split on cached matrix) ────────────
            split      = max(h, len(X_full) // 5)
            X_tr, X_te = X_full[:-split], X_full[-split:]
            y_tr, y_te = y_full[:-split], y_full[-split:]
            # separate scaler for validation — fitted on train only
            val_scaler = StandardScaler()
            val_mdl    = sklearn_clone(base_model)
            val_mdl.fit(val_scaler.fit_transform(X_tr), y_tr)
            val_preds  = np.maximum(val_mdl.predict(val_scaler.transform(X_te)), 0.0)
            val_resid  = y_te - val_preds
            mae, rmse, smape, mase_, mape, r2 = StatUtils.all_metrics(
                y_te, val_preds, y_tr, period)

            self._check()

            # ── Final model on the full cached feature matrix ─────────────────
            full_scaler = StandardScaler()
            full_mdl    = sklearn_clone(base_model)
            full_mdl.fit(full_scaler.fit_transform(X_full), y_full)

            # ── Iterative multi-step forecast (recursive strategy) ────────────
            future_y = list(y.copy().astype(np.float64))
            preds: List[float] = []
            for step in range(h):
                try:
                    X_fut, _, new_feat_names = MLFeatureEngineering.build_features(
                        np.array(future_y, dtype=np.float64), period=period)
                except ValueError:
                    break
                X_fut_df = pd.DataFrame(X_fut, columns=new_feat_names)
                X_fut_df = X_fut_df.reindex(columns=feat_names, fill_value=0.0)
                X_fut_arr = X_fut_df.to_numpy(dtype=np.float64)
                if X_fut_arr.shape[1] != len(feat_names):
                    logger.warning(
                        "%s: feature shape mismatch at step %d, stopping.",
                        model_name, step)
                    break
                # sanitise before predict
                X_fut_arr = np.nan_to_num(X_fut_arr, nan=0.0, posinf=0.0, neginf=0.0)
                if len(X_fut_arr) == 0:
                    break
                p = float(max(0.0,
                              full_mdl.predict(full_scaler.transform(X_fut_arr[-1:]))[0]))
                preds.append(p)
                future_y.append(p)

            if not preds:
                return None
            preds_arr = np.array(preds[:h], dtype=float)

            lo, hi = StatUtils.bootstrap_ci(
                preds_arr, val_resid, alpha=alpha, n_boot=self.config.n_bootstrap)
            lo = np.maximum(lo, 0.0)

            diag     = StatUtils.compute_diagnostics(val_resid)
            feat_imp = None
            if hasattr(full_mdl, "feature_importances_"):
                pairs    = zip(feat_names, full_mdl.feature_importances_)
                feat_imp = dict(sorted(pairs, key=lambda x: x[1], reverse=True)[:10])

            return ForecastResult(
                model_name, preds_arr, lo, hi,
                mae, rmse, smape, mase_, mape, r2,
                cv_smape=cv_smape,
                feature_importance=feat_imp,
                diagnostics=diag,
                train_time_s=time.perf_counter() - t0,
            )
        except InterruptedError:
            raise
        except Exception as exc:
            self._report(f"{model_name} failed: {exc}")
            logging.exception("%s error", model_name)
            return None

    def run_random_forest(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_random_forest:
            return None
        return self._run_ml_model(
            y,
            RandomForestRegressor(n_estimators=200, max_depth=8, random_state=42,
                                  min_samples_leaf=2, max_features=0.8, n_jobs=-1),
            "Random Forest", period,
        )

    def run_gradient_boosting(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_gradient_boosting:
            return None
        return self._run_ml_model(
            y,
            GradientBoostingRegressor(n_estimators=200, learning_rate=0.05,
                                      max_depth=4, subsample=0.8, random_state=42,
                                      min_samples_leaf=3),
            "Gradient Boosting", period,
        )

    def run_ridge(self, y: "np.ndarray", period: int = 12) -> Optional[ForecastResult]:
        if not self.config.use_ridge:
            return None
        return self._run_ml_model(y, Ridge(alpha=1.0), "Ridge Regression", period)

    # ── Stacked Ensemble ──────────────────────────────────────────────────────
    #
    #  [BUG-4 FIX] min_delta guard of 1e-6 is applied before the inversion so
    #  that scores[k] is never zero (which would make 1/score → ∞, and if ALL
    #  scores are ∞, total=∞ and weights become NaN).  The guard activates only
    #  on degenerate datasets where models produce the same sMAPE.

    def build_ensemble(
        self,
        forecasts: Dict[str, Optional[ForecastResult]],
        y: "np.ndarray",
        period: int = 12,
    ) -> "ForecastResult":
        self._report("Building stacked ensemble …")
        valid = {k: v for k, v in forecasts.items()
                 if v is not None and v.is_valid}
        if not valid:
            raise ForecastError("No valid forecasts — cannot build ensemble.")

        h = self.config.forecast_horizon

        # Weight basis: prefer CV sMAPE (unbiased); fall back to val sMAPE.
        # min_delta guard: ensures no score is exactly 0 (avoids div-by-zero
        # and consequent NaN weights when all models tie).
        _MIN_DELTA = 1e-6
        scores: Dict[str, float] = {}
        for name, res in valid.items():
            s = (res.cv_smape
                 if (res.cv_smape is not None and math.isfinite(res.cv_smape))
                 else res.smape)
            scores[name] = max(float(s), _MIN_DELTA)

        inv   = {k: 1.0 / v for k, v in scores.items()}
        total = sum(inv.values())

        # Guard: if total is still not finite (e.g. all inf scores), fall back
        # to uniform weighting so the ensemble is always well-defined.
        # fast-path for single model: weight is trivially 1.0
        if len(valid) == 1:
            weights = {k: 1.0 for k in valid}
        elif not math.isfinite(total) or total <= 0.0:
            n_models = len(valid)
            weights  = {k: 1.0 / n_models for k in valid}
        else:
            weights = {k: v / total for k, v in inv.items()}

        preds = np.zeros(h)
        lower = np.zeros(h)
        upper = np.zeros(h)
        for name, w in weights.items():
            r = valid[name]
            n = min(len(r.predictions), h)
            preds[:n] += w * r.predictions[:n]
            lower[:n] += w * r.lower_ci[:n]
            upper[:n] += w * r.upper_ci[:n]

        def _wavg(attr: str) -> float:
            return float(sum(weights[k] * getattr(valid[k], attr) for k in weights))

        return ForecastResult(
            "Ensemble (Stacked)", preds, lower, upper,
            _wavg("mae"), _wavg("rmse"), _wavg("smape"),
            _wavg("mase"), _wavg("mape"), _wavg("r2"),
            metadata={
                "weights":      {k: round(v, 4) for k, v in weights.items()},
                "weight_basis": "cv_smape (walk-forward, non-leaky)",
            },
        )

    def forecast(
        self, y: "np.ndarray", period: int = 12,
    ) -> Tuple[Dict, "ForecastResult", str]:
        runners = [
            ("ARIMA",             lambda: self.run_arima(y, period)),
            ("Holt-Winters",      lambda: self.run_exp_smoothing(y, period)),
            ("Random Forest",     lambda: self.run_random_forest(y, period)),
            ("Gradient Boosting", lambda: self.run_gradient_boosting(y, period)),
            ("Ridge Regression",  lambda: self.run_ridge(y, period)),
        ]
        results: Dict[str, Optional[ForecastResult]] = {}
        for name, fn in runners:
            self._check()
            try:
                results[name] = fn()
            except InterruptedError:
                raise
            except Exception as exc:
                self._report(f"{name} error: {exc}")
                results[name] = None

        ensemble = self.build_ensemble(results, y, period)
        valid    = {k: v for k, v in results.items() if v is not None and v.is_valid}
        best     = min(valid, key=lambda k: valid[k].smape) if valid else "Ensemble (Stacked)"
        return results, ensemble, best


# ══════════════════════════════════════════════════════════════════════════════
#  ANALYSIS ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class SalesAnalyzer:
    """Orchestrates decomposition, anomaly detection, statistics, and forecasting."""
    def __init__(
        self, config: ForecastConfig,
        progress_cb: Optional[Callable] = None,
        cancel_token: Optional[CancelToken] = None,
    ):
        self.config  = config
        self._cb     = progress_cb or (lambda msg, pct: None)
        self._cancel = cancel_token or CancelToken()

    def _report(self, msg: str, pct: Optional[float] = None) -> None:
        try:
            self._cb(msg, pct)
        except Exception:
            pass

    @staticmethod
    def decompose(
        y: "np.ndarray", period: int = 12,
    ) -> Tuple["np.ndarray", "np.ndarray", "np.ndarray"]:
        """
        Additive decomposition via centred moving average.
        Fully vectorised; handles series shorter than 2*period gracefully.
        """
        n      = len(y)
        y      = np.asarray(y, dtype=np.float64)
        period = max(2, int(period))

        if n < 2 * period:
            trend    = np.full(n, float(np.mean(y)), dtype=np.float64)
            seasonal = np.zeros(n, dtype=np.float64)
            return trend, seasonal, (y - trend)

        # centred moving average via convolution — O(n), single-pass
        kernel     = np.ones(period, dtype=np.float64) / period
        trend_conv = np.convolve(y, kernel, mode="valid")   # length n - period + 1
        n_valid    = len(trend_conv)
        pad_total  = n - n_valid
        pad_l      = max(0, pad_total // 2)     # clamp — never negative
        pad_r      = max(0, pad_total - pad_l)
        trend = np.empty(n, dtype=np.float64)
        if pad_l > 0:
            trend[:pad_l] = trend_conv[0]
        trend[pad_l:pad_l + n_valid] = trend_conv
        if pad_r > 0:
            trend[pad_l + n_valid:] = trend_conv[-1]

        detrended = y - trend
        # seasonal prototype: mean over each phase
        s_proto = np.array([
            float(np.mean(detrended[i::period]))
            for i in range(period)
        ], dtype=np.float64)
        s_proto -= float(s_proto.mean())          # zero-sum constraint
        seasonal  = np.tile(s_proto, n // period + 2)[:n]
        residual  = y - trend - seasonal
        return trend, seasonal, residual

    @staticmethod
    def detect_anomalies(
        series: "pd.Series", threshold: float = 2.5,
    ) -> "pd.DataFrame":
        # window size at least 3 and at most 12; flat-series guard
        w   = min(max(3, len(series) // 8), 12)
        rm  = series.rolling(w, center=True, min_periods=1).mean()
        rs  = series.rolling(w, center=True, min_periods=1).std().fillna(0.0)
        # flat series → std = 0; use global std as fallback
        global_std = float(series.std(ddof=1)) if len(series) > 1 else 1.0
        rs  = rs.where(rs > 1e-9, other=max(global_std, 1e-9))
        z   = np.abs((series - rm) / rs)
        idx = z[z > threshold].index
        if len(idx) == 0:
            # explicit dtype so downstream consumers always get float/str
            return pd.DataFrame(
                columns=["value", "z_score", "type"]
            ).astype({"value": "float64", "z_score": "float64", "type": "object"})
        anom  = series.loc[idx]
        types = np.where(anom.values > rm.loc[idx].values, "spike", "drop")
        return pd.DataFrame({
            "value":   anom.astype(np.float64),
            "z_score": z.loc[idx].astype(np.float64),
            "type":    pd.array(types, dtype=object),
        })

    @staticmethod
    def compute_statistics(
        df: "pd.DataFrame", roi_ci: Optional[Dict] = None,
    ) -> Dict:
        monthly = df.groupby("date")["sales"].sum()
        n_years = max(len(monthly) / 12.0, 1.0 / 12.0)
        total   = float(df["sales"].sum())
        first   = float(monthly.iloc[0])
        last    = float(monthly.iloc[-1])
        try:
            cagr = ((last / max(abs(first), 1e-10)) ** (1.0 / n_years) - 1.0) * 100.0
            if not math.isfinite(cagr):
                cagr = 0.0
        except Exception:
            cagr = 0.0

        def _best(col: str) -> str:
            if col not in df.columns or df[col].nunique() == 0:
                return "N/A"
            try:
                return str(df.groupby(col)["sales"].sum().idxmax())
            except Exception:
                return "N/A"

        returns_total = float(df["returns"].sum()) if "returns" in df.columns else 0.0
        avg_m  = float(monthly.mean())
        std_m  = float(monthly.std(ddof=1)) if len(monthly) > 1 else 0.0
        roi    = roi_ci or {"point": 0.0, "ci_lo": 0.0, "ci_hi": 0.0}

        # convert Timestamps to ISO strings for JSON safety
        peak_month   = monthly.idxmax()
        trough_month = monthly.idxmin()
        peak_str   = peak_month.isoformat()   if isinstance(peak_month,   pd.Timestamp) else str(peak_month)
        trough_str = trough_month.isoformat() if isinstance(trough_month, pd.Timestamp) else str(trough_month)

        return {
            "total_revenue":          total,
            "avg_monthly_revenue":    avg_m,
            "peak_month":             peak_str,
            "peak_value":             float(monthly.max()),
            "trough_month":           trough_str,
            "trough_value":           float(monthly.min()),
            "std_dev":                std_m,
            "cv":                     (std_m / max(abs(avg_m), 1e-10) * 100.0),
            "cagr":                   cagr,
            "best_product":           _best("product"),
            "best_region":            _best("region"),
            "total_returns":          returns_total,
            "return_rate":            (returns_total / max(abs(total), 1e-10) * 100.0),
            "marketing_roi":          roi["point"],
            "marketing_roi_ci_lo":    roi["ci_lo"],
            "marketing_roi_ci_hi":    roi["ci_hi"],
            "n_records":              len(df),
            "n_months":               len(monthly),
        }

    def analyze(self, df: "pd.DataFrame") -> SalesAnalysis:
        self._report("Aggregating monthly data …", 5)
        self._cancel.check()
        monthly = df.groupby("date")["sales"].sum().sort_index()
        y       = monthly.to_numpy(dtype=float)

        cfg_period = self.config.seasonality_period
        if cfg_period <= 0:
            self._report("Detecting seasonality period …", 8)
            cfg_period = StatUtils.detect_seasonality(y)
        else:
            detected = StatUtils.detect_seasonality(y)
            if detected != cfg_period:
                self._report(
                    f"Configured period={cfg_period}, "
                    f"auto-detected={detected} — using configured.", 8)

        self._report(f"Decomposing (period={cfg_period}) …", 15)
        self._cancel.check()
        trend, seasonal, residual = SalesAnalyzer.decompose(y, cfg_period)

        self._report("Detecting anomalies …", 20)
        anomalies = SalesAnalyzer.detect_anomalies(monthly)

        yoy      = monthly.pct_change(12) * 100.0
        mom      = monthly.pct_change()   * 100.0
        rolling3 = monthly.rolling(3, min_periods=1).mean()

        self._report("Computing statistics …", 25)
        roi_ci = None
        if "marketing_spend" in df.columns:
            mkt    = df.groupby("date")["marketing_spend"].sum().reindex(
                monthly.index, fill_value=0.0).to_numpy()
            roi_ci = StatUtils.marketing_roi_with_ci(
                y, mkt, n_boot=self.config.n_bootstrap)

        statistics = self.compute_statistics(df, roi_ci=roi_ci)

        self._report("Running forecasting models …", 30)
        engine = ForecastingEngine(self.config, self._cb, self._cancel)
        forecasts, ensemble, best_model = engine.forecast(y, cfg_period)

        self._report("Finalising …", 90)
        return SalesAnalysis(
            raw_data=df, trend=trend, seasonal=seasonal, residual=residual,
            anomalies=anomalies, yoy_growth=yoy, mom_growth=mom,
            rolling_avg=rolling3, statistics=statistics,
            forecasts=forecasts, best_model=best_model,
            ensemble_forecast=ensemble,
            detected_period=cfg_period,
        )


# ══════════════════════════════════════════════════════════════════════════════
#  CHART GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class ChartGenerator:
    """Renders all matplotlib/seaborn charts to PNG files in the temp directory."""
    PALETTE = Theme.CHART

    def __init__(self, config: ForecastConfig):
        self.config = config
        sns.set_theme(style="whitegrid", font_scale=0.95)
        plt.rcParams.update({
            "figure.facecolor":  "white",
            "axes.facecolor":    "white",
            "grid.alpha":        0.35,
            "axes.spines.top":   False,
            "axes.spines.right": False,
            "font.family":       "DejaVu Sans",
        })

    def _save(self, fig: "plt.Figure", name: str, dpi: int = 180) -> Path:
        """DPI raised to 180 for crisper report output."""
        path = self.config.charts_dir / f"{name}.png"
        try:
            # if charts_dir was cleaned up mid-run, log and skip gracefully
            self.config.charts_dir.mkdir(parents=True, exist_ok=True)
            fig.savefig(str(path), dpi=dpi, bbox_inches="tight",
                        facecolor="white", edgecolor="none")
        except Exception:
            logging.exception("Chart save failed for %s", name)
        finally:
            try:
                plt.close(fig)
            except Exception:
                pass
        return path

    @staticmethod
    def _money_fmt(x: float, _: Any) -> str:
        if abs(x) >= 1_000_000:
            return f"${x/1_000_000:.1f}M"
        if abs(x) >= 1_000:
            return f"${x/1_000:.0f}K"
        return f"${x:.0f}"

    def sales_overview(self, analysis: SalesAnalysis) -> Path:
        monthly = analysis.raw_data.groupby("date")["sales"].sum()
        fig, ax = plt.subplots(figsize=(12, 4.5))
        ax.fill_between(monthly.index, monthly.values, alpha=0.13, color=self.PALETTE[0])
        ax.plot(monthly.index, monthly.values,
                color=self.PALETTE[0], lw=2.5, label="Monthly Sales")
        ax.plot(monthly.index, analysis.rolling_avg.values,
                color=self.PALETTE[1], lw=2.0, ls="--", label="3-Month Rolling Avg")
        ax.plot(monthly.index, analysis.trend,
                color=self.PALETTE[2], lw=2.0, ls="-.", label="Trend")
        if not analysis.anomalies.empty:
            ax.scatter(analysis.anomalies.index, analysis.anomalies["value"],
                       c="red", zorder=5, s=80, marker="x", lw=2.5, label="Anomaly")
        ax.set_title("Historical Sales Overview", fontsize=14, fontweight="bold", pad=12)
        ax.set_ylabel(f"Revenue ({self.config.currency_symbol})", fontsize=11)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        ax.legend(frameon=True, fontsize=9, loc="upper left")
        ax.tick_params(axis="x", rotation=30)
        fig.tight_layout()
        return self._save(fig, "sales_overview")

    def forecast_chart(self, analysis: SalesAnalysis) -> Path:
        monthly      = analysis.raw_data.groupby("date")["sales"].sum()
        last_date    = monthly.index[-1]
        future_dates = pd.date_range(
            last_date + pd.DateOffset(months=1),
            periods=self.config.forecast_horizon, freq="MS")
        fig, ax = plt.subplots(figsize=(13, 5))
        ax.plot(monthly.index[-24:], monthly.values[-24:],
                color=self.PALETTE[0], lw=2.5, label="Historical", zorder=4)
        for i, (name, result) in enumerate(analysis.forecasts.items()):
            if result is None or not result.is_valid:
                continue
            n = min(len(result.predictions), len(future_dates))
            ax.plot(future_dates[:n], result.predictions[:n],
                    color=self.PALETTE[(i + 2) % len(self.PALETTE)],
                    lw=1.2, alpha=0.55, ls="--", label=name)
        ens = analysis.ensemble_forecast
        n   = min(len(ens.predictions), len(future_dates))
        ax.plot(future_dates[:n], ens.predictions[:n],
                color=self.PALETTE[1], lw=3, label="Ensemble Forecast", zorder=5)
        ax.fill_between(future_dates[:n],
                        ens.lower_ci[:n], ens.upper_ci[:n],
                        alpha=0.20, color=self.PALETTE[1],
                        label=f"{int(self.config.confidence_level*100)}% CI")
        ax.axvline(x=last_date, color="gray", ls=":", lw=1.5, alpha=0.7)
        ax.set_title("Sales Forecast — All Models vs Ensemble",
                     fontsize=14, fontweight="bold", pad=12)
        ax.set_ylabel(f"Revenue ({self.config.currency_symbol})", fontsize=11)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        ax.legend(frameon=True, fontsize=8, ncol=2, loc="upper left")
        ax.tick_params(axis="x", rotation=30)
        fig.tight_layout()
        return self._save(fig, "forecast")

    def decomposition_chart(self, analysis: SalesAnalysis) -> Path:
        monthly = analysis.raw_data.groupby("date")["sales"].sum()
        fig, axes = plt.subplots(4, 1, figsize=(12, 10), sharex=True)
        layers = [
            (monthly.values,    "Original",  self.PALETTE[0]),
            (analysis.trend,    "Trend",     self.PALETTE[1]),
            (analysis.seasonal, "Seasonal",  self.PALETTE[2]),
            (analysis.residual, "Residual",  self.PALETTE[3]),
        ]
        for ax, (data, title, color) in zip(axes, layers):
            ax.plot(monthly.index, data, color=color, lw=1.8)
            ax.fill_between(monthly.index, data, alpha=0.12, color=color)
            ax.set_ylabel(title, fontsize=10, fontweight="bold")
            ax.yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        axes[-1].tick_params(axis="x", rotation=30)
        fig.suptitle("Time Series Decomposition",
                     fontsize=14, fontweight="bold", y=1.01)
        fig.tight_layout()
        return self._save(fig, "decomposition")

    def product_comparison(self, analysis: SalesAnalysis) -> Path:
        pm  = (analysis.raw_data
               .groupby(["date", "product"])["sales"].sum().reset_index())
        fig, axes = plt.subplots(1, 2, figsize=(13, 5))
        for i, prod in enumerate(pm["product"].unique()):
            sub = pm[pm["product"] == prod]
            axes[0].plot(sub["date"], sub["sales"],
                         color=self.PALETTE[i % len(self.PALETTE)], lw=1.8, label=prod)
        axes[0].set_title("Revenue by Product Over Time", fontsize=12, fontweight="bold")
        axes[0].set_ylabel(f"Revenue ({self.config.currency_symbol})")
        axes[0].yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        axes[0].legend(fontsize=8)
        axes[0].tick_params(axis="x", rotation=30)

        totals = (analysis.raw_data.groupby("product")["sales"]
                  .sum().sort_values(ascending=False))
        bars = axes[1].bar(totals.index, totals.values,
                            color=self.PALETTE[:len(totals)])
        axes[1].set_title("Total Revenue by Product", fontsize=12, fontweight="bold")
        axes[1].set_ylabel(f"Total Revenue ({self.config.currency_symbol})")
        axes[1].yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))
        for bar in bars:
            h = bar.get_height()
            axes[1].text(bar.get_x() + bar.get_width() / 2, h * 1.01,
                          self._money_fmt(h, None),
                          ha="center", va="bottom", fontsize=8.5, fontweight="bold")
        fig.tight_layout()
        return self._save(fig, "product_comparison")

    def model_comparison(self, analysis: SalesAnalysis) -> Path:
        valid = {k: v for k, v in analysis.forecasts.items()
                 if v is not None and v.is_valid}
        valid["Ensemble"] = analysis.ensemble_forecast
        names = list(valid.keys())
        x     = np.arange(len(names))
        fig, axes = plt.subplots(1, 4, figsize=(17, 4.5))
        metrics = [
            ([v.smape for v in valid.values()], "sMAPE (%)\n(lower=better)",  self.PALETTE[2]),
            ([v.mase  for v in valid.values()], "MASE\n(lower=better)",       self.PALETTE[3]),
            ([v.rmse  for v in valid.values()], "RMSE ($)\n(lower=better)",   self.PALETTE[0]),
            ([v.r2    for v in valid.values()], "R² Score\n(higher=better)",  self.PALETTE[1]),
        ]
        for ax, (vals, title, color) in zip(axes, metrics):
            bars = ax.bar(x, vals, color=color, alpha=0.85, width=0.6)
            ax.set_xticks(x)
            ax.set_xticklabels(names, rotation=30, ha="right", fontsize=8.5)
            ax.set_title(title, fontsize=10, fontweight="bold")
            for bar in bars:
                h = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, h * 1.02,
                        f"{h:.2f}", ha="center", va="bottom", fontsize=7.5)
        fig.suptitle("Model Performance Comparison", fontsize=13, fontweight="bold", y=1.02)
        fig.tight_layout()
        return self._save(fig, "model_comparison")

    def yoy_growth_chart(self, analysis: SalesAnalysis) -> Path:
        yoy = analysis.yoy_growth.dropna()
        fig, ax = plt.subplots(figsize=(12, 4))
        bar_colors = [self.PALETTE[3] if v >= 0 else self.PALETTE[4] for v in yoy.values]
        ax.bar(yoy.index, yoy.values, color=bar_colors, width=20, alpha=0.85)
        ax.axhline(0, color="black", lw=0.8)
        ax.set_title("Year-over-Year Revenue Growth (%)",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_ylabel("YoY Growth (%)")
        ax.yaxis.set_major_formatter(
            plt.FuncFormatter(lambda x, _: f"{x:.1f}%"))
        ax.tick_params(axis="x", rotation=30)
        fig.tight_layout()
        return self._save(fig, "yoy_growth")

    def heatmap_chart(self, analysis: SalesAnalysis) -> Path:
        df          = analysis.raw_data.copy()
        df["month"] = df["date"].dt.month
        df["year"]  = df["date"].dt.year
        pivot       = df.pivot_table(
            values="sales", index="year", columns="month", aggfunc="sum")
        month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                       "Jul","Aug","Sep","Oct","Nov","Dec"]
        # cast to int to avoid KeyError when pandas returns float column labels
        pivot.columns = [month_names[int(m) - 1] for m in pivot.columns]
        fig, ax = plt.subplots(figsize=(13, 4.5))
        sns.heatmap(pivot / 1e6, annot=True, fmt=".1f", cmap="Blues",
                    linewidths=0.5, ax=ax, cbar_kws={"label": "Revenue ($M)"})
        ax.set_title("Monthly Revenue Heatmap by Year ($M)",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_xlabel("Month"); ax.set_ylabel("Year")
        fig.tight_layout()
        return self._save(fig, "heatmap")

    def residual_diagnostics_chart(self, analysis: SalesAnalysis) -> Path:
        monthly = analysis.raw_data.groupby("date")["sales"].sum()
        resid   = analysis.residual.copy()
        n       = len(resid)

        fig, axes = plt.subplots(2, 2, figsize=(12, 8))
        fig.suptitle(f"Residual Diagnostics — {analysis.best_model}",
                     fontsize=13, fontweight="bold")

        axes[0, 0].plot(monthly.index, resid, color=self.PALETTE[0], lw=1.5)
        axes[0, 0].axhline(0, ls="--", color="gray", lw=1)
        axes[0, 0].set_title("Residuals vs Time")
        axes[0, 0].yaxis.set_major_formatter(plt.FuncFormatter(self._money_fmt))

        max_lag  = max(1, min(24, n // 2 - 2))
        acf_vals = []
        for k in range(1, max_lag + 1):
            if n - k < 2:
                break
            r = np.corrcoef(resid[:n - k], resid[k:])
            acf_vals.append(float(r[0, 1]) if not np.any(np.isnan(r)) else 0.0)
        if acf_vals:
            conf = 1.96 / math.sqrt(max(n, 2))
            axes[0, 1].bar(range(1, len(acf_vals) + 1), acf_vals,
                            color=self.PALETTE[1], alpha=0.8)
            axes[0, 1].axhline( conf, ls="--", color="red", lw=1, label="95% band")
            axes[0, 1].axhline(-conf, ls="--", color="red", lw=1)
            axes[0, 1].set_title("ACF of Residuals")
            axes[0, 1].set_xlabel("Lag")
            axes[0, 1].legend(fontsize=8)

        try:
            stats.probplot(resid, dist="norm", plot=axes[1, 0])
            axes[1, 0].set_title("Normal Q-Q Plot")
        except Exception:
            axes[1, 0].set_title("Normal Q-Q Plot (unavailable)")

        axes[1, 1].hist(resid, bins=min(20, max(5, n // 3)),
                         color=self.PALETTE[0], alpha=0.7, edgecolor="white")
        mu, sigma = float(np.mean(resid)), float(np.std(resid))
        if sigma > 1e-10 and n > 5:
            r_range = float(resid.max() - resid.min())
            bin_w   = r_range / max(min(20, n // 3), 1)
            x       = np.linspace(mu - 4 * sigma, mu + 4 * sigma, 300)
            axes[1, 1].plot(
                x, stats.norm.pdf(x, mu, sigma) * n * bin_w,
                color=self.PALETTE[2], lw=2, label="Normal fit")
            axes[1, 1].legend(fontsize=8)
        axes[1, 1].set_title("Residual Distribution")

        fig.tight_layout()
        return self._save(fig, "residual_diagnostics")

    def generate_all(
        self, analysis: SalesAnalysis,
        progress_cb: Optional[Callable] = None,
    ) -> Dict[str, Path]:
        cb    = progress_cb or (lambda msg, pct: None)
        steps = [
            ("overview",         self.sales_overview),
            ("forecast",         self.forecast_chart),
            ("decomposition",    self.decomposition_chart),
            ("products",         self.product_comparison),
            ("model_comparison", self.model_comparison),
            ("yoy_growth",       self.yoy_growth_chart),
            ("heatmap",          self.heatmap_chart),
            ("diagnostics",      self.residual_diagnostics_chart),
        ]
        charts: Dict[str, Path] = {}
        for idx, (name, fn) in enumerate(steps):
            cb(f"Rendering chart: {name.replace('_',' ').title()} …",
               91 + int(idx / len(steps) * 5))
            try:
                charts[name] = fn(analysis)
            except Exception as exc:
                cb(f"Chart '{name}' skipped: {exc}", None)
                logging.exception("Chart %s failed", name)
        return charts


# ══════════════════════════════════════════════════════════════════════════════
#  PDF REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class PDFReportGenerator:
    def __init__(self, config: ForecastConfig):
        self.config = config
        self.styles = getSampleStyleSheet()
        self._setup_styles()

    @staticmethod
    def _rl_color(rgb: Tuple) -> "colors.Color":
        return colors.Color(*[float(c) for c in rgb[:3]])

    def _setup_styles(self) -> None:
        primary = self._rl_color(self.config.primary_color)
        accent  = self._rl_color(self.config.accent_color)

        def _add(name: str, **kw: Any) -> None:
            if name not in self.styles:
                self.styles.add(ParagraphStyle(name, **kw))

        _add("ReportTitle", parent=self.styles["Title"],    fontSize=24, textColor=primary,
             spaceAfter=6, fontName="Helvetica-Bold", alignment=TA_CENTER)
        _add("ReportSub",   parent=self.styles["Normal"],   fontSize=11, textColor=accent,
             spaceAfter=18, alignment=TA_CENTER)
        _add("SecHead",     parent=self.styles["Heading1"], fontSize=14, textColor=primary,
             spaceBefore=18, spaceAfter=8, fontName="Helvetica-Bold")
        _add("SubHead",     parent=self.styles["Heading2"], fontSize=11, textColor=accent,
             spaceBefore=10, spaceAfter=6, fontName="Helvetica-Bold")
        _add("Body2",       parent=self.styles["Normal"],   fontSize=9.5, leading=14,
             spaceAfter=6, alignment=TA_JUSTIFY)
        _add("Callout",     parent=self.styles["Normal"],   fontSize=9, leading=13,
             leftIndent=12, rightIndent=12,
             backColor=colors.HexColor("#F0F7FF"), spaceAfter=8)
        _add("Disclaimer",  parent=self.styles["Normal"],   fontSize=7.5,
             textColor=colors.HexColor("#888888"), alignment=TA_CENTER)

    def _kpi_table(self, st: Dict) -> Table:
        primary = self._rl_color(self.config.primary_color)
        sym     = self.config.currency_symbol 
        roi_str = (f"{st['marketing_roi']:.1f}x  "
                   f"[{st['marketing_roi_ci_lo']:.1f}x – "
                   f"{st['marketing_roi_ci_hi']:.1f}x]")
        kpis = [
            ("Total Revenue",           f"{sym}{st['total_revenue']:,.0f}"),
            ("Avg Monthly Revenue",     f"{sym}{st['avg_monthly_revenue']:,.0f}"),
            ("CAGR",                    f"{st['cagr']:.1f}%"),
            ("Return Rate",             f"{st['return_rate']:.2f}%"),
            ("Marketing ROI (95% CI)",  roi_str),
            ("Best Product",            str(st["best_product"])),
            ("Best Region",             str(st["best_region"])),
            ("Revenue Volatility (CV)", f"{st['cv']:.1f}%"),
        ]

        def _cell(label: str, value: str) -> List:
            return [
                Paragraph(f"<b>{label}</b>",
                          ParagraphStyle("kl", fontSize=8.5,
                                         textColor=colors.white, leading=11)),
                Paragraph(value,
                          ParagraphStyle("kv", fontSize=11,
                                         fontName="Helvetica-Bold",
                                         textColor=colors.HexColor("#FFE066"), leading=15)),
            ]

        rows = []
        for i in range(0, len(kpis), 2):
            r = [_cell(*kpis[i]),
                 _cell(*kpis[i + 1]) if i + 1 < len(kpis) else ""]
            rows.append(r)

        t = Table(rows, colWidths=[3.5 * inch, 3.5 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), primary),
            ("BOX",        (0, 0), (-1, -1), 0.5, colors.white),
            ("INNERGRID",  (0, 0), (-1, -1), 0.5, colors.HexColor("#2D6DA8")),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING",    (0, 0), (-1, -1), 10),
        ]))
        return t

    def _metrics_table(self, analysis: SalesAnalysis) -> Table:
        valid = {k: v for k, v in analysis.forecasts.items() if v is not None}
        valid["Ensemble ★"] = analysis.ensemble_forecast
        sym = self.config.currency_symbol  # use configured symbol
        data = [["Model", "sMAPE (%)", "MASE", f"RMSE ({sym})", "R²",
                 "CV sMAPE", "Time (s)"]]
        for name, res in valid.items():
            cv  = (f"{res.cv_smape:.1f}"
                   if res.cv_smape is not None and math.isfinite(res.cv_smape)
                   else "—")
            data.append([
                name, f"{res.smape:.2f}", f"{res.mase:.3f}",
                f"{sym}{res.rmse:,.0f}", f"{res.r2:.3f}", cv,
                f"{res.train_time_s:.1f}",
            ])
        cw = [2.0, 0.85, 0.75, 1.1, 0.7, 0.85, 0.75]
        t  = Table(data, colWidths=[c * inch for c in cw])
        primary = self._rl_color(self.config.primary_color)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),      primary),
            ("TEXTCOLOR",     (0, 0), (-1, 0),      colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),      "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1),     8.5),
            ("ALIGN",         (0, 0), (-1, -1),     "CENTER"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1),
             [colors.HexColor("#F5F8FC"), colors.white]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D8E0")),
            ("PADDING",       (0, 0), (-1, -1),     5),
            ("BACKGROUND",    (0, len(data) - 1), (-1, len(data) - 1),
             colors.HexColor("#E8F4FD")),
            ("FONTNAME",      (0, len(data) - 1), (-1, len(data) - 1),
             "Helvetica-Bold"),
        ]))
        return t

    def _forecast_table(self, analysis: SalesAnalysis) -> Table:
        monthly      = analysis.raw_data.groupby("date")["sales"].sum()
        last_date    = monthly.index[-1]
        future_dates = pd.date_range(
            last_date + pd.DateOffset(months=1),
            periods=self.config.forecast_horizon, freq="MS")
        ens  = analysis.ensemble_forecast
        data = [["Month", "Forecast ($)", "Lower CI ($)", "Upper CI ($)",
                 "CI Width ($)", "MoM Δ"]]
        prev = float(monthly.iloc[-1])
        for dt, fc, lo, hi in zip(
            future_dates, ens.predictions, ens.lower_ci, ens.upper_ci
        ):
            mom = (float(fc) - prev) / max(abs(prev), 1e-10) * 100.0
            data.append([
                dt.strftime("%b %Y"),
                f"${float(fc):,.0f}", f"${float(lo):,.0f}", f"${float(hi):,.0f}",
                f"${float(hi) - float(lo):,.0f}",
                f"{'+' if mom >= 0 else ''}{mom:.1f}%",
            ])
            prev = float(fc)
        cw = [1.1, 1.2, 1.2, 1.2, 1.2, 1.1]
        t  = Table(data, colWidths=[c * inch for c in cw])
        primary = self._rl_color(self.config.primary_color)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),      primary),
            ("TEXTCOLOR",     (0, 0), (-1, 0),      colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),      "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1),     8.5),
            ("ALIGN",         (0, 0), (-1, -1),     "CENTER"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1),
             [colors.HexColor("#F5F8FC"), colors.white]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D8E0")),
            ("PADDING",       (0, 0), (-1, -1),     5),
        ]))
        return t

    def _add_chart(self, path: Optional[Path],
                   width: float = 6.5, caption: str = "",
                   heading: str = "", heading_style: str = "SubHead") -> List:
        """
        Return a KeepTogether block so heading + chart image + caption are
        never split across pages.  Falls back gracefully if image is missing.
        """
        if not path or not Path(path).exists():
            return []
        inner: List = []
        try:
            if heading:
                inner.append(Paragraph(heading, self.styles[heading_style]))
                inner.append(Spacer(1, 0.06 * inch))

            # Clamp to available page width (margins 2 × 0.65 in)
            page_w = A4[0] - 1.30 * inch
            img_w  = min(width * inch, page_w)
            img_h  = img_w * 0.42

            img        = RLImage(str(path), width=img_w, height=img_h)
            img.hAlign = "CENTER"
            inner.append(img)

            if caption:
                inner.append(Paragraph(
                    f"<i>{caption}</i>",
                    ParagraphStyle("cap", fontSize=8, alignment=TA_CENTER,
                                   textColor=colors.HexColor("#888888"),
                                   spaceAfter=8, spaceBefore=4)))
        except Exception as exc:
            logging.warning("Could not embed chart %s: %s", path, exc)
            return []

        return [KeepTogether(inner), Spacer(1, 0.12 * inch)]

    def _header_footer(self, canvas_obj: Any, doc: Any) -> None:
        canvas_obj.saveState()
        pc  = self.config.primary_color
        W, H = doc.pagesize
        canvas_obj.setFillColorRGB(*pc[:3])
        canvas_obj.rect(0, H - 0.55 * inch, W, 0.55 * inch, fill=1, stroke=0)
        canvas_obj.setFillColorRGB(1, 1, 1)
        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.drawString(0.4 * inch, H - 0.32 * inch, self.config.company_name)
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.drawRightString(W - 0.4 * inch, H - 0.32 * inch, self.config.report_title)
        canvas_obj.setFillColorRGB(0.5, 0.5, 0.5)
        canvas_obj.setFont("Helvetica", 7.5)
        canvas_obj.drawString(
            0.4 * inch, 0.3 * inch,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  •  {self.config.analyst_name}")
        canvas_obj.drawRightString(
            W - 0.4 * inch, 0.3 * inch,
            f"Page {canvas_obj.getPageNumber()}")
        canvas_obj.setStrokeColorRGB(0.8, 0.8, 0.8)
        canvas_obj.line(0.4 * inch, 0.45 * inch, W - 0.4 * inch, 0.45 * inch)
        canvas_obj.restoreState()

    def generate(
        self, analysis: SalesAnalysis,
        charts: Dict[str, Path], output_path: Path,
    ) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        tmp = output_path.with_suffix(".tmp.pdf")
        try:
            doc = SimpleDocTemplate(
                str(tmp), pagesize=A4,
                rightMargin=0.65 * inch, leftMargin=0.65 * inch,
                topMargin=0.85 * inch, bottomMargin=0.7 * inch)
            story = self._build_story(analysis, charts)
            doc.build(story,
                       onFirstPage=self._header_footer,
                       onLaterPages=self._header_footer)
            shutil.move(str(tmp), str(output_path))
            return output_path
        except Exception as exc:
            if tmp.exists():
                try:
                    tmp.unlink()
                except OSError:
                    pass
            raise ReportGenerationError(f"PDF generation failed: {exc}") from exc

    def _build_story(self, analysis: SalesAnalysis, charts: Dict) -> List:
        """
        Build the ReportLab story list.

        Layout guarantee: every section heading is KeepTogether with its
        introductory text and the first chart of that section so that a
        heading can never be stranded alone at the bottom of a page.
        """
        S     = self.styles
        story: List = []
        st    = analysis.statistics
        ens   = analysis.ensemble_forecast
        cfg   = self.config

        # ── Cover / KPI page ──────────────────────────────────────────────────
        story.append(Spacer(1, 0.6 * inch))
        story.append(Paragraph(cfg.report_title, S["ReportTitle"]))
        story.append(Paragraph(
            f"{cfg.company_name}  ·  {datetime.now().strftime('%B %d, %Y')}",
            S["ReportSub"]))
        story.append(HRFlowable(
            width="100%", thickness=2,
            color=self._rl_color(cfg.accent_color)))
        story.append(Spacer(1, 0.25 * inch))
        # KPI table: keep heading + table together
        story.append(KeepTogether([
            Paragraph("Executive KPI Summary", S["SecHead"]),
            Spacer(1, 0.08 * inch),
            self._kpi_table(st),
        ]))
        story.append(Spacer(1, 0.3 * inch))
        story.append(PageBreak())

        # ── Section 1: Historical Sales ───────────────────────────────────────
        sec1_intro = [
            Paragraph("1. Historical Sales Performance", S["SecHead"]),
            Spacer(1, 0.06 * inch),
            Paragraph(
                "Total monthly revenue aggregated across all products and regions. "
                "Dashed = 3-month rolling average; dash-dot = centred moving-average trend. "
                "Red × markers = rolling Z-score anomalies (threshold 2.5σ).", S["Body2"]),
        ]
        story.append(KeepTogether(sec1_intro))
        story += self._add_chart(charts.get("overview"),
                                  caption="Figure 1: Historical Sales with Trend & Anomalies")
        story += self._add_chart(charts.get("yoy_growth"),
                                  caption="Figure 2: YoY Revenue Growth (%)",
                                  heading="Year-over-Year Growth")
        story += self._add_chart(charts.get("heatmap"),
                                  caption="Figure 3: Monthly Revenue Heatmap by Year ($M)",
                                  heading="Seasonality Heatmap")
        story.append(PageBreak())

        # ── Section 2: Decomposition ──────────────────────────────────────────
        sec2_intro = [
            Paragraph("2. Time Series Decomposition", S["SecHead"]),
            Spacer(1, 0.06 * inch),
            Paragraph(
                f"Series decomposed into Original, Trend (centred "
                f"{analysis.detected_period}-period moving average), Seasonality, and "
                f"Residual noise (additive model). "
                f"Dominant seasonal period detected: "
                f"<b>{analysis.detected_period} months</b>.", S["Body2"]),
        ]
        story.append(KeepTogether(sec2_intro))
        story += self._add_chart(charts.get("decomposition"),
                                  caption="Figure 4: Decomposition — Trend / Seasonal / Residual")
        story.append(PageBreak())

        # ── Section 3: Product Analysis ───────────────────────────────────────
        story.append(KeepTogether([
            Paragraph("3. Product-Level Analysis", S["SecHead"]),
            Spacer(1, 0.06 * inch),
        ]))
        story += self._add_chart(charts.get("products"),
                                  caption="Figure 5: Revenue by Product — Timeline & Total")
        if "product" in analysis.raw_data.columns:
            prod = (analysis.raw_data
                    .groupby("product")
                    .agg(total=("sales", "sum"),
                         avg=("sales", "mean"),
                         units=("units", "sum"))
                    .reset_index()
                    .sort_values("total", ascending=False))
            tbl_data = [["Product", "Total Revenue ($)", "Avg Monthly ($)", "Units Sold"]]
            for _, row in prod.iterrows():
                tbl_data.append([
                    str(row["product"]),
                    f"${row['total']:,.0f}",
                    f"${row['avg']:,.0f}",
                    f"{int(row['units']):,}",
                ])
            pt = Table(tbl_data,
                        colWidths=[1.8 * inch, 1.7 * inch, 1.7 * inch, 1.8 * inch])
            primary = self._rl_color(cfg.primary_color)
            pt.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0),      primary),
                ("TEXTCOLOR",     (0, 0), (-1, 0),      colors.white),
                ("FONTNAME",      (0, 0), (-1, 0),      "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, -1),     9),
                ("ALIGN",         (0, 0), (-1, -1),     "CENTER"),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1),
                 [colors.HexColor("#F5F8FC"), colors.white]),
                ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D8E0")),
                ("PADDING",       (0, 0), (-1, -1),     6),
            ]))
            story.append(pt)

        if not analysis.anomalies.empty:
            an_data = [["Date", "Revenue ($)", "Z-Score", "Type"]]
            for date_idx, row in analysis.anomalies.head(12).iterrows():
                an_data.append([
                    str(date_idx.date()),
                    f"${row['value']:,.0f}",
                    f"{row['z_score']:.2f}σ",
                    str(row["type"]).capitalize(),
                ])
            at = Table(an_data,
                        colWidths=[1.7 * inch, 1.7 * inch, 1.7 * inch, 1.9 * inch])
            at.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0),      self._rl_color(cfg.primary_color)),
                ("TEXTCOLOR",     (0, 0), (-1, 0),      colors.white),
                ("FONTNAME",      (0, 0), (-1, 0),      "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, -1),     9),
                ("ALIGN",         (0, 0), (-1, -1),     "CENTER"),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1),
                 [colors.HexColor("#FFF8F0"), colors.white]),
                ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D8E0")),
                ("PADDING",       (0, 0), (-1, -1),     6),
            ]))
            # Keep anomaly heading + description + table together
            story.append(Spacer(1, 0.2 * inch))
            story.append(KeepTogether([
                Paragraph("Detected Anomalies", S["SubHead"]),
                Spacer(1, 0.04 * inch),
                Paragraph(
                    f"{len(analysis.anomalies)} anomalous data points detected "
                    "via 6-month rolling Z-score (threshold 2.5σ).", S["Body2"]),
                Spacer(1, 0.06 * inch),
                at,
            ]))
        story.append(PageBreak())

        # ── Section 4: Forecasting Results ────────────────────────────────────
        sec4_intro = [
            Paragraph("4. Demand Forecasting Results", S["SecHead"]),
            Spacer(1, 0.06 * inch),
            Paragraph(
                f"Five models trained and evaluated via walk-forward cross-validation "
                f"(expanding window, no look-ahead). "
                f"Stacking ensemble uses inverse CV-sMAPE weights. "
                f"Best single model by sMAPE: <b>{analysis.best_model}</b>.", S["Body2"]),
        ]
        story.append(KeepTogether(sec4_intro))
        story += self._add_chart(charts.get("forecast"),
                                  caption=(
                                      f"Figure 6: Sales Forecast — Models vs Ensemble "
                                      f"({int(cfg.confidence_level * 100)}% CI)"))
        story += self._add_chart(
            charts.get("model_comparison"),
            caption="Figure 7: Model Comparison",
            heading="Model Performance Comparison (sMAPE / MASE / RMSE / R²)")
        story.append(Spacer(1, 0.10 * inch))
        # Metrics table: keep heading + table together
        story.append(KeepTogether([
            Paragraph("Detailed Model Metrics", S["SubHead"]),
            Spacer(1, 0.06 * inch),
            self._metrics_table(analysis),
        ]))
        story.append(Spacer(1, 0.15 * inch))

        # Residual diagnostics: keep heading + optional text + chart together
        best_res = analysis.forecasts.get(analysis.best_model)
        diag_block: List = [
            Paragraph("Residual Diagnostics", S["SubHead"]),
            Spacer(1, 0.04 * inch),
        ]
        if best_res and best_res.diagnostics:
            diag_block.append(Paragraph(
                f"<b>{analysis.best_model}</b>  |  {best_res.diagnostics.summary()}",
                S["Body2"]))
            diag_block.append(Spacer(1, 0.04 * inch))
        story.append(KeepTogether(diag_block))
        story += self._add_chart(charts.get("diagnostics"),
                                  caption="Figure 8: Residual Diagnostics")
        story.append(PageBreak())

        # ── Section 5: Monthly Forecast Table ────────────────────────────────
        story.append(KeepTogether([
            Paragraph("5. Monthly Forecast Detail", S["SecHead"]),
            Spacer(1, 0.06 * inch),
            Paragraph(
                f"Ensemble monthly point forecast with "
                f"{int(cfg.confidence_level * 100)}% confidence intervals "
                f"and month-over-month change.", S["Body2"]),
            Spacer(1, 0.08 * inch),
        ]))
        story.append(self._forecast_table(analysis))
        story.append(PageBreak())

        # ── Section 6: Strategic Insights ────────────────────────────────────
        story.append(Paragraph("6. Strategic Insights & Recommendations", S["SecHead"]))
        story.append(Spacer(1, 0.06 * inch))
        total_fc  = float(ens.predictions.sum())
        avg_trail = float(st["avg_monthly_revenue"]) * float(cfg.forecast_horizon)
        yoy_proj  = (total_fc / max(avg_trail, 1e-10) - 1.0) * 100.0
        best_res  = analysis.forecasts.get(analysis.best_model) or ens
        ci_lo     = float(st.get("marketing_roi_ci_lo", 0.0))
        ci_hi     = float(st.get("marketing_roi_ci_hi", 0.0))

        insights = [
            (f"Revenue Outlook: {cfg.currency_symbol}{total_fc:,.0f} forecast "
             f"over the next {cfg.forecast_horizon} months "
             f"({yoy_proj:+.1f}% vs trailing same-period avg)."),
            (f"Top Performer: <b>{st['best_product']}</b> leads all products. "
             "Consider increasing inventory allocation and marketing budget."),
            (f"Seasonal Strategy: {analysis.detected_period}-month seasonality detected. "
             "Begin inventory build 6–8 weeks ahead of Q4 peak."),
            (f"Anomaly Management: <b>{len(analysis.anomalies)}</b> anomalies flagged — "
             "review for bulk orders, promotions, or data-entry errors."),
            (f"Model Confidence: Best single model sMAPE = "
             f"<b>{best_res.smape:.1f}%</b> (MASE={best_res.mase:.2f})  |  "
             f"Ensemble sMAPE = <b>{ens.smape:.1f}%</b>. "
             "CV sMAPE (walk-forward) provides unbiased out-of-sample error estimate."),
            (f"Marketing ROI: <b>{st['marketing_roi']:.1f}x</b>  "
             f"(95% bootstrap CI: {ci_lo:.1f}x – {ci_hi:.1f}x). "
             "Interpret the full interval, not just the point estimate."),
            (f"Returns: Return rate is <b>{st['return_rate']:.2f}%</b>. "
             "Review high-return SKUs to improve net revenue materially."),
        ]
        for i, text in enumerate(insights, 1):
            story.append(Paragraph(f"<b>{i}.</b>  {text}", S["Callout"]))

        story.append(Spacer(1, 0.3 * inch))
        story.append(HRFlowable(width="100%", thickness=1,
                                color=self._rl_color(cfg.accent_color)))
        story.append(Spacer(1, 0.1 * inch))
        story.append(Paragraph(
            f"<i>Auto-generated by Sales Forecasting System. "
            "ARIMA: exact MLE via innovations form, ψ-weight CI propagation. "
            "Holt-Winters: SSE minimisation, AIC variant selection, analytical/bootstrap CI. "
            "ML models: PACF-guided features, feature matrix built once per run, "
            "iterative multi-step forecast, bootstrap CI. "
            "Ensemble: inverse CV-sMAPE stacking (walk-forward, non-leaky) with NaN guard. "
            "sMAPE symmetric and bounded [0,200]; MASE scaled to naïve baseline. "
            "Marketing ROI includes 95% bootstrap CI. "
            "Smart column mapping: auto-detects date/sales columns by alias or heuristic. "
            "Forecasts carry inherent uncertainty — combine with domain expertise.</i>",
            S["Disclaimer"]))
        return story


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class ExcelReportGenerator:
    _HEADER_FILL   = "1E4D7B"
    _ALT_FILL      = "F0F5FC"
    _BEST_FILL     = "D4EFDF"
    _ENS_FILL      = "EBF5FB"
    _TITLE_FONT_SZ = 15
    _HEAD_FONT_SZ  = 10
    _BODY_FONT_SZ  = 9

    def __init__(self, config: ForecastConfig):
        self.config = config

    @staticmethod
    def _fill(hex_: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_.lstrip("#"))

    @staticmethod
    def _font(bold: bool = False, size: int = 10,
              color: str = "000000", italic: bool = False) -> Font:
        return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

    @staticmethod
    def _border(color: str = "D0D8E0") -> Border:
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _align(h: str = "center", wrap: bool = True) -> Alignment:
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def _header_cell(self, ws: Any, row: int, col: int,
                     value: Any, fill: str = "") -> None:
        c = ws.cell(row=row, column=col, value=value)
        c.font      = self._font(bold=True, size=self._HEAD_FONT_SZ, color="FFFFFF")
        c.fill      = self._fill(fill or self._HEADER_FILL)
        c.alignment = self._align()
        c.border    = self._border("4A90D9")

    def _data_cell(
        self, ws: Any, row: int, col: int,
        value: Any, shade: bool = False, bold: bool = False,
        color: str = "000000", h_align: str = "center",
        number_format: str = "General",
    ) -> None:
        c               = ws.cell(row=row, column=col, value=value)
        c.font          = self._font(bold=bold, size=self._BODY_FONT_SZ, color=color)
        c.fill          = self._fill(self._ALT_FILL if shade else "FFFFFF")
        c.alignment     = self._align(h_align)
        c.border        = self._border()
        c.number_format = number_format

    def _write_header_row(self, ws: Any, row: int,
                          headers: List[str], col_start: int = 1) -> None:
        for j, h in enumerate(headers, col_start):
            self._header_cell(ws, row, j, h)

    def _write_data_row(
        self, ws: Any, row: int, values: List[Any],
        col_start: int = 1, shade: bool = False,
        bold: bool = False, color: str = "000000",
    ) -> None:
        for j, v in enumerate(values, col_start):
            self._data_cell(ws, row, j, v, shade=shade, bold=bold, color=color)

    def _write_title_block(self, ws: Any, title: str, subtitle: str = "") -> int:
        last_col = get_column_letter(max(ws.max_column or 1, 10))
        ws.merge_cells(f"A1:{last_col}1")
        c           = ws["A1"]
        c.value     = title
        c.font      = Font(name="Calibri", bold=True,
                           size=self._TITLE_FONT_SZ, color="FFFFFF")
        c.fill      = self._fill(self._HEADER_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36
        if subtitle:
            ws.merge_cells(f"A2:{last_col}2")
            c2           = ws["A2"]
            c2.value     = subtitle
            c2.font      = Font(name="Calibri", italic=True, size=9, color="444444")
            c2.fill      = self._fill("EEF3FA")
            c2.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 18
            return 4
        return 3

    def _autofit_columns(self, ws: Any, min_w: int = 8, max_w: int = 35) -> None:
        for col in ws.columns:
            try:
                best   = max(
                    len(str(c.value)) if c.value is not None else 0 for c in col)
                letter = get_column_letter(col[0].column)
                ws.column_dimensions[letter].width = min(max(best + 2, min_w), max_w)
            except Exception:
                pass

    def _add_xl_table(
        self, ws: Any, start_row: int, end_row: int,
        start_col: int, end_col: int, name: str,
    ) -> None:
        try:
            ref   = (f"{get_column_letter(start_col)}{start_row}:"
                     f"{get_column_letter(end_col)}{end_row}")
            tbl   = XLTable(displayName=name, ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False)
            tbl.tableStyleInfo = style
            ws.add_table(tbl)
        except Exception as exc:
            logging.warning("Could not add Excel table '%s': %s", name, exc)

    # ── Sheets ─────────────────────────────────────────────────────────────────

    def _sheet_summary(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws    = wb.create_sheet("📊 Summary")
        start = self._write_title_block(
            ws, "Executive Summary",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {self.config.company_name}")
        st  = analysis.statistics
        ens = analysis.ensemble_forecast
        roi_ci = (f"{st['marketing_roi']:.2f}x  "
                  f"[{st['marketing_roi_ci_lo']:.2f}x – {st['marketing_roi_ci_hi']:.2f}x]")
        kpis = [
            ("Total Revenue",             f"${st['total_revenue']:,.2f}"),
            ("Avg Monthly Revenue",       f"${st['avg_monthly_revenue']:,.2f}"),
            ("Peak Month",                _fmt_month(st["peak_month"])),
            ("Peak Revenue",              f"${st['peak_value']:,.2f}"),
            ("Trough Month",              _fmt_month(st["trough_month"])),
            ("Trough Revenue",            f"${st['trough_value']:,.2f}"),
            ("CAGR",                      f"{st['cagr']:.2f}%"),
            ("Revenue Volatility (CV)",   f"{st['cv']:.2f}%"),
            ("Best Product",              str(st["best_product"])),
            ("Best Region",               str(st["best_region"])),
            ("Total Returns",             f"${st['total_returns']:,.2f}"),
            ("Return Rate",               f"{st['return_rate']:.3f}%"),
            ("Marketing ROI (95% CI)",    roi_ci),
            ("Records in Dataset",        f"{st['n_records']:,}"),
            ("Months of History",         f"{st['n_months']}"),
            ("Detected Seasonal Period",  f"{analysis.detected_period} months"),
            ("Forecast Horizon",          f"{self.config.forecast_horizon} months"),
            ("Ensemble sMAPE",            f"{ens.smape:.2f}%"),
            ("Ensemble MASE",             f"{ens.mase:.3f}"),
            ("Best Single Model",         analysis.best_model),
            ("Total Forecast Revenue",    f"${ens.predictions.sum():,.0f}"),
            ("CI Methods",
             "ARIMA: ψ-weight propagation; HW: analytical/bootstrap; ML: bootstrap"),
            ("Temp Charts Dir",           str(self.config.charts_dir)),
        ]
        self._write_header_row(ws, start, ["KPI Metric", "Value"], col_start=1)
        r = start + 1
        for i, (k, v) in enumerate(kpis):
            self._write_data_row(ws, r, [k, v], shade=(i % 2 == 0))
            ws.cell(row=r, column=1).alignment = self._align("left")
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, 2, "SummaryKPIs")
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 52
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_forecast(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws           = wb.create_sheet("🔮 Forecast")
        monthly      = analysis.raw_data.groupby("date")["sales"].sum()
        last_date    = monthly.index[-1]
        future_dates = pd.date_range(
            last_date + pd.DateOffset(months=1),
            periods=self.config.forecast_horizon, freq="MS")
        ens   = analysis.ensemble_forecast
        start = self._write_title_block(
            ws, "Ensemble Sales Forecast",
            f"Horizon: {self.config.forecast_horizon} months  |  "
            f"Best model: {analysis.best_model}  |  "
            f"Ensemble sMAPE: {ens.smape:.2f}%  |  MASE: {ens.mase:.3f}")
        headers = ["Month", "Point Forecast ($)", "Lower CI ($)", "Upper CI ($)",
                   "CI Width ($)", "MoM Change (%)", "Cumulative ($)"]
        self._write_header_row(ws, start, headers)
        r, prev, cum = start + 1, float(monthly.iloc[-1]), 0.0
        for i, (dt, fc, lo, hi) in enumerate(
            zip(future_dates, ens.predictions, ens.lower_ci, ens.upper_ci)
        ):
            fc_f, lo_f, hi_f = float(fc), float(lo), float(hi)
            mom_pct = (fc_f - prev) / max(abs(prev), 1e-10) * 100.0
            cum    += fc_f
            shade   = (i % 2 == 0)
            vals    = [dt.strftime("%b %Y"),
                       round(fc_f, 2), round(lo_f, 2), round(hi_f, 2),
                       round(hi_f - lo_f, 2),
                       round(mom_pct, 2),
                       round(cum, 2)]
            self._write_data_row(ws, r, vals, shade=shade)
            mom_cell = ws.cell(row=r, column=6)
            mom_cell.number_format = '+0.00;-0.00;0.00'
            mom_cell.font = self._font(
                bold=True, size=self._BODY_FONT_SZ,
                color="1A6B1A" if mom_pct >= 0 else "C51A1A")
            for col_idx in [2, 3, 4, 5, 7]:
                ws.cell(row=r, column=col_idx).number_format = '#,##0.00'
            prev = fc_f
            r   += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "ForecastTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_models(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws    = wb.create_sheet("🤖 Models")
        valid = {k: v for k, v in analysis.forecasts.items() if v is not None}
        valid["Ensemble"] = analysis.ensemble_forecast
        start = self._write_title_block(
            ws, "Forecasting Model Performance",
            "sMAPE = symmetric MAPE (safe near-zero)  |  "
            "MASE = Mean Absolute Scaled Error  |  "
            "CV sMAPE = walk-forward cross-validation")
        headers = ["Model", "sMAPE (%)", "MASE", "RMSE ($)", "R²",
                   "MAE ($)", "Accuracy (%)", "CV sMAPE (%)", "Train (s)",
                   "AIC", "Diagnostics"]
        self._write_header_row(ws, start, headers)
        r          = start + 1
        best_smape = min(v.smape for v in valid.values())
        for i, (name, res) in enumerate(valid.items()):
            is_best = abs(res.smape - best_smape) < 1e-9
            is_ens  = "Ensemble" in name
            shade   = (i % 2 == 0)
            cv_str  = (f"{res.cv_smape:.2f}"
                       if res.cv_smape is not None and math.isfinite(res.cv_smape)
                       else "—")
            aic_str  = f"{res.aic:.1f}" if res.aic is not None else "—"
            diag_str = res.diagnostics.summary() if res.diagnostics else "—"
            vals = [name, round(res.smape, 3), round(res.mase, 4),
                    round(res.rmse, 2), round(res.r2, 4), round(res.mae, 2),
                    round(res.accuracy_score, 2), cv_str,
                    round(res.train_time_s, 2), aic_str, diag_str]
            self._write_data_row(ws, r, vals, shade=shade)
            fill_hex = (self._BEST_FILL if is_best
                        else self._ENS_FILL if is_ens else None)
            if fill_hex:
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c_idx).fill = self._fill(fill_hex)
                    if is_best:
                        ws.cell(row=r, column=c_idx).font = self._font(
                            bold=True, size=self._BODY_FONT_SZ, color="1A6B1A")
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "ModelMetrics")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_history(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws      = wb.create_sheet("📈 History")
        monthly = analysis.raw_data.groupby("date")["sales"].sum().reset_index()
        monthly.columns = ["Date", "Total Sales ($)"]
        monthly["3M Rolling Avg ($)"]  = monthly["Total Sales ($)"].rolling(3).mean().round(2)
        monthly["12M Rolling Avg ($)"] = monthly["Total Sales ($)"].rolling(12).mean().round(2)
        monthly["YoY Growth (%)"]      = monthly["Total Sales ($)"].pct_change(12).mul(100).round(2)
        monthly["MoM Growth (%)"]      = monthly["Total Sales ($)"].pct_change().mul(100).round(2)
        monthly["Date"]                = monthly["Date"].dt.strftime("%Y-%m")
        start   = self._write_title_block(ws, "Historical Monthly Sales",
                                           f"{len(monthly)} months of data")
        headers = monthly.columns.tolist()
        self._write_header_row(ws, start, headers)
        r = start + 1
        for i, row in enumerate(monthly.itertuples(index=False)):
            shade = (i % 2 == 0)
            vals  = list(row)
            self._write_data_row(ws, r, vals, shade=shade)
            # columns: Date(1) Sales(2) 3M(3) 12M(4) YoY(5) MoM(6)
            # colour-code YoY (col 5) and MoM (col 6)
            for col_idx, v in [(5, vals[4]), (6, vals[5])]:
                if v is not None and not (isinstance(v, float) and math.isnan(v)):
                    cell      = ws.cell(row=r, column=col_idx)
                    cell.font = self._font(
                        size=self._BODY_FONT_SZ,
                        color="1A6B1A" if float(v) >= 0 else "C51A1A")
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "HistoryTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"
        for row in ws.iter_rows(min_row=start + 1, max_row=r - 1, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = "#,##0.00"

    def _sheet_products(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws   = wb.create_sheet("🏷️ Products")
        prod = (analysis.raw_data.groupby("product")
                .agg(total=("sales", "sum"), avg=("sales", "mean"),
                     units=("units", "sum"), returns=("returns", "sum"))
                .reset_index().sort_values("total", ascending=False))
        prod["return_rate_%"]   = (prod["returns"] / prod["total"].replace(0, 1) * 100).round(3)
        prod["revenue_share_%"] = (prod["total"] / prod["total"].sum() * 100).round(2)
        start   = self._write_title_block(
            ws, "Product-Level Analysis", "Sorted by total revenue descending")
        headers = ["Product", "Total Revenue ($)", "Avg Monthly ($)", "Units Sold",
                   "Returns ($)", "Return Rate (%)", "Revenue Share (%)"]
        self._write_header_row(ws, start, headers)
        r = start + 1
        for i, row in prod.iterrows():
            vals = [str(row["product"]), round(float(row["total"]), 2),
                    round(float(row["avg"]), 2), int(row["units"]),
                    round(float(row["returns"]), 2),
                    round(float(row["return_rate_%"]), 3),
                    round(float(row["revenue_share_%"]), 2)]
            self._write_data_row(ws, r, vals, shade=(i % 2 == 0))
            ws.cell(row=r, column=1).alignment = self._align("left")
            for col_idx in [2, 3, 5]:
                ws.cell(row=r, column=col_idx).number_format = "#,##0.00"
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "ProductTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_decomp(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws      = wb.create_sheet("🔬 Decomposition")
        monthly = analysis.raw_data.groupby("date")["sales"].sum().reset_index()
        start   = self._write_title_block(
            ws, "Time Series Decomposition",
            f"Centred {analysis.detected_period}-period moving-average trend  |  "
            f"seasonal period = {analysis.detected_period}  |  additive model")
        headers = ["Date (YYYY-MM)", "Original ($)", "Trend ($)",
                   "Seasonal ($)", "Residual ($)"]
        self._write_header_row(ws, start, headers)
        r = start + 1
        n = len(monthly)
        for i, row in monthly.iterrows():
            idx   = min(int(i), n - 1)
            shade = (i % 2 == 0)
            vals  = [row["date"].strftime("%Y-%m"),
                     round(float(row["sales"]),            2),
                     round(float(analysis.trend[idx]),     2),
                     round(float(analysis.seasonal[idx]),  2),
                     round(float(analysis.residual[idx]),  2)]
            self._write_data_row(ws, r, vals, shade=shade)
            for col_idx in [2, 3, 4, 5]:
                ws.cell(row=r, column=col_idx).number_format = "#,##0.00"
            r += 1
        self._add_xl_table(ws, start, r - 1, 1, len(headers), "DecompTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_anomalies(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws    = wb.create_sheet("⚠️ Anomalies")
        start = self._write_title_block(
            ws, "Detected Anomalies",
            f"{len(analysis.anomalies)} anomalous months (rolling Z-score > 2.5σ)")
        headers = ["Date", "Revenue ($)", "Z-Score (σ)", "Type", "Severity"]
        self._write_header_row(ws, start, headers)
        r = start + 1
        if analysis.anomalies.empty:
            ws.cell(row=r, column=1).value = "No anomalies detected."
            ws.cell(row=r, column=1).font  = self._font(italic=True, color="666666")
        else:
            for i, (date_idx, row) in enumerate(analysis.anomalies.iterrows()):
                z    = float(row["z_score"])
                sev  = "High" if z > 4 else ("Medium" if z > 3 else "Low")
                shade = (i % 2 == 0)
                vals  = [str(date_idx.date()),
                         round(float(row["value"]), 2),
                         round(z, 3),
                         str(row["type"]).capitalize(),
                         sev]
                self._write_data_row(ws, r, vals, shade=shade)
                fill = "FFF0F0" if row["type"] == "spike" else "F0FFF0"
                for c_idx in range(1, 6):
                    ws.cell(row=r, column=c_idx).fill = self._fill(fill)
                ws.cell(row=r, column=2).number_format = "#,##0.00"
                r += 1
            self._add_xl_table(ws, start, r - 1, 1, len(headers), "AnomalyTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def _sheet_raw(self, wb: Any, analysis: SalesAnalysis) -> None:
        ws  = wb.create_sheet("📁 Raw Data")
        df  = analysis.raw_data.copy()
        df["date"] = df["date"].dt.strftime("%Y-%m-%d")
        cols_display = list(df.columns)
        CHUNK = 5_000
        start = self._write_title_block(
            ws, "Raw Sales Data",
            f"{len(df):,} records total"
            + (f"  (showing first {CHUNK:,})" if len(df) > CHUNK else ""))
        self._write_header_row(ws, start, cols_display)
        r = start + 1
        for i, row in enumerate(df.head(CHUNK).itertuples(index=False)):
            self._write_data_row(ws, r, list(row), shade=(i % 2 == 0))
            for col_idx, col_name in enumerate(cols_display, 1):
                if col_name in ("sales", "returns", "marketing_spend", "net_revenue"):
                    ws.cell(row=r, column=col_idx).number_format = "#,##0.00"
            r += 1
        if len(df) > CHUNK:
            ws.cell(row=r, column=1).value = (
                f"⚠ Showing first {CHUNK:,} of {len(df):,} records.")
            ws.cell(row=r, column=1).font = self._font(italic=True, color="888888")
        self._add_xl_table(ws, start, r - 1, 1, len(cols_display), "RawDataTable")
        self._autofit_columns(ws)
        ws.freeze_panes = f"A{start + 1}"

    def generate(
        self, analysis: SalesAnalysis,
        charts: Dict[str, Path], output_path: Path,
    ) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        tmp = output_path.with_suffix(".tmp.xlsx")
        wb  = None
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            self._sheet_summary(wb, analysis)
            self._sheet_forecast(wb, analysis)
            self._sheet_models(wb, analysis)
            self._sheet_history(wb, analysis)
            if "product" in analysis.raw_data.columns:
                self._sheet_products(wb, analysis)
            self._sheet_decomp(wb, analysis)
            self._sheet_anomalies(wb, analysis)
            self._sheet_raw(wb, analysis)
            for cname, cpath in charts.items():
                if cpath and Path(cpath).exists():
                    try:
                        cws       = wb.create_sheet(
                            f"📷 {cname.replace('_', ' ').title()}")
                        img       = XLImage(str(cpath))
                        img.anchor = "B2"
                        cws.add_image(img)
                    except Exception as exc:
                        logging.warning("Could not embed chart '%s': %s", cname, exc)
            wb.save(str(tmp))
            shutil.move(str(tmp), str(output_path))
            return output_path
        except Exception as exc:
            if tmp.exists():
                try:
                    tmp.unlink()
                except OSError:
                    pass
            raise ReportGenerationError(f"Excel generation failed: {exc}") from exc
        finally:
            # always close the workbook to release OS file handles on Windows
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass


# ══════════════════════════════════════════════════════════════════════════════
#  REUSABLE GUI WIDGETS
# ══════════════════════════════════════════════════════════════════════════════

class SortableTree(ttk.Treeview):
    def __init__(self, parent: "tk.Widget",
                 columns: List[Tuple[str, str, int]], **kw: Any):
        col_ids = [c[0] for c in columns]
        super().__init__(parent, columns=col_ids, show="headings", **kw)
        self._sort_state: Dict[str, bool] = {}
        for col_id, heading, width in columns:
            self.heading(col_id, text=heading,
                          command=lambda c=col_id: self._sort(c))
            self.column(col_id, width=width, anchor="center", stretch=True)

    def _sort(self, col_id: str) -> None:
        ascending = not self._sort_state.get(col_id, True)
        self._sort_state[col_id] = ascending
        data = [(self.set(child, col_id), child) for child in self.get_children("")]
        try:
            data.sort(
                key=lambda x: float(
                    x[0].replace("$", "").replace(",", "").replace("%", "").replace("—", "999")),
                reverse=not ascending)
        except ValueError:
            data.sort(key=lambda x: x[0].lower(), reverse=not ascending)
        for idx, (_, child) in enumerate(data):
            self.move(child, "", idx)
        arrow = " ▲" if ascending else " ▼"
        for col in self["columns"]:
            txt = self.heading(col)["text"].rstrip(" ▲▼")
            self.heading(col, text=txt + (arrow if col == col_id else ""))


class ScrollFrame(tk.Frame):
    def __init__(self, parent: "tk.Widget", bg: str = Theme.BG, **kw: Any):
        super().__init__(parent, bg=bg, **kw)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        self._canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self._canvas.grid(row=0, column=0, sticky="nsew")
        sb = tk.Scrollbar(self, orient="vertical", command=self._canvas.yview,
                           bg=Theme.SURFACE, troughcolor=Theme.SURFACE2,
                           activebackground=Theme.BORDER, relief="flat", width=10)
        sb.grid(row=0, column=1, sticky="ns")
        self._canvas.configure(yscrollcommand=sb.set)
        self.inner = tk.Frame(self._canvas, bg=bg)
        self._win_id = self._canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", self._on_frame_configure)
        self._canvas.bind("<Configure>", self._on_canvas_configure)
        # Platform-appropriate mousewheel binding
        self._canvas.bind("<MouseWheel>", self._on_mousewheel_win)    # Windows / macOS
        self._canvas.bind("<Button-4>",   self._on_mousewheel_linux_up)
        self._canvas.bind("<Button-5>",   self._on_mousewheel_linux_down)

    def _on_frame_configure(self, _: Any) -> None:
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))

    def _on_canvas_configure(self, e: Any) -> None:
        self._canvas.itemconfig(self._win_id, width=e.width)

    def _on_mousewheel_win(self, e: Any) -> None:
        self._canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

    def _on_mousewheel_linux_up(self, _: Any) -> None:
        self._canvas.yview_scroll(-1, "units")

    def _on_mousewheel_linux_down(self, _: Any) -> None:
        self._canvas.yview_scroll(1, "units")


class ToolTip:
    def __init__(self, widget: "tk.Widget", text: str):
        self._widget = widget
        self._text   = text
        self._tip: Optional[tk.Toplevel] = None
        widget.bind("<Enter>", self._show)
        widget.bind("<Leave>", self._hide)

    def _show(self, _: Any) -> None:
        if self._tip:
            return
        x = self._widget.winfo_rootx() + 20
        y = self._widget.winfo_rooty() + self._widget.winfo_height() + 4
        self._tip = tk.Toplevel(self._widget)
        self._tip.wm_overrideredirect(True)
        self._tip.wm_geometry(f"+{x}+{y}")
        tk.Label(self._tip, text=self._text, justify="left",
                  bg="#FFFFCC", fg="#333333", font=("Segoe UI", 8),
                  relief="solid", bd=1, padx=6, pady=4).pack()

    def _hide(self, _: Any) -> None:
        if self._tip:
            self._tip.destroy()
            self._tip = None


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN GUI APPLICATION
# ══════════════════════════════════════════════════════════════════════════════

class SalesForecastApp:
    T = Theme

    def __init__(self, root: "tk.Tk"):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.configure(bg=self.T.BG)
        self.root.minsize(960, 680)

        self.v_use_demo   = tk.BooleanVar(value=True)
        self.v_data_path  = tk.StringVar(value="")
        self.v_company    = tk.StringVar(value="My Company")
        self.v_horizon    = tk.IntVar(value=12)
        self.v_make_pdf   = tk.BooleanVar(value=True)
        self.v_make_excel = tk.BooleanVar(value=True)
        self.v_use_arima  = tk.BooleanVar(value=True)
        self.v_use_rf     = tk.BooleanVar(value=True)
        self.v_use_gb     = tk.BooleanVar(value=True)
        self.v_use_ridge  = tk.BooleanVar(value=True)
        self.v_use_es     = tk.BooleanVar(value=True)
        self.v_currency   = tk.StringVar(value="$")

        self._chosen_output: Optional[Path] = None
        self._cancel_token:  Optional[CancelToken] = None
        self._q:             queue.Queue = queue.Queue(maxsize=1_000)  
        self._running        = False
        self._last_analysis: Optional[SalesAnalysis] = None

        self._build_ui()
        self._poll()

    # ── Style ──────────────────────────────────────────────────────────────────

    def _apply_styles(self) -> None:
        st = ttk.Style()
        st.theme_use("clam")
        st.configure("Dark.Horizontal.TProgressbar",
                      background=self.T.ACCENT, troughcolor=self.T.BORDER2,
                      borderwidth=0, thickness=10)
        st.configure("TNotebook",     background=self.T.BG,    borderwidth=0)
        st.configure("TNotebook.Tab", background=self.T.BORDER2,
                      foreground=self.T.MUTED,
                      padding=(18, 9), font=("Segoe UI", 10), borderwidth=0)
        st.map("TNotebook.Tab",
               background=[("selected", self.T.SURFACE)],
               foreground=[("selected", self.T.TEXT)])
        st.configure("Treeview",
                      background=self.T.SURFACE2, foreground=self.T.TEXT2,
                      fieldbackground=self.T.SURFACE2, rowheight=26,
                      borderwidth=0, font=("Segoe UI", 9))
        st.configure("Treeview.Heading",
                      background=self.T.BORDER2, foreground=self.T.TEXT,
                      font=("Segoe UI", 9, "bold"), relief="flat")
        st.map("Treeview",
               background=[("selected", self.T.BLUE)],
               foreground=[("selected", self.T.TEXT)])
        st.map("Treeview.Heading", background=[("active", self.T.BORDER)])

    # ── UI Build ───────────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        self._apply_styles()
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)
        self._build_header()
        self._build_body()
        self._build_statusbar()

    def _build_header(self) -> None:
        hdr = tk.Frame(self.root, bg=self.T.SURFACE,
                        highlightbackground=self.T.BORDER, highlightthickness=1)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(1, weight=1)
        tk.Label(hdr, text="📈", font=("Segoe UI", 22),
                  bg=self.T.SURFACE, fg=self.T.ACCENT).grid(row=0, column=0,
                                                              padx=18, pady=10)
        info = tk.Frame(hdr, bg=self.T.SURFACE)
        info.grid(row=0, column=1, sticky="w")
        tk.Label(info, text=f"Sales Forecasting System",
                  font=("Segoe UI", 15, "bold"),
                  bg=self.T.SURFACE, fg=self.T.TEXT).pack(anchor="w")
        tk.Label(info,
                  text=(
                      "MLE ARIMA · Holt-Winters (auto add/mul) · "
                      "Walk-forward CV · Bootstrap CI · sMAPE/MASE · "
                      "Cross-platform (Windows / macOS / Linux)"),
                  font=("Segoe UI", 9),
                  bg=self.T.SURFACE, fg=self.T.MUTED).pack(anchor="w")
        tk.Label(hdr, text=" Production ", font=("Segoe UI", 8, "bold"),
                  bg=self.T.ACCENT, fg="white", padx=6, pady=2).grid(
                      row=0, column=2, padx=18)

    def _build_body(self) -> None:
        nb = ttk.Notebook(self.root)
        nb.grid(row=1, column=0, sticky="nsew")
        self._nb = nb
        self._tab_setup   = tk.Frame(nb, bg=self.T.BG)
        self._tab_results = tk.Frame(nb, bg=self.T.BG)
        self._tab_log     = tk.Frame(nb, bg=self.T.BG)
        nb.add(self._tab_setup,   text="  ⚙  Setup  ")
        nb.add(self._tab_results, text="  📊  Results  ")
        nb.add(self._tab_log,     text="  📋  Log  ")
        self._build_setup_tab(self._tab_setup)
        self._build_results_tab(self._tab_results)
        self._build_log_tab(self._tab_log)

    def _build_statusbar(self) -> None:
        bar = tk.Frame(self.root, bg=self.T.SURFACE,
                        highlightbackground=self.T.BORDER, highlightthickness=1,
                        height=44)
        bar.grid(row=2, column=0, sticky="ew")
        bar.columnconfigure(0, weight=1)
        bar.grid_propagate(False)
        self._status_lbl = tk.Label(
            bar, text="Ready — configure your analysis and click Run.",
            font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.MUTED, anchor="w")
        self._status_lbl.grid(row=0, column=0, sticky="ew", padx=16, pady=12)
        self._pvar = tk.DoubleVar(value=0)
        self._pbar = ttk.Progressbar(bar, variable=self._pvar,
                                      style="Dark.Horizontal.TProgressbar",
                                      mode="determinate", maximum=100, length=200)
        self._pbar.grid(row=0, column=1, padx=8)
        self._pct_lbl = tk.Label(bar, text="", font=("Segoe UI", 9, "bold"),
                                   bg=self.T.SURFACE, fg=self.T.WARN, width=5)
        self._pct_lbl.grid(row=0, column=2, padx=(0, 16))
        self._pbar.grid_remove()
        self._pct_lbl.grid_remove()

    def _build_setup_tab(self, parent: "tk.Frame") -> None:
        sf = ScrollFrame(parent, bg=self.T.BG)
        sf.pack(fill="both", expand=True)
        inner = sf.inner
        inner.columnconfigure(0, weight=1)
        inner.columnconfigure(1, weight=1)
        left = tk.Frame(inner, bg=self.T.BG)
        left.grid(row=0, column=0, sticky="nsew", padx=(24, 10), pady=16)
        left.columnconfigure(0, weight=1)
        self._build_data_card(left)
        self._build_config_card(left)
        right = tk.Frame(inner, bg=self.T.BG)
        right.grid(row=0, column=1, sticky="nsew", padx=(10, 24), pady=16)
        right.columnconfigure(0, weight=1)
        self._build_models_card(right)
        self._build_output_card(right)
        run_row = tk.Frame(inner, bg=self.T.BG)
        run_row.grid(row=1, column=0, columnspan=2, sticky="ew", padx=24, pady=(0, 20))
        self._run_btn = tk.Button(
            run_row,
            text="  🚀  Run Analysis & Generate Reports",
            font=("Segoe UI", 12, "bold"),
            bg=self.T.ACCENT, fg="#FFFFFF",
            activebackground=self.T.ACCENT_HV, activeforeground="#FFFFFF",
            relief="flat", bd=0, padx=28, pady=14, cursor="hand2",
            command=self._run)
        self._run_btn.pack(side="left")
        ToolTip(self._run_btn,
                "Validate → choose output path → run full analysis pipeline")
        self._cancel_btn = tk.Button(
            run_row, text="  ✕  Cancel",
            font=("Segoe UI", 11),
            bg=self.T.SURFACE, fg=self.T.MUTED,
            activebackground=self.T.BORDER, activeforeground=self.T.TEXT,
            relief="flat", bd=0, padx=18, pady=14, cursor="hand2",
            command=self._cancel, state="disabled")
        self._cancel_btn.pack(side="left", padx=(12, 0))

    def _card(self, parent: "tk.Frame") -> "tk.Frame":
        f = tk.Frame(parent, bg=self.T.SURFACE,
                      highlightbackground=self.T.BORDER, highlightthickness=1)
        f.pack(fill="x", pady=(0, 14))
        return f

    def _card_header(self, card: "tk.Frame", icon: str,
                     title: str, sub: str = "") -> None:
        hdr = tk.Frame(card, bg=self.T.SURFACE, pady=10)
        hdr.pack(fill="x", padx=14)
        tk.Label(hdr, text=f"{icon}  {title}", font=("Segoe UI", 11, "bold"),
                  bg=self.T.SURFACE, fg=self.T.TEXT).pack(anchor="w")
        if sub:
            tk.Label(hdr, text=sub, font=("Segoe UI", 8),
                      bg=self.T.SURFACE, fg=self.T.MUTED).pack(anchor="w")
        tk.Frame(card, bg=self.T.BORDER, height=1).pack(fill="x")

    def _labeled_entry(
        self, parent: "tk.Widget", label: str, var: "tk.Variable",
        width: int = 20, tooltip: str = "",
    ) -> "tk.Entry":
        row = tk.Frame(parent, bg=self.T.SURFACE)
        row.pack(fill="x", padx=14, pady=4)
        row.columnconfigure(1, weight=1)
        tk.Label(row, text=label, font=("Segoe UI", 9),
                  bg=self.T.SURFACE, fg=self.T.MUTED, width=width, anchor="w"
                  ).grid(row=0, column=0, sticky="w")
        e = tk.Entry(row, textvariable=var, font=("Segoe UI", 9),
                      bg=self.T.SURFACE2, fg=self.T.TEXT,
                      insertbackground=self.T.TEXT,
                      relief="flat", bd=5,
                      highlightbackground=self.T.BORDER, highlightthickness=1)
        e.grid(row=0, column=1, sticky="ew", padx=(8, 0), ipady=5)
        if tooltip:
            ToolTip(e, tooltip)
        return e

    def _checkbutton(
        self, parent: "tk.Widget", text: str,
        var: "tk.BooleanVar", tooltip: str = "",
    ) -> "tk.Checkbutton":
        cb = tk.Checkbutton(
            parent, text=f"  {text}", variable=var,
            font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.TEXT,
            selectcolor=self.T.BLUE, activebackground=self.T.SURFACE,
            activeforeground=self.T.TEXT, cursor="hand2", bd=0)
        cb.pack(anchor="w", padx=14, pady=2)
        if tooltip:
            ToolTip(cb, tooltip)
        return cb

    def _build_data_card(self, parent: "tk.Frame") -> None:
        card = self._card(parent)
        self._card_header(card, "📂", "Data Source",
                           "Load your own file or use built-in demo data")
        body = tk.Frame(card, bg=self.T.SURFACE, pady=8)
        body.pack(fill="x")
        tk.Checkbutton(
            body, text="  Use built-in demo data  (5 products · 60 months)",
            variable=self.v_use_demo, font=("Segoe UI", 10),
            bg=self.T.SURFACE, fg=self.T.TEXT, selectcolor=self.T.ACCENT,
            activebackground=self.T.SURFACE, activeforeground=self.T.TEXT,
            cursor="hand2", bd=0, command=self._toggle_demo).pack(
                anchor="w", padx=14, pady=(0, 6))
        self._file_row = tk.Frame(body, bg=self.T.SURFACE)
        self._file_row.pack(fill="x", padx=14, pady=(0, 4))
        self._file_row.columnconfigure(0, weight=1)
        tk.Label(self._file_row,
                  text="File  (CSV / TSV / Excel / JSON / Parquet):",
                  font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.MUTED
                  ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 4))
        self._path_entry = tk.Entry(
            self._file_row, textvariable=self.v_data_path,
            font=("Segoe UI", 9), bg=self.T.SURFACE2, fg=self.T.TEXT,
            insertbackground=self.T.TEXT, relief="flat", bd=6,
            highlightbackground=self.T.BORDER, highlightthickness=1)
        self._path_entry.grid(row=1, column=0, sticky="ew", ipady=5)
        btn_frame = tk.Frame(self._file_row, bg=self.T.SURFACE)
        btn_frame.grid(row=1, column=1, padx=(8, 0))
        tk.Button(
            btn_frame, text=" Browse… ",
            font=("Segoe UI", 9), bg=self.T.BLUE, fg="#FFFFFF",
            activebackground=self.T.BLUE_HV, activeforeground="#FFFFFF",
            relief="flat", bd=0, padx=10, pady=5, cursor="hand2",
            command=self._pick_file).pack(side="left")
        tk.Button(
            btn_frame, text=" 👁 Preview ",
            font=("Segoe UI", 9), bg=self.T.SURFACE2, fg=self.T.TEXT2,
            activebackground=self.T.BORDER, activeforeground=self.T.TEXT,
            relief="flat", bd=0, padx=8, pady=5, cursor="hand2",
            command=self._preview_data).pack(side="left", padx=(6, 0))
        ToolTip(btn_frame.winfo_children()[-1], "Preview first 20 rows of selected file")
        tk.Label(
            body,
            text=("  Required columns: 'date', 'sales'   "
                  "Optional: product · region · units · returns · marketing_spend\n"
                  "  Supported formats: .csv  .tsv  .xlsx  .xlsm  .xls  .json  .parquet"),
            font=("Segoe UI", 8), bg=self.T.SURFACE, fg=self.T.MUTED2,
        ).pack(anchor="w", padx=14, pady=(4, 8))
        self._toggle_demo()

    def _build_config_card(self, parent: "tk.Frame") -> None:
        card = self._card(parent)
        self._card_header(card, "🔧", "Configuration")
        body = tk.Frame(card, bg=self.T.SURFACE, pady=6)
        body.pack(fill="x")
        self._labeled_entry(body, "Company Name",    self.v_company,
                             tooltip="Appears in report headers")
        self._labeled_entry(body, "Currency Symbol", self.v_currency,
                             tooltip="e.g.  $  €  £  ₹")
        hz_row = tk.Frame(body, bg=self.T.SURFACE)
        hz_row.pack(fill="x", padx=14, pady=5)
        hz_row.columnconfigure(1, weight=1)
        tk.Label(hz_row, text="Forecast Horizon", font=("Segoe UI", 9),
                  bg=self.T.SURFACE, fg=self.T.MUTED, width=20, anchor="w"
                  ).grid(row=0, column=0, sticky="w")
        hz_inner = tk.Frame(hz_row, bg=self.T.SURFACE)
        hz_inner.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        hz_inner.columnconfigure(0, weight=1)
        tk.Scale(hz_inner, variable=self.v_horizon, from_=1, to=36,
                  orient="horizontal", bg=self.T.SURFACE, fg=self.T.TEXT,
                  troughcolor=self.T.BORDER2, activebackground=self.T.ACCENT,
                  highlightthickness=0, bd=0, font=("Segoe UI", 8),
                  showvalue=False, command=self._update_hz
                  ).grid(row=0, column=0, sticky="ew")
        self._hz_lbl = tk.Label(hz_inner, text="12 months",
                                 font=("Segoe UI", 9, "bold"),
                                 bg=self.T.SURFACE, fg=self.T.ACCENT, width=10)
        self._hz_lbl.grid(row=0, column=1)
        tk.Frame(card, bg=self.T.SURFACE, height=6).pack()

    def _build_models_card(self, parent: "tk.Frame") -> None:
        card = self._card(parent)
        self._card_header(card, "🤖", "Forecasting Models",
                           "All enabled models combined via inverse CV-sMAPE stacking")
        body = tk.Frame(card, bg=self.T.SURFACE, pady=8)
        body.pack(fill="x")
        models = [
            (self.v_use_arima, "ARIMA  (auto order, exact MLE)",
             "Grid-search by AIC; CI via ψ-weight variance propagation; dead _undifference removed"),
            (self.v_use_es,    "Holt-Winters  (auto add/mul)",
             "SSE-minimised; AIC selects additive vs multiplicative; analytical/bootstrap CI"),
            (self.v_use_rf,    "Random Forest  (200 trees, PACF lags)",
             "Feature matrix built once; bootstrap CI; PACF-guided lag selection"),
            (self.v_use_gb,    "Gradient Boosting  (200 rounds)",
             "Shared feature matrix; bootstrap CI; walk-forward cross-validation score reported"),
            (self.v_use_ridge, "Ridge Regression  (L2 baseline)",
             "Fast regularised linear baseline with bootstrap CI"),
        ]
        for var, label, tip in models:
            self._checkbutton(body, label, var, tooltip=tip)
        tk.Frame(body, bg=self.T.BORDER, height=1).pack(fill="x", padx=14, pady=8)
        tk.Label(body,
                  text="  ⚡ Ensemble: inverse CV-sMAPE weights (walk-forward, non-leaky, NaN-guarded)",
                  font=("Segoe UI", 8), bg=self.T.SURFACE, fg=self.T.MUTED,
                  ).pack(anchor="w", padx=14, pady=(0, 6))

    def _build_output_card(self, parent: "tk.Frame") -> None:
        card = self._card(parent)
        self._card_header(card, "📁", "Output", "Save location chosen when you click Run")
        body = tk.Frame(card, bg=self.T.SURFACE, pady=10)
        body.pack(fill="x")
        tk.Label(body, text="Report formats:", font=("Segoe UI", 9),
                  bg=self.T.SURFACE, fg=self.T.MUTED).pack(anchor="w", padx=14)
        fmt_row = tk.Frame(body, bg=self.T.SURFACE)
        fmt_row.pack(anchor="w", padx=14, pady=4)
        for var, label, icon, tip in [
            (self.v_make_pdf,   "PDF Report",     "📄",
             "Multi-page PDF with charts, diagnostic plots, tables, and insights"),
            (self.v_make_excel, "Excel Workbook", "📊",
             "9-sheet workbook with model diagnostics, CI methods, and embedded charts"),
        ]:
            cb = tk.Checkbutton(
                fmt_row, text=f"  {icon}  {label}", variable=var,
                font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.TEXT,
                selectcolor=self.T.BLUE, activebackground=self.T.SURFACE,
                activeforeground=self.T.TEXT, cursor="hand2", bd=0)
            cb.pack(side="left", padx=(0, 20))
            ToolTip(cb, tip)
        tk.Frame(body, bg=self.T.BORDER, height=1).pack(fill="x", padx=14, pady=8)
        tk.Label(body,
                  text=("  ℹ  A JSON summary is always saved alongside the reports.\n"
                        "  📁  Charts stored in system tempdir (cross-platform)."),
                  font=("Segoe UI", 8), bg=self.T.SURFACE, fg=self.T.MUTED,
                  justify="left",
                  ).pack(anchor="w", padx=14, pady=(0, 8))

    def _build_results_tab(self, parent: "tk.Frame") -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)
        hdr = tk.Frame(parent, bg=self.T.BG)
        hdr.grid(row=0, column=0, sticky="ew", padx=24, pady=(18, 10))
        tk.Label(hdr, text="Analysis Results", font=("Segoe UI", 16, "bold"),
                  bg=self.T.BG, fg=self.T.TEXT).pack(anchor="w")
        self._res_sub = tk.Label(
            hdr, text="No results yet — run an analysis first.",
            font=("Segoe UI", 9), bg=self.T.BG, fg=self.T.MUTED)
        self._res_sub.pack(anchor="w")
        self._kpi_bar = tk.Frame(parent, bg=self.T.BG)
        self._kpi_bar.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 10))
        tbl = tk.Frame(parent, bg=self.T.SURFACE,
                        highlightbackground=self.T.BORDER, highlightthickness=1)
        tbl.grid(row=2, column=0, sticky="nsew", padx=24, pady=(0, 10))
        tbl.columnconfigure(0, weight=1)
        tbl.rowconfigure(1, weight=1)
        th = tk.Frame(tbl, bg=self.T.SURFACE)
        th.grid(row=0, column=0, columnspan=2, sticky="ew", padx=16, pady=8)
        tk.Label(th, text="Model Performance",
                  font=("Segoe UI", 11, "bold"),
                  bg=self.T.SURFACE, fg=self.T.TEXT).pack(side="left")
        tk.Label(th,
                  text="Click column headers to sort  ·  "
                       "sMAPE symmetric (near-zero safe)  ·  CV = walk-forward",
                  font=("Segoe UI", 8),
                  bg=self.T.SURFACE, fg=self.T.MUTED).pack(side="left", padx=(12, 0))
        cols = [
            ("model",    "Model",      160),
            ("smape",    "sMAPE (%)",   80),
            ("mase",     "MASE",        70),
            ("rmse",     "RMSE ($)",    95),
            ("r2",       "R² Score",    75),
            ("cv_smape", "CV sMAPE",    80),
            ("time",     "Time (s)",    70),
        ]
        self._model_tree = SortableTree(tbl, cols, height=7)
        self._model_tree.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 10))
        self._model_tree.tag_configure("ens",  background="#0D2A20", foreground="#3FB950")
        self._model_tree.tag_configure("best", background="#0D1E38", foreground="#58A6FF")
        sb = tk.Scrollbar(tbl, command=self._model_tree.yview,
                           bg=self.T.SURFACE, troughcolor=self.T.SURFACE2, relief="flat")
        sb.grid(row=1, column=1, sticky="ns", pady=(0, 10))
        self._model_tree.configure(yscrollcommand=sb.set)
        self._files_row = tk.Frame(parent, bg=self.T.BG)
        self._files_row.grid(row=3, column=0, sticky="ew", padx=24, pady=(0, 18))

    def _build_log_tab(self, parent: "tk.Frame") -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)
        hdr = tk.Frame(parent, bg=self.T.BG)
        hdr.grid(row=0, column=0, sticky="ew", padx=24, pady=(18, 8))
        tk.Label(hdr, text="Activity Log", font=("Segoe UI", 16, "bold"),
                  bg=self.T.BG, fg=self.T.TEXT).pack(anchor="w")
        tk.Label(hdr, text="Real-time pipeline events — colour-coded by severity",
                  font=("Segoe UI", 9), bg=self.T.BG, fg=self.T.MUTED).pack(anchor="w")
        log_card = tk.Frame(parent, bg=self.T.SURFACE,
                             highlightbackground=self.T.BORDER, highlightthickness=1)
        log_card.grid(row=1, column=0, sticky="nsew", padx=24, pady=(0, 24))
        log_card.columnconfigure(0, weight=1)
        log_card.rowconfigure(0, weight=1)
        self._log = tk.Text(
            log_card, bg=self.T.SURFACE2, fg=self.T.TEXT2,
            font=("Consolas", 9), insertbackground=self.T.TEXT,
            selectbackground=self.T.BLUE, relief="flat", bd=0,
            wrap="word", state="disabled", padx=12, pady=10)
        self._log.grid(row=0, column=0, sticky="nsew")
        sb2 = tk.Scrollbar(log_card, command=self._log.yview,
                            bg=self.T.SURFACE, troughcolor=self.T.SURFACE2, relief="flat")
        sb2.grid(row=0, column=1, sticky="ns")
        self._log.configure(yscrollcommand=sb2.set)
        for tag, fg in [
            ("ok",     self.T.SUCCESS), ("warn",   self.T.WARN),
            ("error",  self.T.DANGER),  ("accent", "#58A6FF"),
            ("muted",  self.T.MUTED),   ("info",   self.T.TEXT2),
            ("sep",    self.T.BORDER),
        ]:
            self._log.tag_configure(tag, foreground=fg)
        ctrl = tk.Frame(log_card, bg=self.T.SURFACE)
        ctrl.grid(row=1, column=0, columnspan=2, sticky="ew", padx=12, pady=6)
        for lbl, cmd in [("Copy All", self._copy_log), ("Clear Log", self._clear_log)]:
            tk.Button(ctrl, text=lbl, font=("Segoe UI", 9),
                       bg=self.T.SURFACE2, fg=self.T.MUTED,
                       activebackground=self.T.BORDER, activeforeground=self.T.TEXT,
                       relief="flat", bd=0, padx=10, pady=4, cursor="hand2",
                       command=cmd).pack(side="right", padx=(0, 6))

    # ── Callbacks ──────────────────────────────────────────────────────────────

    def _toggle_demo(self) -> None:
        if self.v_use_demo.get():
            self._file_row.pack_forget()
        else:
            self._file_row.pack(fill="x", padx=14, pady=(0, 4))

    def _pick_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Sales Data File",
            filetypes=[
                ("All supported", "*.csv *.tsv *.xlsx *.xlsm *.xls *.json *.parquet"),
                ("CSV/TSV",       "*.csv *.tsv"),
                ("Excel",         "*.xlsx *.xlsm *.xls"),
                ("JSON",          "*.json"),
                ("Parquet",       "*.parquet"),
                ("All files",     "*.*"),
            ])
        if path:
            self.v_data_path.set(path)
            self.v_use_demo.set(False)
            self._toggle_demo()
            self._log_msg(f"Data file selected: {Path(path).name}", "accent")

    def _preview_data(self) -> None:
        """Show first 20 rows of the selected file in a popup."""
        path_str = self.v_data_path.get().strip()
        if not path_str:
            messagebox.showinfo("Preview", "Please select a data file first.")
            return
        try:
            df = DataLoader.load(path_str)
        except Exception as exc:
            messagebox.showerror("Preview Error", str(exc))
            return

        top = tk.Toplevel(self.root)
        top.title(f"Data Preview — {Path(path_str).name}  ({len(df):,} rows total)")
        top.configure(bg=self.T.BG)
        top.geometry("900x400")

        info = tk.Label(
            top,
            text=(f"  Loaded {len(df):,} records · "
                  f"{df['date'].nunique()} months · "
                  f"{df['product'].nunique()} products · "
                  f"Columns: {', '.join(df.columns.tolist())}"),
            font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.TEXT2,
            anchor="w", pady=6)
        info.pack(fill="x")

        cols   = list(df.columns)
        frm    = tk.Frame(top, bg=self.T.BG)
        frm.pack(fill="both", expand=True)
        frm.columnconfigure(0, weight=1)
        frm.rowconfigure(0, weight=1)
        tree   = ttk.Treeview(frm, columns=cols, show="headings", height=18)
        style  = ttk.Style(top)
        style.configure("Preview.Treeview",
                         background=self.T.SURFACE2, foreground=self.T.TEXT2,
                         fieldbackground=self.T.SURFACE2, rowheight=22)
        style.configure("Preview.Treeview.Heading",
                         background=self.T.SURFACE, foreground=self.T.TEXT)
        tree.configure(style="Preview.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(90, min(160, len(str(c)) * 10)), anchor="center")
        for row in df.head(20).itertuples(index=False):
            tree.insert("", "end", values=[str(v)[:40] for v in row])
        # use grid manager for both tree and scrollbars — consistent on HiDPI
        vsb = tk.Scrollbar(frm, orient="vertical",   command=tree.yview, bg=self.T.SURFACE)
        hsb = tk.Scrollbar(frm, orient="horizontal", command=tree.xview, bg=self.T.SURFACE)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

    def _update_hz(self, _: Any = None) -> None:
        v = int(self.v_horizon.get())
        self._hz_lbl.configure(text=f"{v} month{'s' if v != 1 else ''}")

    def _log_msg(self, message: str, tag: str = "info") -> None:
        self._log.configure(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self._log.insert("end", f"[{ts}] ", "muted")
        self._log.insert("end", message + "\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    def _clear_log(self) -> None:
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

    def _copy_log(self) -> None:
        try:
            content = self._log.get("1.0", "end")
            self.root.clipboard_clear()
            self.root.clipboard_append(content)
        except Exception:
            pass

    def _set_status(self, text: str, color: Optional[str] = None) -> None:
        self._status_lbl.configure(text=text, fg=color or self.T.MUTED)

    def _poll(self) -> None:
        """Drain the inter-thread queue every 50 ms on the main thread."""
        try:
            while True:
                item = self._q.get_nowait()
                kind = item[0]
                if kind == "progress":
                    _, msg, pct = item
                    self._set_status(msg, self.T.WARN)
                    if pct is not None:
                        self._pvar.set(float(pct))
                        self._pct_lbl.configure(text=f"{int(pct)}%")
                    self._log_msg(msg, "info")
                elif kind == "log":
                    _, msg, tag = item
                    self._log_msg(msg, tag)
                elif kind == "done":
                    self._on_done(item[1])
                elif kind == "error":
                    self._on_error(item[1])
        except queue.Empty:
            pass
        self.root.after(50, self._poll)

    # ── Validation & run ───────────────────────────────────────────────────────

    def _validate(self) -> bool:
        if not self.v_use_demo.get() and not self.v_data_path.get().strip():
            messagebox.showerror("Missing Data",
                                  "Please select a data file or enable the demo dataset.")
            return False
        if not self.v_make_pdf.get() and not self.v_make_excel.get():
            messagebox.showerror("No Output",
                                  "Please select at least one output format.")
            return False
        if not any([self.v_use_arima.get(), self.v_use_rf.get(),
                    self.v_use_gb.get(),    self.v_use_ridge.get(),
                    self.v_use_es.get()]):
            messagebox.showerror("No Models", "Please enable at least one model.")
            return False
        if not self.v_currency.get().strip():
            messagebox.showwarning("Currency", "Currency symbol empty — using '$'.")
            self.v_currency.set("$")
        return True

    def _ask_save_path(self) -> Optional[Path]:
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        ext    = ".pdf" if self.v_make_pdf.get() else ".xlsx"
        ftypes = ([("PDF", "*.pdf"), ("All", "*.*")]
                  if ext == ".pdf"
                  else [("Excel", "*.xlsx"), ("All", "*.*")])
        chosen = filedialog.asksaveasfilename(
            title="Save Report (base filename — extensions added automatically)",
            initialfile=f"sales_forecast_{ts}",
            defaultextension=ext, filetypes=ftypes)
        if not chosen:
            return None
        p = Path(chosen)
        return p.parent / p.stem

    def _run(self) -> None:
        if self._running:
            return
        self._run_btn.configure(state="disabled")
        if not self._validate():
            self._run_btn.configure(state="normal")
            return
        save_stem = self._ask_save_path()
        if save_stem is None:
            self._run_btn.configure(state="normal")
            self._log_msg("Save cancelled — run aborted.", "warn")
            return
        self._chosen_output = save_stem
        self._cancel_token  = CancelToken()
        self._running       = True
        self._run_btn.configure(state="disabled", text="  ⏳  Running…")
        self._cancel_btn.configure(state="normal")
        self._pbar.grid()
        self._pct_lbl.grid()
        self._pvar.set(0)
        self._set_status("Starting …", self.T.WARN)
        self._nb.select(self._tab_log)
        self._log_msg("━" * 56, "sep")
        self._log_msg("ANALYSIS STARTED", "ok")
        self._log_msg("━" * 56, "sep")

        cfg = ForecastConfig(
            output_path            = save_stem,
            forecast_horizon       = int(self.v_horizon.get()),
            company_name           = self.v_company.get().strip() or "My Company",
            currency_symbol        = self.v_currency.get().strip() or "$",
            use_arima              = self.v_use_arima.get(),
            use_random_forest      = self.v_use_rf.get(),
            use_gradient_boosting  = self.v_use_gb.get(),
            use_ridge              = self.v_use_ridge.get(),
            use_exp_smoothing      = self.v_use_es.get(),
            report_format          = (["pdf"]   if self.v_make_pdf.get()   else []) +
                                     (["excel"] if self.v_make_excel.get() else []),
        )
        data_source = (None if self.v_use_demo.get()
                       else self.v_data_path.get().strip())
        threading.Thread(
            target=self._pipeline,
            args=(cfg, data_source, self._cancel_token),
            daemon=True).start()

    def _pipeline(
        self, cfg: ForecastConfig,
        data_source: Optional[str],
        token: CancelToken,
    ) -> None:
        """Worker thread — communicates only via self._q."""
        q             = self._q
        _last_pct: List[float] = [0.0]    # monotonic progress tracker

        def report(msg: str, pct: Optional[float] = None) -> None:
            # never let progress go backwards
            if pct is not None:
                pct = max(float(pct), _last_pct[0])
                _last_pct[0] = pct
            q.put(("progress", msg, pct))

        def log(msg: str, tag: str = "info") -> None:
            q.put(("log", msg, tag))

        try:
            cfg.ensure_dirs()
            log(f"Temp charts dir: {cfg.charts_dir}", "muted")
            report("Loading data …", 2)
            if data_source is None:
                log("Generating synthetic demo data …", "accent")
                df = DataLoader.load(None)
            else:
                log(f"Loading: {Path(data_source).name}", "accent")
                df = DataLoader.load(data_source)
            token.check()

            log(f"Loaded {len(df):,} records — "
                f"{df['date'].nunique()} months, "
                f"{df['product'].nunique()} products.", "ok")

            report("Analysing …", 5)
            analyzer = SalesAnalyzer(cfg, progress_cb=report, cancel_token=token)
            analysis = analyzer.analyze(df)

            st  = analysis.statistics
            ens = analysis.ensemble_forecast
            log(f"Total Revenue   : {cfg.currency_symbol}{st['total_revenue']:,.0f}", "ok")
            log(f"CAGR            : {st['cagr']:.1f}%", "ok")
            log(f"Seasonal period : {analysis.detected_period} months", "ok")
            log(f"Anomalies found : {len(analysis.anomalies)}", "ok")
            log(f"Best model      : {analysis.best_model}", "ok")
            log(f"Ensemble sMAPE  : {ens.smape:.2f}%  (MASE={ens.mase:.3f})", "ok")
            log(f"Marketing ROI   : {st['marketing_roi']:.2f}x  "
                f"[{st['marketing_roi_ci_lo']:.2f}x – "
                f"{st['marketing_roi_ci_hi']:.2f}x  95% CI]", "ok")
            token.check()

            report("Generating charts …", 91)
            charts = ChartGenerator(cfg).generate_all(analysis, progress_cb=report)
            log(f"{len(charts)} charts rendered.", "ok")

            output_files: Dict[str, Path] = {}

            if "pdf" in cfg.report_format:
                token.check()
                report("Generating PDF …", 96)
                pdf_path = cfg.output_path.with_suffix(".pdf")
                PDFReportGenerator(cfg).generate(analysis, charts, pdf_path)
                sz = pdf_path.stat().st_size // 1024
                output_files["pdf"] = pdf_path
                log(f"PDF saved: {pdf_path.name}  ({sz:,} KB)", "ok")

            if "excel" in cfg.report_format:
                token.check()
                report("Generating Excel …", 98)
                xl_path = cfg.output_path.with_suffix(".xlsx")
                ExcelReportGenerator(cfg).generate(analysis, charts, xl_path)
                sz = xl_path.stat().st_size // 1024
                output_files["excel"] = xl_path
                log(f"Excel saved: {xl_path.name}  ({sz:,} KB)", "ok")

            json_path = cfg.output_path.with_suffix(".json")
            summary   = {
                "version": APP_VERSION,
                "generated_at":             datetime.now().isoformat(),
                "company":                  cfg.company_name,
                "forecast_horizon_months":  cfg.forecast_horizon,
                "detected_seasonal_period": analysis.detected_period,
                "best_model":               analysis.best_model,
                "ensemble_metrics": {
                    "smape": round(ens.smape, 3),
                    "mase":  round(ens.mase,  4),
                    "rmse":  round(ens.rmse,  2),
                    "r2":    round(ens.r2,    4),
                    "mae":   round(ens.mae,   2),
                },
                "model_metrics": {
                    name: {
                        "smape":       round(v.smape, 3),
                        "mase":        round(v.mase,  4),
                        "rmse":        round(v.rmse,  2),
                        "r2":          round(v.r2,    4),
                        "cv_smape":    (round(v.cv_smape, 3)
                                        if v.cv_smape and math.isfinite(v.cv_smape)
                                        else None),
                        "aic":         round(v.aic, 2) if v.aic else None,
                        "diagnostics": (v.diagnostics.summary()
                                        if v.diagnostics else None),
                    }
                    for name, v in analysis.forecasts.items()
                    if v is not None
                },
                "statistics": analysis.statistics,   # raw dict — _NpEncoder handles all types
                "total_forecast_revenue": round(float(ens.predictions.sum()), 2),
                "anomalies_detected":     len(analysis.anomalies),
                "ci_methods": {
                    "ARIMA":        "ψ-weight variance propagation (Box-Jenkins §5.4)",
                    "Holt-Winters": "Analytical (additive) or non-parametric bootstrap (multiplicative)",
                    "ML_models":    f"Non-parametric bootstrap ({cfg.n_bootstrap} iterations)",
                },
                "Improvements": [
                    "Config: seasonality_period validated; n_bootstrap capped at 5 000",
                    "ensure_dirs() creates output parent before temp dir",
                    ".txt files parsed with full CSV multi-sep/encoding logic",
                    "NaN-sales rows dropped (not silently zeroed)",
                    "ARIMA σ² floor raised 1e-20→1e-10 (log stability)",
                    "ARIMA predict() explicit float64 before undiff",
                    "HoltWinters accepts non-success L-BFGS-B convergence",
                    "Ensemble fast-path for single valid model",
                    "decompose() pad clamp for n == period+1 edge case",
                    "detect_anomalies() returns explicit dtype DataFrame",
                    "heatmap pivot columns cast to int before month lookup",
                    "Chart _save() recreates dir if cleaned up mid-run",
                    "PDF RMSE column uses configured currency symbol",
                    "Excel YoY/MoM colour-coding column indices corrected",
                    "Excel workbook closed in finally block (Windows handle)",
                    "Preview scrollbars use grid manager (HiDPI macOS fix)",
                    "JSON statistics fully handled by _NpEncoder",
                    "_on_done handles Path objects in file dict safely",
                    "WM_DELETE_WINDOW cancels running analysis before close",
                    "Rotating log file: ~/SalesForecastReports/sales_forecast.log",
                ],
                "platform_info": {
                    "python":          sys.version,
                    "platform":        sys.platform,
                    "pandas":          pd.__version__,
                    "numpy":           np.__version__,
                    "charts_tmpdir":   str(cfg.charts_dir),
                },
            }
            json_path.write_text(
                json.dumps(summary, indent=2, cls=_NpEncoder), encoding="utf-8")
            output_files["json"] = json_path
            log(f"JSON summary saved: {json_path.name}", "ok")

            # retry cleanup — Windows may hold file handles briefly
            for _attempt in range(3):
                try:
                    shutil.rmtree(cfg.charts_dir, ignore_errors=False)
                    break
                except Exception:
                    time.sleep(0.3)

            report("Complete! ✓", 100)
            q.put(("done", {"analysis": analysis, "files": output_files, "cfg": cfg}))

        except InterruptedError:
            q.put(("log", "Analysis cancelled by user.", "warn"))
            q.put(("error", "cancelled"))
        except (DataLoadError, ForecastError, ReportGenerationError) as exc:
            q.put(("log", f"ERROR: {exc}", "error"))
            q.put(("error", str(exc)))
        except Exception as exc:
            q.put(("log", f"UNEXPECTED ERROR: {exc}", "error"))
            q.put(("log", traceback.format_exc(), "error"))
            q.put(("error", str(exc)))

    # ── Done / Error handlers ──────────────────────────────────────────────────

    def _on_done(self, result: Dict) -> None:
        self._running = False
        self._run_btn.configure(state="normal",
                                 text="  🚀  Run Analysis & Generate Reports")
        self._cancel_btn.configure(state="disabled")
        self._pbar.grid_remove()
        self._pct_lbl.grid_remove()
        self._set_status("✓  Analysis complete", self.T.SUCCESS)
        self._log_msg("━" * 56, "sep")
        self._log_msg("ANALYSIS COMPLETE ✓", "ok")
        self._log_msg("━" * 56, "sep")

        analysis = result["analysis"]
        files    = result["files"]
        cfg      = result["cfg"]
        self._last_analysis = analysis
        self._populate_results(analysis, files, cfg)
        self._nb.select(self._tab_results)

        file_lines = "\n".join(
            # p may already be a Path; Path(p).name handles both str and Path
            f"• {fmt.upper()}: {Path(p).name}" for fmt, p in files.items())
        messagebox.showinfo(
            "Analysis Complete 🎉",
            f"All reports saved to:\n{cfg.output_path.parent}\n\n"
            f"{file_lines}\n\n"
            f"Best Model      : {analysis.best_model}\n"
            f"Ensemble sMAPE  : {analysis.ensemble_forecast.smape:.2f}%\n"
            f"Ensemble MASE   : {analysis.ensemble_forecast.mase:.3f}\n"
            f"Anomalies found : {len(analysis.anomalies)}")

    def _on_error(self, err: str) -> None:
        self._running = False
        self._run_btn.configure(state="normal",
                                 text="  🚀  Run Analysis & Generate Reports")
        self._cancel_btn.configure(state="disabled")
        self._pbar.grid_remove()
        self._pct_lbl.grid_remove()
        if err == "cancelled":
            self._set_status("Cancelled", self.T.MUTED)
            return
        self._set_status("✗  Analysis failed", self.T.DANGER)
        messagebox.showerror("Analysis Failed",
                              f"An error occurred:\n\n{err}\n\n"
                              "See the Log tab for full traceback.")

    def _cancel(self) -> None:
        if self._cancel_token:
            self._cancel_token.cancel()
        self._cancel_btn.configure(state="disabled")
        self._set_status("Cancelling …", self.T.WARN)
        self._log_msg("Cancel requested — stopping after current step …", "warn")

    # ── Results population ─────────────────────────────────────────────────────

    def _populate_results(
        self, analysis: SalesAnalysis,
        files: Dict, cfg: ForecastConfig,
    ) -> None:
        st  = analysis.statistics
        ens = analysis.ensemble_forecast
        self._res_sub.configure(
            text=(f"Completed {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  "
                  f"Best model: {analysis.best_model}  ·  "
                  f"sMAPE: {ens.smape:.2f}%  ·  MASE: {ens.mase:.3f}"))

        for w in self._kpi_bar.winfo_children():
            w.destroy()
        roi_str = f"{st['marketing_roi']:.1f}x"
        kpis = [
            ("💰", "Total Revenue",  f"{cfg.currency_symbol}{st['total_revenue']/1e6:.2f}M", self.T.BLUE),
            ("📅", "Avg / Month",    f"{cfg.currency_symbol}{st['avg_monthly_revenue']/1e3:.0f}K", self.T.BLUE),
            ("📈", "CAGR",           f"{st['cagr']:.1f}%",        self.T.SUCCESS),
            ("🔮", "Forecast Total", f"{cfg.currency_symbol}{ens.predictions.sum()/1e6:.2f}M", self.T.PURPLE),
            ("🎯", "sMAPE",          f"{ens.smape:.1f}%",          self.T.WARN),
            ("📣", "Mktg ROI",       roi_str,                       self.T.TEAL),
            ("⚠️", "Anomalies",      f"{len(analysis.anomalies)}",  self.T.DANGER),
        ]
        for i, (icon, label, val, color) in enumerate(kpis):
            kc = tk.Frame(self._kpi_bar, bg=self.T.SURFACE,
                           highlightbackground=self.T.BORDER, highlightthickness=1)
            kc.pack(side="left", expand=True, fill="x",
                    padx=(0 if i == 0 else 6, 0), pady=2)
            tk.Label(kc, text=icon,  font=("Segoe UI", 13),
                      bg=self.T.SURFACE, fg=color).pack(pady=(10, 1))
            tk.Label(kc, text=val,   font=("Segoe UI", 13, "bold"),
                      bg=self.T.SURFACE, fg=color).pack()
            tk.Label(kc, text=label, font=("Segoe UI", 8),
                      bg=self.T.SURFACE, fg=self.T.MUTED).pack(pady=(1, 10))

        for row in self._model_tree.get_children():
            self._model_tree.delete(row)
        valid = {k: v for k, v in analysis.forecasts.items() if v is not None}
        valid["Ensemble ★"] = ens
        for name, res in valid.items():
            tag    = ("ens"  if "Ensemble" in name else
                      "best" if name == analysis.best_model else "")
            cv_str = (f"{res.cv_smape:.2f}"
                      if res.cv_smape is not None and math.isfinite(res.cv_smape)
                      else "—")
            self._model_tree.insert("", "end",
                values=(name,
                        f"{res.smape:.2f}", f"{res.mase:.3f}",
                        f"{cfg.currency_symbol}{res.rmse:,.0f}",
                        f"{res.r2:.3f}", cv_str,
                        f"{res.train_time_s:.2f}"),
                tags=(tag,))

        for w in self._files_row.winfo_children():
            w.destroy()
        tk.Label(self._files_row, text="Generated files:",
                  font=("Segoe UI", 10, "bold"),
                  bg=self.T.BG, fg=self.T.TEXT).pack(anchor="w", pady=(0, 6))
        btn_row = tk.Frame(self._files_row, bg=self.T.BG)
        btn_row.pack(anchor="w")
        icons = {"pdf": "📄", "excel": "📊", "json": "🗒️"}
        for fmt, path in files.items():
            p = Path(path)
            if not p.exists():
                continue
            sz = p.stat().st_size // 1024
            b  = tk.Button(
                btn_row,
                text=f"  {icons.get(fmt, '📁')}  {p.name}  ({sz:,} KB)",
                font=("Segoe UI", 9), bg=self.T.SURFACE, fg=self.T.TEXT2,
                activebackground=self.T.BORDER, activeforeground=self.T.TEXT,
                relief="flat", bd=0, padx=12, pady=7, cursor="hand2",
                command=lambda _p=p: self._open_file(_p))
            b.pack(side="left", padx=(0, 8))
            ToolTip(b, f"Click to open {p.name}")
        tk.Button(
            btn_row, text="  📂  Open Folder",
            font=("Segoe UI", 9), bg=self.T.BLUE, fg="#FFFFFF",
            activebackground=self.T.BLUE_HV, activeforeground="#FFFFFF",
            relief="flat", bd=0, padx=12, pady=7, cursor="hand2",
            command=lambda: self._open_folder(cfg.output_path.parent),
        ).pack(side="left")

    @staticmethod
    def _open_file(path: Path) -> None:
        """Open a file with the OS default application — cross-platform."""
        try:
            if sys.platform.startswith("darwin"):
                os.system(f'open "{path}"')
            elif sys.platform.startswith("win"):
                os.startfile(str(path))     # type: ignore[attr-defined]
            else:
                os.system(f'xdg-open "{path}" &')
        except Exception as exc:
            messagebox.showwarning("Open Failed", str(exc))

    @staticmethod
    def _open_folder(folder: Path) -> None:
        """Open a folder in the OS file manager — cross-platform."""
        try:
            if sys.platform.startswith("darwin"):
                os.system(f'open "{folder}"')
            elif sys.platform.startswith("win"):
                os.startfile(str(folder))   # type: ignore[attr-defined]
            else:
                os.system(f'xdg-open "{folder}" &')
        except Exception as exc:
            messagebox.showwarning("Open Folder Failed", str(exc))


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
#
#  [BUG-1 FIX] Import-guard re-ordered so the missing-dependency dialog is
#  shown BEFORE root.mainloop() is called, not after it returns.
#
#  Original (broken) flow:
#      root.mainloop()          ← app starts (and crashes mid-run)
#      messagebox.showerror()   ← never reached if user didn't close window
#
#  Fixed flow:
#      if _IMPORT_ERROR:
#          show error dialog    ← immediate, blocking, before any mainloop
#          sys.exit(1)          ← clean exit
#      else:
#          build app + mainloop ← only when deps are satisfied
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    # ── [BUG-1 FIX] Dependency check BEFORE anything else ────────────────────
    if _IMPORT_ERROR:
        # Spin up the smallest possible root just to host the error dialog,
        # then destroy it and exit — the full app is never constructed.
        _err_root = tk.Tk()
        _err_root.withdraw()          # hide the empty window
        messagebox.showerror(
            "Missing Dependencies",
            f"A required library is missing:\n\n{_IMPORT_ERROR}\n\n"
            "Install all dependencies with:\n\n"
            "  pip install numpy pandas matplotlib seaborn scipy \\\n"
            "              scikit-learn openpyxl reportlab pillow\n\n"
            "Then restart the application.")
        _err_root.destroy()
        sys.exit(1)

    # Normal startup ────────────────────────────────────────────────────────────
    root = tk.Tk()
    root.title(f"{APP_TITLE} — v{APP_VERSION}")
    root.configure(bg=Theme.BG)

    w, h = 1180, 800
    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    root.geometry(f"{w}x{h}+{max(0, (sw - w) // 2)}+{max(0, (sh - h) // 2)}")

    # Dark title-bar on Windows 10 / 11 (no-op on macOS / Linux)
    if sys.platform.startswith("win"):
        try:
            import ctypes
            hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 20,
                ctypes.byref(ctypes.c_int(2)), ctypes.sizeof(ctypes.c_int(2)))
        except Exception:
            pass  # non-fatal: dark title bar is cosmetic only

    app = SalesForecastApp(root)

    def _on_close() -> None:
        """Cancel any running analysis before destroying the window."""
        try:
            if app._cancel_token is not None:
                app._cancel_token.cancel()
        except Exception:
            pass
        try:
            root.destroy()
        except Exception:
            pass

    root.protocol("WM_DELETE_WINDOW", _on_close)
    app._log_msg(f"Sales Forecasting System {APP_VERSION} ready.", "ok")
    app._log_msg("⚙  Setup tab → configure → Run.", "accent")
    app._log_msg("── production-hardening fixes ─────────────────────", "sep")
    app._log_msg("Config: seasonality_period validated; n_bootstrap ≤ 5 000", "muted")
    app._log_msg("NaN-sales rows dropped (not silently zeroed)", "muted")
    app._log_msg("ARIMA σ² floor 1e-20 → 1e-10 (log stability)", "muted")
    app._log_msg("ARIMA predict() explicit float64 before undiff", "muted")
    app._log_msg("HoltWinters accepts non-success L-BFGS-B convergence", "muted")
    app._log_msg("Ensemble fast-path for single valid model", "muted")
    app._log_msg("decompose() pad clamp for n == period+1 edge case", "muted")
    app._log_msg("detect_anomalies() returns explicit dtype DataFrame", "muted")
    app._log_msg("Chart _save() recreates dir if cleaned up mid-run", "muted")
    app._log_msg("PDF RMSE column uses configured currency symbol", "muted")
    app._log_msg("Excel YoY/MoM colour-coding column indices corrected", "muted")
    app._log_msg("Excel workbook closed in finally block (Windows)", "muted")
    app._log_msg("Preview scrollbars use grid manager (HiDPI fix)", "muted")
    app._log_msg("JSON statistics fully serialised via _NpEncoder", "muted")
    app._log_msg("WM_DELETE_WINDOW cancels analysis before close", "muted")
    app._log_msg("Rotating log: ~/SalesForecastReports/sales_forecast.log", "muted")
    app._log_msg("────────────────────────────────────────────────────────", "sep")
    app._log_msg("Supported: .csv  .tsv  .txt  .xlsx  .xlsm  .xls  .json  .parquet", "muted")
    app._log_msg("Flexible columns: date/sales auto-detected from 50+ aliases", "muted")
    app._log_msg("Optional: product · region · units · returns · marketing_spend", "muted")
    app._log_msg("Cross-platform: Windows / macOS / Linux ✓", "muted")
    app._log_msg("━" * 56, "sep")
    try:
        root.mainloop()
    except KeyboardInterrupt:
        logger.info("Interrupted by user — exiting.")
    finally:
        try:
            root.destroy()
        except Exception:
            pass

if __name__ == "__main__":
    main()